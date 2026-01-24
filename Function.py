# ==================== ENHANCED AISC 360-16 STEEL DESIGN WEB APP ====================
# Version: 7.0 - UI/UX with Advanced Export Capabilities
# New Features: PDF/Excel Export, Modern UI, Enhanced Visualizations

import streamlit as st

# ==================== PAGE CONFIGURATION - MUST BE FIRST ====================
st.set_page_config(
    page_title="AISC 360-16 Steel Design",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== IMPORTS ====================
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.patches import Rectangle
import math
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime
import io
import base64
from io import BytesIO
import streamlit.components.v1 as components

# Add these imports at the top with other imports
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib import colors as rl_colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt

# ==================== PDF GENERATION WITH FIXED FORMATTING ====================
from reportlab.platypus import PageTemplate, Frame, BaseDocTemplate, KeepTogether
from reportlab.lib.units import inch, cm
from reportlab.pdfgen import canvas

from reportlab.platypus import PageTemplate, Frame, BaseDocTemplate, KeepTogether, Flowable
from reportlab.lib.units import inch, cm
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

class EquationBox(Flowable):
    """Custom flowable for equation boxes with proper padding and no overlap"""
    def __init__(self, text, width):
        Flowable.__init__(self)
        self.text = text
        self.width = width
        self.height = 0
        
    def wrap(self, availWidth, availHeight):
        # Calculate required height with padding
        self.width = availWidth
        # Estimate height based on text length (adjust as needed)
        lines = len(self.text) / 80 + 1
        self.height = max(30, lines * 15 + 20)  # Minimum 30pt, with padding
        return (self.width, self.height)
    
    def draw(self):
        # Draw blue background box
        self.canv.setFillColor(rl_colors.HexColor('#E7F3FF'))
        self.canv.setStrokeColor(rl_colors.HexColor('#2196F3'))
        self.canv.setLineWidth(1)
        self.canv.roundRect(0, 0, self.width, self.height, 5, fill=1, stroke=1)
        
        # Draw text with proper padding
        self.canv.setFillColor(rl_colors.HexColor('#1565C0'))
        self.canv.setFont('Courier', 9)
        
        # Word wrap text
        words = self.text.split()
        lines = []
        current_line = []
        current_width = 0
        max_width = self.width - 20  # 10pt padding on each side
        
        for word in words:
            word_width = self.canv.stringWidth(word + ' ', 'Courier', 9)
            if current_width + word_width <= max_width:
                current_line.append(word)
                current_width += word_width
            else:
                lines.append(' '.join(current_line))
                current_line = [word]
                current_width = word_width
        
        if current_line:
            lines.append(' '.join(current_line))
        
        # Draw lines
        y_position = self.height - 15
        for line in lines:
            self.canv.drawString(10, y_position, line)
            y_position -= 12


# Optional: AgGrid (only if installed)
try:
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
    AGGRID_AVAILABLE = True
except ImportError:
    AGGRID_AVAILABLE = False

# PDF Export Libraries (silent import, warnings shown later)
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Excel Export Libraries (silent import, warnings shown later)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference, LineChart
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter  # ADD THIS LINE
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class NumberedCanvas(canvas.Canvas):
    """Custom canvas for page numbers and headers"""
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.setFillColor(rl_colors.grey)
        
        # Header
        self.line(0.75*inch, letter[1] - 0.6*inch, letter[0] - 0.75*inch, letter[1] - 0.6*inch)
        self.setFont("Helvetica-Bold", 10)
        self.setFillColor(rl_colors.HexColor('#667eea'))
        self.drawString(0.75*inch, letter[1] - 0.5*inch, "AISC 360-16 Steel Design - Calculation Report")
        
        # Footer
        self.setFont("Helvetica", 9)
        self.setFillColor(rl_colors.grey)
        self.line(0.75*inch, 0.6*inch, letter[0] - 0.75*inch, 0.6*inch)
        self.drawRightString(letter[0] - 0.75*inch, 0.4*inch, 
                            f"Page {self._pageNumber} of {page_count}")
        self.drawString(0.75*inch, 0.4*inch, 
                       f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
# ==================== HELPER FUNCTIONS ====================

def safe_scalar(value):
    """Convert pandas Series or array to scalar value"""
    if hasattr(value, 'iloc'):
        return float(value.iloc[0])
    elif hasattr(value, '__len__') and not isinstance(value, str):
        return float(value[0])
    return float(value)

def format_number(value, decimals=2):
    """Format number with specified decimals"""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "—"
    return f"{value:,.{decimals}f}"

def format_equation_result(value, decimals=2, unit=""):
    """Format equation result with unit"""
    formatted = format_number(value, decimals)
    if unit:
        return f"{formatted} {unit}"
    return formatted

# ==================== SECTION CLASSIFICATION ====================

def classify_section_flange(bf, tf, tw, d, Fy, E=200000):
    """
    Classify section flange per AISC 360-16 Table B4.1b
    Returns: 'Compact', 'Noncompact', or 'Slender'
    """
    # Flange slenderness ratio
    lambda_f = bf / (2 * tf)
    
    # Limiting ratios (Table B4.1b Case 10 - I-shaped rolled sections)
    lambda_pf = 0.38 * math.sqrt(E / Fy)  # Compact limit
    lambda_rf = 1.0 * math.sqrt(E / Fy)   # Noncompact limit
    
    if lambda_f <= lambda_pf:
        return 'Compact', lambda_f, lambda_pf, lambda_rf
    elif lambda_f <= lambda_rf:
        return 'Noncompact', lambda_f, lambda_pf, lambda_rf
    else:
        return 'Slender', lambda_f, lambda_pf, lambda_rf

def classify_section_web(h, tw, Fy, E=200000):
    """
    Classify section web per AISC 360-16 Table B4.1b
    Returns: 'Compact', 'Noncompact', or 'Slender'
    """
    # Web slenderness ratio
    lambda_w = h / tw
    
    # Limiting ratios (Table B4.1b Case 15 - Webs in flexure)
    lambda_pw = 3.76 * math.sqrt(E / Fy)  # Compact limit
    lambda_rw = 5.70 * math.sqrt(E / Fy)  # Noncompact limit
    
    if lambda_w <= lambda_pw:
        return 'Compact', lambda_w, lambda_pw, lambda_rw
    elif lambda_w <= lambda_rw:
        return 'Noncompact', lambda_w, lambda_pw, lambda_rw
    else:
        return 'Slender', lambda_w, lambda_pw, lambda_rw

def get_overall_classification(flange_class, web_class):
    """Get overall section classification"""
    classes = ['Compact', 'Noncompact', 'Slender']
    return classes[max(classes.index(flange_class), classes.index(web_class))]

# ==================== FLEXURAL STRENGTH (AISC F2) ====================

def calculate_flexural_strength(section_props, Lb, Cb=1.0, E=200000):
    """
    Calculate flexural strength per AISC 360-16 Chapter F2
    Returns dict with all calculation steps
    """
    Fy = section_props['Fy']
    Zx = section_props['Zx']
    Sx = section_props['Sx']
    Iy = section_props['Iy']
    ry = section_props['ry']
    J = section_props.get('J', 0)
    Cw = section_props.get('Cw', 0)
    ho = section_props.get('ho', section_props.get('d', 0) - section_props.get('tf', 0))
    rts = section_props.get('rts', ry)
    
    # Convert units if needed (assuming cm to mm)
    Zx_mm3 = Zx * 1000  # cm³ to mm³
    Sx_mm3 = Sx * 1000
    Iy_mm4 = Iy * 10000  # cm⁴ to mm⁴
    ry_mm = ry * 10
    Lb_mm = Lb * 1000  # m to mm
    J_mm4 = J * 10000 if J else 1
    Cw_mm6 = Cw * 1000000 if Cw else 1
    ho_mm = ho * 10 if ho else 1
    rts_mm = rts * 10 if rts else ry_mm
    c = 1.0  # For doubly symmetric I-shapes
    
    # Plastic moment
    Mp = Fy * Zx_mm3 / 1e6  # kN·m
    
    # Limiting lengths
    Lp = 1.76 * ry_mm * math.sqrt(E / Fy)  # mm
    
    # Lr calculation
    try:
        term1 = (J_mm4 * c) / (Sx_mm3 * ho_mm)
        term2 = math.sqrt(term1**2 + 6.76 * (0.7 * Fy / E)**2)
        Lr = 1.95 * rts_mm * (E / (0.7 * Fy)) * math.sqrt(term1 + term2)
    except:
        Lr = Lp * 3  # Fallback estimate
    
    # Determine limit state
    if Lb_mm <= Lp:
        # Case (a): Yielding - Compact section, full plastic moment
        Mn = Mp
        limit_state = "Yielding (Lb ≤ Lp)"
        case = "F2-1"
        Fcr = Fy  # Not applicable but set for reference
    elif Lb_mm <= Lr:
        # Case (b): Inelastic LTB
        Mn = Cb * (Mp - (Mp - 0.7 * Fy * Sx_mm3 / 1e6) * (Lb_mm - Lp) / (Lr - Lp))
        Mn = min(Mn, Mp)
        limit_state = "Inelastic LTB (Lp < Lb ≤ Lr)"
        case = "F2-2"
        Fcr = Cb * (Fy - (Fy - 0.7 * Fy) * (Lb_mm - Lp) / (Lr - Lp))
    else:
        # Case (c): Elastic LTB
        try:
            term_a = (Cb * math.pi**2 * E) / (Lb_mm / rts_mm)**2
            term_b = math.sqrt(1 + 0.078 * (J_mm4 * c / (Sx_mm3 * ho_mm)) * (Lb_mm / rts_mm)**2)
            Fcr = term_a * term_b
        except:
            Fcr = 0.7 * Fy
        Mn = Fcr * Sx_mm3 / 1e6
        Mn = min(Mn, Mp)
        limit_state = "Elastic LTB (Lb > Lr)"
        case = "F2-3"
    
    # Design strength
    phi = 0.90
    phi_Mn = phi * Mn
    
    return {
        'Mp': Mp,
        'Mn': Mn,
        'phi': phi,
        'phi_Mn': phi_Mn,
        'Lp': Lp / 1000,  # Convert back to m
        'Lr': Lr / 1000,
        'Lb': Lb,
        'Cb': Cb,
        'Fcr': Fcr,
        'limit_state': limit_state,
        'case': case,
        'Fy': Fy,
        'Zx': Zx,
        'Sx': Sx
    }

# ==================== AXIAL STRENGTH (AISC E3) ====================

def calculate_compression_strength(section_props, KL_x, KL_y, E=200000):
    """
    Calculate compression strength per AISC 360-16 Chapter E3
    Returns dict with all calculation steps
    """
    Fy = section_props['Fy']
    Ag = section_props['Ag']
    rx = section_props['rx']
    ry = section_props['ry']
    
    # Convert units
    Ag_mm2 = Ag * 100  # cm² to mm²
    rx_mm = rx * 10  # cm to mm
    ry_mm = ry * 10
    KLx_mm = KL_x * 1000  # m to mm
    KLy_mm = KL_y * 1000
    
    # Slenderness ratios
    lambda_x = KLx_mm / rx_mm
    lambda_y = KLy_mm / ry_mm
    lambda_governing = max(lambda_x, lambda_y)
    governing_axis = 'x' if lambda_x >= lambda_y else 'y'
    
    # Elastic buckling stress
    Fe = (math.pi**2 * E) / lambda_governing**2
    
    # Limiting slenderness
    lambda_limit = 4.71 * math.sqrt(E / Fy)
    
    # Critical stress
    if lambda_governing <= lambda_limit:
        # Inelastic buckling (E3-2)
        Fcr = 0.658**(Fy / Fe) * Fy
        buckling_mode = "Inelastic"
        equation = "E3-2"
    else:
        # Elastic buckling (E3-3)
        Fcr = 0.877 * Fe
        buckling_mode = "Elastic"
        equation = "E3-3"
    
    # Nominal and design strength
    Pn = Fcr * Ag_mm2 / 1000  # kN
    phi = 0.90
    phi_Pn = phi * Pn
    
    return {
        'Pn': Pn,
        'phi': phi,
        'phi_Pn': phi_Pn,
        'Fcr': Fcr,
        'Fe': Fe,
        'lambda_x': lambda_x,
        'lambda_y': lambda_y,
        'lambda_governing': lambda_governing,
        'lambda_limit': lambda_limit,
        'governing_axis': governing_axis,
        'buckling_mode': buckling_mode,
        'equation': equation,
        'KL_x': KL_x,
        'KL_y': KL_y,
        'Ag': Ag,
        'Fy': Fy
    }

def calculate_tension_strength(section_props, E=200000):
    """
    Calculate tension strength per AISC 360-16 Chapter D
    Returns dict with all calculation steps
    """
    Fy = section_props['Fy']
    Fu = section_props.get('Fu', 1.2 * Fy)  # Estimate if not provided
    Ag = section_props['Ag']
    Ae = section_props.get('Ae', 0.85 * Ag)  # Effective area, default 85%
    
    # Convert units
    Ag_mm2 = Ag * 100
    Ae_mm2 = Ae * 100
    
    # Yielding on gross section (D2-1)
    Pn_yield = Fy * Ag_mm2 / 1000  # kN
    phi_yield = 0.90
    
    # Rupture on net section (D2-2)
    Pn_rupture = Fu * Ae_mm2 / 1000  # kN
    phi_rupture = 0.75
    
    # Design strength (governing)
    phi_Pn_yield = phi_yield * Pn_yield
    phi_Pn_rupture = phi_rupture * Pn_rupture
    
    if phi_Pn_yield <= phi_Pn_rupture:
        governing = "Yielding (D2-1)"
        phi_Pn = phi_Pn_yield
        Pn = Pn_yield
        phi = phi_yield
    else:
        governing = "Rupture (D2-2)"
        phi_Pn = phi_Pn_rupture
        Pn = Pn_rupture
        phi = phi_rupture
    
    return {
        'Pn_yield': Pn_yield,
        'Pn_rupture': Pn_rupture,
        'phi_yield': phi_yield,
        'phi_rupture': phi_rupture,
        'phi_Pn_yield': phi_Pn_yield,
        'phi_Pn_rupture': phi_Pn_rupture,
        'Pn': Pn,
        'phi': phi,
        'phi_Pn': phi_Pn,
        'governing': governing,
        'Ag': Ag,
        'Ae': Ae,
        'Fy': Fy,
        'Fu': Fu
    }

# ==================== COMBINED FORCES (AISC H1) ====================

def calculate_interaction(Pu, phi_Pn, Mux, phi_Mnx, Muy=0, phi_Mny=None, is_tension=False):
    """
    Calculate combined force interaction per AISC 360-16 Chapter H1
    Returns dict with all calculation steps
    """
    if phi_Mny is None:
        phi_Mny = phi_Mnx * 0.5  # Estimate for minor axis
    
    Pr = abs(Pu)
    Pc = phi_Pn
    Mrx = abs(Mux)
    Mcx = phi_Mnx
    Mry = abs(Muy)
    Mcy = phi_Mny
    
    # Axial ratio
    axial_ratio = Pr / Pc if Pc > 0 else 999
    
    if is_tension or Pu < 0:
        # Tension + Bending: Linear interaction
        moment_ratio_x = Mrx / Mcx if Mcx > 0 else 0
        moment_ratio_y = Mry / Mcy if Mcy > 0 else 0
        
        interaction_ratio = axial_ratio + moment_ratio_x + moment_ratio_y
        equation_used = "H1-1 (Modified for Tension)"
        equation_text = "Pu/φPn + Mux/φMnx + Muy/φMny ≤ 1.0"
    else:
        # Compression + Bending: AISC H1
        if axial_ratio >= 0.2:
            # Equation H1-1a
            moment_term = (Mrx / Mcx + Mry / Mcy) if Mcx > 0 and Mcy > 0 else 0
            interaction_ratio = axial_ratio + (8.0 / 9.0) * moment_term
            equation_used = "H1-1a"
            equation_text = "Pu/φPn + (8/9)(Mux/φMnx + Muy/φMny) ≤ 1.0"
        else:
            # Equation H1-1b
            moment_term = (Mrx / Mcx + Mry / Mcy) if Mcx > 0 and Mcy > 0 else 0
            interaction_ratio = axial_ratio / 2.0 + moment_term
            equation_used = "H1-1b"
            equation_text = "Pu/(2φPn) + (Mux/φMnx + Muy/φMny) ≤ 1.0"
    
    design_ok = interaction_ratio <= 1.0
    
    return {
        'Pu': Pu,
        'phi_Pn': phi_Pn,
        'Mux': Mux,
        'phi_Mnx': phi_Mnx,
        'Muy': Muy,
        'phi_Mny': phi_Mny,
        'axial_ratio': axial_ratio,
        'interaction_ratio': interaction_ratio,
        'equation_used': equation_used,
        'equation_text': equation_text,
        'design_ok': design_ok,
        'is_tension': is_tension or Pu < 0
    }

# ==================== HTML REPORT GENERATOR ====================

class SteelDesignReportGenerator:
    """Generate steel member design reports in HTML format"""
    
    def __init__(self, project_info=None):
        self.project_info = project_info or {}
        self.members = []
        self.report_date = datetime.now().strftime("%Y-%m-%d")
        
    def add_member(self, member_data):
        """Add a member to the report"""
        self.members.append(member_data)
    
    def _get_css_styles(self):
        """Return CSS styles for the report"""
        return """
        <style>
            * { box-sizing: border-box; }
            body { 
                font-family: 'Segoe UI', Arial, sans-serif; 
                font-size: 10pt; 
                line-height: 1.4;
                color: #333;
                max-width: 210mm;
                margin: 0 auto;
                padding: 15mm;
            }
            .report-header {
                text-align: center;
                border-bottom: 2px solid #2c3e50;
                padding-bottom: 10px;
                margin-bottom: 20px;
            }
            .report-title { 
                font-size: 16pt; 
                font-weight: bold; 
                color: #2c3e50;
                margin: 0;
            }
            .report-subtitle {
                font-size: 11pt;
                color: #666;
                margin: 5px 0;
            }
            .member-section {
                page-break-inside: avoid;
                margin-bottom: 25px;
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 15px;
                background: #fafafa;
            }
            .member-header {
                background: #2c3e50;
                color: white;
                padding: 8px 12px;
                margin: -15px -15px 15px -15px;
                border-radius: 5px 5px 0 0;
                font-size: 12pt;
                font-weight: bold;
            }
            .section-title {
                font-size: 11pt;
                font-weight: bold;
                color: #2c3e50;
                border-bottom: 1px solid #2c3e50;
                padding-bottom: 3px;
                margin: 15px 0 10px 0;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin: 10px 0;
                font-size: 9pt;
            }
            th, td {
                border: 1px solid #ccc;
                padding: 6px 8px;
                text-align: left;
                vertical-align: middle;
            }
            th {
                background: #ecf0f1;
                font-weight: bold;
                color: #2c3e50;
            }
            .calc-table th:first-child { width: 25%; }
            .calc-table th:nth-child(2) { width: 30%; }
            .calc-table th:nth-child(3) { width: 25%; }
            .calc-table th:nth-child(4) { width: 20%; }
            .equation { 
                font-family: 'Cambria Math', 'Times New Roman', serif;
                font-style: italic;
            }
            .result { 
                font-weight: bold; 
                color: #2c3e50;
            }
            .pass { 
                background: #d4edda; 
                color: #155724;
                font-weight: bold;
            }
            .fail { 
                background: #f8d7da; 
                color: #721c24;
                font-weight: bold;
            }
            .warning {
                background: #fff3cd;
                color: #856404;
                font-weight: bold;
            }
            .governing {
                background: #cce5ff;
                font-weight: bold;
            }
            .conclusion-box {
                border: 2px solid #2c3e50;
                border-radius: 5px;
                padding: 12px;
                margin-top: 15px;
            }
            .conclusion-pass {
                border-color: #28a745;
                background: #d4edda;
            }
            .conclusion-fail {
                border-color: #dc3545;
                background: #f8d7da;
            }
            .unit { 
                font-size: 8pt; 
                color: #666; 
            }
            .note {
                font-size: 8pt;
                color: #666;
                font-style: italic;
                margin: 5px 0;
            }
            @media print {
                body { padding: 10mm; }
                .member-section { page-break-inside: avoid; }
            }
        </style>
        """
    
    def _generate_member_geometry_table(self, member):
        """Generate Member & Geometry Summary table"""
        html = """
        <div class="section-title">1. Member & Geometry Summary</div>
        <table>
            <tr>
                <th>Parameter</th>
                <th>Value</th>
                <th>Parameter</th>
                <th>Value</th>
            </tr>
        """
        
        params = [
            ('Member No.', member.get('member_no', '—')),
            ('Member Type', member.get('member_type', '—')),
            ('Length (m)', format_number(member.get('length', 0), 2)),
            ('K Factor', format_number(member.get('K', 1.0), 2)),
            ('KL (m)', format_number(member.get('KL', 0), 2)),
            ('Lb (m)', format_number(member.get('Lb', 0), 2)),
        ]
        
        for i in range(0, len(params), 2):
            p1 = params[i]
            p2 = params[i+1] if i+1 < len(params) else ('', '')
            html += f"""
            <tr>
                <td><b>{p1[0]}</b></td>
                <td>{p1[1]}</td>
                <td><b>{p2[0]}</b></td>
                <td>{p2[1]}</td>
            </tr>
            """
        
        html += "</table>"
        return html
    
    def _generate_section_properties_table(self, member):
        """Generate Section Properties & Classification table"""
        section = member.get('section_props', {})
        classification = member.get('classification', 'Compact')
        
        html = """
        <div class="section-title">2. Section Properties & Classification</div>
        <table>
            <tr>
                <th>Property</th>
                <th>Value</th>
                <th>Property</th>
                <th>Value</th>
            </tr>
        """
        
        props = [
            ('Section', member.get('section_name', '—')),
            ('Ag (cm²)', format_number(section.get('Ag', 0), 2)),
            ('Ix (cm⁴)', format_number(section.get('Ix', 0), 1)),
            ('Iy (cm⁴)', format_number(section.get('Iy', 0), 1)),
            ('Sx (cm³)', format_number(section.get('Sx', 0), 1)),
            ('Zx (cm³)', format_number(section.get('Zx', 0), 1)),
            ('rx (cm)', format_number(section.get('rx', 0), 2)),
            ('ry (cm)', format_number(section.get('ry', 0), 2)),
            ('Fy (MPa)', format_number(section.get('Fy', 0), 0)),
            ('E (MPa)', '200,000'),
            ('Classification', f"<b>{classification}</b>"),
            ('', ''),
        ]
        
        for i in range(0, len(props), 2):
            p1 = props[i]
            p2 = props[i+1] if i+1 < len(props) else ('', '')
            html += f"""
            <tr>
                <td><b>{p1[0]}</b></td>
                <td>{p1[1]}</td>
                <td><b>{p2[0]}</b></td>
                <td>{p2[1]}</td>
            </tr>
            """
        
        html += "</table>"
        return html
    
    def _generate_flexural_strength_table(self, member):
        """Generate Flexural Strength Calculation table"""
        flex = member.get('flexural_results', {})
        section = member.get('section_props', {})
        classification = member.get('classification', 'Compact')
        
        # Determine which section modulus to use
        use_Zx = classification == 'Compact'
        S_used = section.get('Zx', 0) if use_Zx else section.get('Sx', 0)
        S_label = 'Zx' if use_Zx else 'Sx'
        
        html = f"""
        <div class="section-title">3. Flexural Strength Calculation (AISC F2)</div>
        <p class="note">Limit State: {flex.get('limit_state', '—')} | Equation: {flex.get('case', '—')}</p>
        <table class="calc-table">
            <tr>
                <th>Item</th>
                <th>Expression</th>
                <th>Substitution</th>
                <th>Result</th>
            </tr>
            <tr>
                <td>Limiting Length Lp</td>
                <td class="equation">Lp = 1.76·ry·√(E/Fy)</td>
                <td>1.76 × {format_number(section.get('ry', 0)*10, 1)} × √(200000/{format_number(section.get('Fy', 0), 0)})</td>
                <td class="result">{format_number(flex.get('Lp', 0), 3)} m</td>
            </tr>
            <tr>
                <td>Limiting Length Lr</td>
                <td class="equation">Lr (per AISC F2-6)</td>
                <td>—</td>
                <td class="result">{format_number(flex.get('Lr', 0), 3)} m</td>
            </tr>
            <tr>
                <td>Plastic Moment Mp</td>
                <td class="equation">Mp = Fy·Zx</td>
                <td>{format_number(section.get('Fy', 0), 0)} × {format_number(section.get('Zx', 0)*1000, 0)} / 10⁶</td>
                <td class="result">{format_number(flex.get('Mp', 0), 2)} kN·m</td>
            </tr>
            <tr>
                <td>Nominal Moment Mn</td>
                <td class="equation">Mn = Fy·{S_label} (or LTB)</td>
                <td>Per {flex.get('case', 'F2-1')}</td>
                <td class="result">{format_number(flex.get('Mn', 0), 2)} kN·m</td>
            </tr>
            <tr>
                <td>Resistance Factor</td>
                <td class="equation">φb = 0.90</td>
                <td>—</td>
                <td class="result">0.90</td>
            </tr>
            <tr>
                <td><b>Design Strength φMn</b></td>
                <td class="equation">φMn = φb × Mn</td>
                <td>0.90 × {format_number(flex.get('Mn', 0), 2)}</td>
                <td class="result"><b>{format_number(flex.get('phi_Mn', 0), 2)} kN·m</b></td>
            </tr>
        </table>
        <p class="note">Note: {'Using Zx for compact section' if use_Zx else 'Using Sx for noncompact/slender section per F2'}</p>
        """
        return html
    
    def _generate_compression_strength_table(self, member):
        """Generate Compression Strength Calculation table"""
        comp = member.get('compression_results', {})
        section = member.get('section_props', {})
        
        html = f"""
        <div class="section-title">4. Compression Strength Calculation (AISC E3)</div>
        <p class="note">Buckling Mode: {comp.get('buckling_mode', '—')} | Equation: {comp.get('equation', '—')} | Governing Axis: {comp.get('governing_axis', '—')}</p>
        <table class="calc-table">
            <tr>
                <th>Item</th>
                <th>Expression</th>
                <th>Substitution</th>
                <th>Result</th>
            </tr>
            <tr>
                <td>Slenderness (x-axis)</td>
                <td class="equation">λx = KLx / rx</td>
                <td>{format_number(member.get('KL', 0)*1000, 0)} / {format_number(section.get('rx', 0)*10, 1)}</td>
                <td class="result">{format_number(comp.get('lambda_x', 0), 1)}</td>
            </tr>
            <tr>
                <td>Slenderness (y-axis)</td>
                <td class="equation">λy = KLy / ry</td>
                <td>{format_number(member.get('KL', 0)*1000, 0)} / {format_number(section.get('ry', 0)*10, 1)}</td>
                <td class="result">{format_number(comp.get('lambda_y', 0), 1)}</td>
            </tr>
            <tr>
                <td>Limiting Slenderness</td>
                <td class="equation">4.71·√(E/Fy)</td>
                <td>4.71 × √(200000/{format_number(section.get('Fy', 0), 0)})</td>
                <td class="result">{format_number(comp.get('lambda_limit', 0), 1)}</td>
            </tr>
            <tr>
                <td>Euler Stress Fe</td>
                <td class="equation">Fe = π²E / λ²</td>
                <td>π² × 200000 / {format_number(comp.get('lambda_governing', 0), 1)}²</td>
                <td class="result">{format_number(comp.get('Fe', 0), 1)} MPa</td>
            </tr>
            <tr>
                <td>Critical Stress Fcr</td>
                <td class="equation">{comp.get('equation', 'E3-2/E3-3')}</td>
                <td>Per AISC E3</td>
                <td class="result">{format_number(comp.get('Fcr', 0), 1)} MPa</td>
            </tr>
            <tr>
                <td>Nominal Strength Pn</td>
                <td class="equation">Pn = Fcr × Ag</td>
                <td>{format_number(comp.get('Fcr', 0), 1)} × {format_number(section.get('Ag', 0)*100, 0)} / 1000</td>
                <td class="result">{format_number(comp.get('Pn', 0), 1)} kN</td>
            </tr>
            <tr>
                <td>Resistance Factor</td>
                <td class="equation">φc = 0.90</td>
                <td>—</td>
                <td class="result">0.90</td>
            </tr>
            <tr>
                <td><b>Design Strength φPn</b></td>
                <td class="equation">φPn = φc × Pn</td>
                <td>0.90 × {format_number(comp.get('Pn', 0), 1)}</td>
                <td class="result"><b>{format_number(comp.get('phi_Pn', 0), 1)} kN</b></td>
            </tr>
        </table>
        """
        return html
    
    def _generate_tension_strength_table(self, member):
        """Generate Tension Strength Calculation table"""
        tens = member.get('tension_results', {})
        section = member.get('section_props', {})
        
        html = f"""
        <div class="section-title">4. Tension Strength Calculation (AISC D2)</div>
        <p class="note">Governing Limit State: {tens.get('governing', '—')}</p>
        <table class="calc-table">
            <tr>
                <th>Item</th>
                <th>Expression</th>
                <th>Substitution</th>
                <th>Result</th>
            </tr>
            <tr>
                <td>Yielding on Gross Section</td>
                <td class="equation">Pn = Fy × Ag</td>
                <td>{format_number(section.get('Fy', 0), 0)} × {format_number(section.get('Ag', 0)*100, 0)} / 1000</td>
                <td class="result">{format_number(tens.get('Pn_yield', 0), 1)} kN</td>
            </tr>
            <tr>
                <td>Design Strength (Yield)</td>
                <td class="equation">φPn = 0.90 × Pn</td>
                <td>0.90 × {format_number(tens.get('Pn_yield', 0), 1)}</td>
                <td class="result">{format_number(tens.get('phi_Pn_yield', 0), 1)} kN</td>
            </tr>
            <tr>
                <td>Rupture on Net Section</td>
                <td class="equation">Pn = Fu × Ae</td>
                <td>{format_number(tens.get('Fu', 0), 0)} × {format_number(tens.get('Ae', 0)*100, 0)} / 1000</td>
                <td class="result">{format_number(tens.get('Pn_rupture', 0), 1)} kN</td>
            </tr>
            <tr>
                <td>Design Strength (Rupture)</td>
                <td class="equation">φPn = 0.75 × Pn</td>
                <td>0.75 × {format_number(tens.get('Pn_rupture', 0), 1)}</td>
                <td class="result">{format_number(tens.get('phi_Pn_rupture', 0), 1)} kN</td>
            </tr>
            <tr>
                <td><b>Design Strength φPn</b></td>
                <td class="equation">Min(φPn_yield, φPn_rupture)</td>
                <td>Governing: {tens.get('governing', '—')}</td>
                <td class="result"><b>{format_number(tens.get('phi_Pn', 0), 1)} kN</b></td>
            </tr>
        </table>
        """
        return html
    
    def _generate_load_combination_table(self, member):
        """Generate Load Combination Summary table"""
        loads = member.get('loads', [])
        
        html = """
        <div class="section-title">5. Load Combination Summary</div>
        <table>
            <tr>
                <th>Load Comb.</th>
                <th>Pu (kN)</th>
                <th>Mux (kN·m)</th>
                <th>Muy (kN·m)</th>
                <th>Load Type</th>
            </tr>
        """
        
        for load in loads:
            Pu = load.get('Pu', 0)
            load_type = "Compression" if Pu > 0 else ("Tension" if Pu < 0 else "—")
            html += f"""
            <tr>
                <td>{load.get('LC', '—')}</td>
                <td>{format_number(Pu, 1)}</td>
                <td>{format_number(load.get('Mux', 0), 2)}</td>
                <td>{format_number(load.get('Muy', 0), 2)}</td>
                <td>{load_type}</td>
            </tr>
            """
        
        html += "</table>"
        return html
    
    def _generate_interaction_table(self, member):
        """Generate Combined Force Check table"""
        interactions = member.get('interaction_results', [])
        
        if not interactions:
            return ""
        
        html = """
        <div class="section-title">6. Combined Force Check (AISC H1)</div>
        <table>
            <tr>
                <th>LC</th>
                <th>Equation</th>
                <th>Substitution</th>
                <th>Ratio</th>
                <th>Status</th>
            </tr>
        """
        
        # Find governing (max ratio)
        max_ratio = max(i.get('interaction_ratio', 0) for i in interactions)
        
        for inter in interactions:
            ratio = inter.get('interaction_ratio', 0)
            is_governing = (ratio == max_ratio)
            status_class = 'pass' if ratio <= 1.0 else 'fail'
            row_class = 'governing' if is_governing else ''
            status = "OK" if ratio <= 1.0 else "NG"
            
            # Build substitution string
            Pu = inter.get('Pu', 0)
            phi_Pn = inter.get('phi_Pn', 1)
            Mux = inter.get('Mux', 0)
            phi_Mnx = inter.get('phi_Mnx', 1)
            
            if inter.get('is_tension', False):
                subst = f"|{format_number(Pu, 1)}|/{format_number(phi_Pn, 1)} + {format_number(Mux, 2)}/{format_number(phi_Mnx, 2)}"
            else:
                axial_ratio = abs(Pu) / phi_Pn if phi_Pn > 0 else 0
                if axial_ratio >= 0.2:
                    subst = f"{format_number(abs(Pu), 1)}/{format_number(phi_Pn, 1)} + (8/9)({format_number(Mux, 2)}/{format_number(phi_Mnx, 2)})"
                else:
                    subst = f"{format_number(abs(Pu), 1)}/(2×{format_number(phi_Pn, 1)}) + {format_number(Mux, 2)}/{format_number(phi_Mnx, 2)}"
            
            gov_marker = " ★" if is_governing else ""
            
            html += f"""
            <tr class="{row_class}">
                <td>{inter.get('LC', '—')}{gov_marker}</td>
                <td class="equation">{inter.get('equation_used', '—')}</td>
                <td>{subst}</td>
                <td class="result">{format_number(ratio, 3)}</td>
                <td class="{status_class}">{status}</td>
            </tr>
            """
        
        html += "</table>"
        html += '<p class="note">★ = Governing Load Combination</p>'
        return html
    
    def _generate_conclusion(self, member):
        """Generate Design Conclusion section"""
        interactions = member.get('interaction_results', [])
        
        if not interactions:
            max_ratio = 0
            governing_lc = "—"
            governing_state = "—"
        else:
            max_idx = max(range(len(interactions)), key=lambda i: interactions[i].get('interaction_ratio', 0))
            max_inter = interactions[max_idx]
            max_ratio = max_inter.get('interaction_ratio', 0)
            governing_lc = max_inter.get('LC', '—')
            governing_state = max_inter.get('equation_used', '—')
        
        design_ok = max_ratio <= 1.0
        status_class = 'conclusion-pass' if design_ok else 'conclusion-fail'
        status_text = 'PASS ✓' if design_ok else 'FAIL ✗'
        
        html = f"""
        <div class="section-title">7. Design Conclusion</div>
        <div class="conclusion-box {status_class}">
            <table>
                <tr>
                    <th>Item</th>
                    <th>Result</th>
                </tr>
                <tr>
                    <td>Governing Load Combination</td>
                    <td><b>LC-{governing_lc}</b></td>
                </tr>
                <tr>
                    <td>Governing Limit State</td>
                    <td><b>{governing_state}</b></td>
                </tr>
                <tr>
                    <td>Maximum Strength Ratio</td>
                    <td><b>{format_number(max_ratio, 3)}</b></td>
                </tr>
                <tr>
                    <td>Design Status</td>
                    <td class="{'pass' if design_ok else 'fail'}" style="font-size: 12pt;"><b>{status_text}</b></td>
                </tr>
            </table>
        </div>
        """
        return html
    
    def generate_member_report(self, member):
        """Generate complete report section for one member"""
        member_type = member.get('member_type', 'Beam-Column')
        
        html = f"""
        <div class="member-section">
            <div class="member-header">
                Member: {member.get('member_no', '—')} | Section: {member.get('section_name', '—')} | Type: {member_type}
            </div>
        """
        
        # 1. Member & Geometry Summary
        html += self._generate_member_geometry_table(member)
        
        # 2. Section Properties & Classification
        html += self._generate_section_properties_table(member)
        
        # 3. Flexural Strength (if applicable)
        if member_type in ['Beam', 'Beam-Column']:
            html += self._generate_flexural_strength_table(member)
        
        # 4. Axial Strength
        if member_type in ['Column', 'Beam-Column']:
            html += self._generate_compression_strength_table(member)
        elif member_type == 'Tension Member':
            html += self._generate_tension_strength_table(member)
        
        # 5. Load Combination Summary
        html += self._generate_load_combination_table(member)
        
        # 6. Combined Force Check (if beam-column)
        if member_type in ['Beam-Column', 'Column']:
            html += self._generate_interaction_table(member)
        
        # 7. Design Conclusion
        html += self._generate_conclusion(member)
        
        html += "</div>"
        return html
    
    def generate_full_report(self):
        """Generate complete HTML report for all members"""
        html = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Steel Design Report</title>
            {self._get_css_styles()}
        </head>
        <body>
            <div class="report-header">
                <h1 class="report-title">STEEL MEMBER DESIGN REPORT</h1>
                <p class="report-subtitle">AISC 360-16 Specification for Structural Steel Buildings</p>
                <p class="report-subtitle">
                    Project: {self.project_info.get('name', '—')} | 
                    Date: {self.report_date} |
                    Prepared by: {self.project_info.get('engineer', '—')}
                </p>
            </div>
        """
        
        # Generate report for each member
        for member in self.members:
            html += self.generate_member_report(member)
        
        html += """
        </body>
        </html>
        """
        
        return html


# ==================== STREAMLIT INTEGRATION FUNCTIONS ====================

def create_member_data(member_no, section_name, section_props, member_type, 
                       length, K, KL, Lb, loads, classification='Compact'):
    """
    Create a complete member data dictionary for report generation
    
    Parameters:
    - member_no: Member identifier
    - section_name: Section designation (e.g., "W14x22")
    - section_props: Dict with section properties (Ag, Ix, Iy, Sx, Zx, rx, ry, Fy, etc.)
    - member_type: 'Beam', 'Column', 'Beam-Column', or 'Tension Member'
    - length: Member length (m)
    - K: Effective length factor
    - KL: Effective length (m)
    - Lb: Unbraced length for flexure (m)
    - loads: List of dicts with 'LC', 'Pu', 'Mux', 'Muy'
    - classification: 'Compact', 'Noncompact', or 'Slender'
    
    Returns complete member dict with all calculations
    """
    
    member = {
        'member_no': member_no,
        'section_name': section_name,
        'section_props': section_props,
        'member_type': member_type,
        'length': length,
        'K': K,
        'KL': KL,
        'Lb': Lb,
        'loads': loads,
        'classification': classification
    }
    
    # Calculate flexural strength
    if member_type in ['Beam', 'Beam-Column']:
        member['flexural_results'] = calculate_flexural_strength(
            section_props, Lb, Cb=1.0
        )
    
    # Calculate axial strength
    if member_type in ['Column', 'Beam-Column']:
        member['compression_results'] = calculate_compression_strength(
            section_props, KL, KL
        )
    elif member_type == 'Tension Member':
        member['tension_results'] = calculate_tension_strength(section_props)
    
    # Calculate interaction for each load combination
    interaction_results = []
    
    for load in loads:
        LC = load.get('LC', 1)
        Pu = load.get('Pu', 0)  # kN
        Mux = load.get('Mux', 0)  # kN·m
        Muy = load.get('Muy', 0)  # kN·m
        
        # Get capacities
        if member_type == 'Tension Member':
            phi_Pn = member.get('tension_results', {}).get('phi_Pn', 1)
            phi_Mnx = 0
            phi_Mny = 0
            is_tension = True
        else:
            phi_Pn = member.get('compression_results', {}).get('phi_Pn', 1)
            phi_Mnx = member.get('flexural_results', {}).get('phi_Mn', 1)
            phi_Mny = phi_Mnx * 0.5  # Estimate for minor axis
            is_tension = Pu < 0
        
        if member_type in ['Beam-Column', 'Column', 'Tension Member']:
            inter = calculate_interaction(
                Pu, phi_Pn, Mux, phi_Mnx, Muy, phi_Mny, is_tension
            )
            inter['LC'] = LC
            interaction_results.append(inter)
        elif member_type == 'Beam':
            # Beam only - check moment ratio
            ratio = abs(Mux) / phi_Mnx if phi_Mnx > 0 else 999
            interaction_results.append({
                'LC': LC,
                'Pu': 0,
                'phi_Pn': 0,
                'Mux': Mux,
                'phi_Mnx': phi_Mnx,
                'Muy': Muy,
                'phi_Mny': phi_Mny,
                'interaction_ratio': ratio,
                'equation_used': 'Mu/φMn',
                'equation_text': 'Mu/φMn ≤ 1.0',
                'design_ok': ratio <= 1.0,
                'is_tension': False
            })
    
    member['interaction_results'] = interaction_results
    
    return member


# ==================== EXAMPLE USAGE / DEMO ====================

def demo_report():
    """Generate a demo report"""
    
    # Sample section properties (W14x22 equivalent in metric)
    section_props = {
        'Ag': 41.8,      # cm²
        'Ix': 8247,      # cm⁴
        'Iy': 533,       # cm⁴
        'Sx': 467,       # cm³
        'Zx': 524,       # cm³
        'rx': 14.0,      # cm
        'ry': 3.57,      # cm
        'Fy': 345,       # MPa (Gr. 50)
        'Fu': 450,       # MPa
        'J': 16.4,       # cm⁴
        'd': 35.0,       # cm
        'tf': 1.07,      # cm
        'ho': 33.9,      # cm
        'rts': 3.91      # cm
    }
    
    # Sample loads for a beam-column
    loads = [
        {'LC': 1, 'Pu': 450, 'Mux': 85, 'Muy': 0},
        {'LC': 2, 'Pu': 520, 'Mux': 95, 'Muy': 0},
        {'LC': 3, 'Pu': -180, 'Mux': 65, 'Muy': 0},  # Tension case
        {'LC': 4, 'Pu': 580, 'Mux': 110, 'Muy': 0},
        {'LC': 5, 'Pu': 490, 'Mux': 88, 'Muy': 0},
    ]
    
    # Create member data
    member1 = create_member_data(
        member_no='B-101',
        section_name='W14x22',
        section_props=section_props,
        member_type='Beam-Column',
        length=4.0,
        K=1.0,
        KL=4.0,
        Lb=2.0,
        loads=loads,
        classification='Compact'
    )
    
    # Create tension member
    tension_loads = [
        {'LC': 1, 'Pu': -320, 'Mux': 0, 'Muy': 0},
        {'LC': 2, 'Pu': -410, 'Mux': 0, 'Muy': 0},
        {'LC': 3, 'Pu': -285, 'Mux': 0, 'Muy': 0},
    ]
    
    member2 = create_member_data(
        member_no='T-201',
        section_name='W8x18',
        section_props={
            'Ag': 34.2, 'Ix': 3040, 'Iy': 310, 'Sx': 278, 'Zx': 316,
            'rx': 9.42, 'ry': 3.00, 'Fy': 345, 'Fu': 450, 'Ae': 29.1
        },
        member_type='Tension Member',
        length=6.0,
        K=1.0,
        KL=6.0,
        Lb=6.0,
        loads=tension_loads,
        classification='Compact'
    )
    
    # Generate report
    generator = SteelDesignReportGenerator(
        project_info={
            'name': 'Sample Steel Building',
            'engineer': 'Design Engineer'
        }
    )
    
    generator.add_member(member1)
    generator.add_member(member2)
    
    return generator.generate_full_report()


if __name__ == "__main__":
    # Generate demo report and save
    html_report = demo_report()
    
    with open('steel_design_report_demo.html', 'w', encoding='utf-8') as f:
        f.write(html_report)
    
    print("Demo report generated: steel_design_report_demo.html")

# ==================== AISC CLASSIFICATION FUNCTIONS ====================
def classify_section_flexure(df, df_mat, section, material):
    """
    Classify section elements for flexural design per AISC 360-16 Table B4.1b
    Returns: dict with flange and web classifications
    """
    try:
        # Get section properties
        bf = safe_scalar(df.loc[section, 'bf [mm]'])
        tf = safe_scalar(df.loc[section, 'tf [mm]'])
        
        # For web
        if 'ho [mm]' in df.columns:
            h = safe_scalar(df.loc[section, 'ho [mm]'])
        else:
            d = safe_scalar(df.loc[section, 'd [mm]'])
            h = d - 2 * tf
        
        tw = safe_scalar(df.loc[section, 'tw [mm]'])
        
        # Material properties
        Fy = safe_scalar(df_mat.loc[material, "Yield Point (ksc)"])
        E = safe_scalar(df_mat.loc[material, "E"])
        
        # Flange slenderness (AISC Table B4.1b Case 10 - Flanges of I-sections)
        lambda_f = (bf / 2.0) / tf
        lambda_pf = 0.38 * safe_sqrt(E / Fy)
        lambda_rf = 1.0 * safe_sqrt(E / Fy)
        
        if lambda_f <= lambda_pf:
            flange_class = "Compact"
        elif lambda_f <= lambda_rf:
            flange_class = "Non-compact"
        else:
            flange_class = "Slender"
        
        # Web slenderness (AISC Table B4.1b Case 15 - Webs in flexural compression)
        lambda_w = h / tw
        lambda_pw = 3.76 * safe_sqrt(E / Fy)
        lambda_rw = 5.70 * safe_sqrt(E / Fy)
        
        if lambda_w <= lambda_pw:
            web_class = "Compact"
        elif lambda_w <= lambda_rw:
            web_class = "Non-compact"
        else:
            web_class = "Slender"
        
        return {
            'flange_class': flange_class,
            'flange_lambda': lambda_f,
            'flange_lambda_p': lambda_pf,
            'flange_lambda_r': lambda_rf,
            'web_class': web_class,
            'web_lambda': lambda_w,
            'web_lambda_p': lambda_pw,
            'web_lambda_r': lambda_rw
        }
        
    except Exception as e:
        st.error(f"Error in flexural classification: {e}")
        return None

def classify_section_compression(df, df_mat, section, material):
    """
    Classify section elements for compression design per AISC 360-16 Table B4.1a
    Returns: dict with overall classification
    """
    try:
        # Get section properties
        bf = safe_scalar(df.loc[section, 'bf [mm]'])
        tf = safe_scalar(df.loc[section, 'tf [mm]'])
        
        d = safe_scalar(df.loc[section, 'd [mm]'])
        tw = safe_scalar(df.loc[section, 'tw [mm]'])
        h = d - 2 * tf
        
        # Material properties
        Fy = safe_scalar(df_mat.loc[material, "Yield Point (ksc)"])
        E = safe_scalar(df_mat.loc[material, "E"])
        
        # Flange slenderness (AISC Table B4.1a Case 1 - Flanges of I-sections)
        lambda_f = (bf / 2.0) / tf
        lambda_r_flange = 0.56 * safe_sqrt(E / Fy)
        
        # Web slenderness (AISC Table B4.1a Case 5 - Webs of doubly symmetric I-sections)
        lambda_w = h / tw
        lambda_r_web = 1.49 * safe_sqrt(E / Fy)
        
        flange_slender = lambda_f > lambda_r_flange
        web_slender = lambda_w > lambda_r_web
        
        if flange_slender or web_slender:
            overall_class = "Slender"
            limiting_element = []
            if flange_slender:
                limiting_element.append("Flange")
            if web_slender:
                limiting_element.append("Web")
            limiting = " & ".join(limiting_element)
        else:
            overall_class = "Non-slender"
            limiting = "N/A"
        
        return {
            'overall_class': overall_class,
            'limiting_element': limiting,
            'flange_lambda': lambda_f,
            'flange_lambda_r': lambda_r_flange,
            'flange_slender': flange_slender,
            'web_lambda': lambda_w,
            'web_lambda_r': lambda_r_web,
            'web_slender': web_slender
        }
        
    except Exception as e:
        st.error(f"Error in compression classification: {e}")
        return None

def generate_excel_report(df, df_mat, section, material, analysis_results, design_params):
    """Generate comprehensive Excel calculation report with formatting"""
    if not EXCEL_AVAILABLE:
        return None
    
    buffer = BytesIO()
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="667EEA")
    subtitle_font = Font(bold=True, size=14, color="2c3e50")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Compact/Non-compact fill colors
    compact_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    noncompact_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    slender_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    # Get classifications
    flex_class = classify_section_flexure(df, df_mat, section, material)
    comp_class = classify_section_compression(df, df_mat, section, material)
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Design Summary"
    
    # Title
    ws_summary['A1'] = "AISC 360-16 STEEL DESIGN CALCULATION REPORT"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].alignment = center_align
    ws_summary.merge_cells('A1:D1')
    
    # Header Info
    row = 3
    ws_summary[f'A{row}'] = "Report Generated:"
    ws_summary[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row += 1
    ws_summary[f'A{row}'] = "Section:"
    ws_summary[f'B{row}'] = section
    row += 1
    ws_summary[f'A{row}'] = "Material Grade:"
    ws_summary[f'B{row}'] = material
    
    # Format header
    for r in range(3, 6):
        ws_summary[f'A{r}'].font = Font(bold=True)
        ws_summary[f'A{r}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    # Material Properties
    row = 8
    ws_summary[f'A{row}'] = "MATERIAL PROPERTIES"
    ws_summary[f'A{row}'].font = subtitle_font
    ws_summary.merge_cells(f'A{row}:C{row}')
    
    row += 1
    headers = ['Property', 'Value', 'Unit']
    for col, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    Fy = safe_scalar(df_mat.loc[material, "Yield Point (ksc)"])
    Fu = safe_scalar(df_mat.loc[material, "Tensile Strength (ksc)"])
    E = safe_scalar(df_mat.loc[material, "E"])
    
    mat_props = [
        ['Yield Strength (Fy)', f'{Fy:.1f}', 'kgf/cm²'],
        ['Tensile Strength (Fu)', f'{Fu:.1f}', 'kgf/cm²'],
        ['Modulus of Elasticity (E)', f'{E:.0f}', 'kgf/cm²']
    ]
    
    for prop_row in mat_props:
        row += 1
        for col, value in enumerate(prop_row, start=1):
            cell = ws_summary.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = center_align if col > 1 else left_align
    
    # ========== COMPREHENSIVE SECTION PROPERTIES ==========
    row += 3
    ws_summary[f'A{row}'] = "COMPLETE SECTION PROPERTIES"
    ws_summary[f'A{row}'].font = subtitle_font
    ws_summary.merge_cells(f'A{row}:C{row}')
    
    row += 1
    for col, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Comprehensive property list
    property_list = [
        ('d [mm]', 'Overall Depth'),
        ('bf [mm]', 'Flange Width'),
        ('tw [mm]', 'Web Thickness'),
        ('tf [mm]', 'Flange Thickness'),
        ('r [mm]', 'Fillet Radius'),
        ('A [cm2]', 'Cross-sectional Area'),
        ('Ix [cm4]', 'Moment of Inertia X-axis'),
        ('Iy [cm4]', 'Moment of Inertia Y-axis'),
        ('rx [cm]', 'Radius of Gyration X-axis'),
        ('ry [cm]', 'Radius of Gyration Y-axis'),
        ('Sx [cm3]', 'Elastic Section Modulus X'),
        ('Sy [cm3]', 'Elastic Section Modulus Y'),
        ('Zx [cm3]', 'Plastic Section Modulus X'),
        ('Zy [cm3]', 'Plastic Section Modulus Y'),
        ('h/tw', 'Web Slenderness Ratio'),
        ('0.5bf/tf', 'Flange Slenderness Ratio'),
        ('ho [mm]', 'Distance Between Flange Centroids'),
        ('j [cm4]', 'Torsional Constant'),
        ('cw [10^6 cm6]', 'Warping Constant'),
        ('rts [cm6]', 'Effective Radius'),
        ('Lp [cm]', 'Limiting Length for Plastic'),
        ('Lr [cm]', 'Limiting Length for Inelastic LTB')
    ]
    
    weight_col = 'Unit Weight [kg/m]' if 'Unit Weight [kg/m]' in df.columns else 'w [kg/m]'
    if weight_col in df.columns:
        weight = safe_scalar(df.loc[section, weight_col])
        row += 1
        ws_summary[f'A{row}'] = "Unit Weight"
        ws_summary[f'B{row}'] = f'{weight:.2f}'
        ws_summary[f'C{row}'] = 'kg/m'
        for col in range(1, 4):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center_align if col > 1 else left_align
    
    for prop_key, prop_description in property_list:
        if prop_key in df.columns:
            row += 1
            value = safe_scalar(df.loc[section, prop_key])
            unit = prop_key.split('[')[1].replace(']', '') if '[' in prop_key else ''
            
            ws_summary[f'A{row}'] = prop_description
            ws_summary[f'B{row}'] = f'{value:.3f}' if value < 100 else f'{value:.2f}'
            ws_summary[f'C{row}'] = unit
            
            for col in range(1, 4):
                cell = ws_summary.cell(row=row, column=col)
                cell.border = border
                cell.alignment = center_align if col > 1 else left_align
    
    # ========== FLEXURAL CLASSIFICATION ==========
    if flex_class:
        row += 3
        ws_summary[f'A{row}'] = "FLEXURAL MEMBER CLASSIFICATION (AISC Table B4.1b)"
        ws_summary[f'A{row}'].font = subtitle_font
        ws_summary.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws_summary[f'A{row}'] = "Element"
        ws_summary[f'B{row}'] = "Classification"
        ws_summary[f'C{row}'] = "Details"
        for col in range(1, 4):
            cell = ws_summary.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Flange classification
        row += 1
        ws_summary[f'A{row}'] = "Flange"
        ws_summary[f'B{row}'] = flex_class['flange_class']
        ws_summary[f'C{row}'] = f"λ={flex_class['flange_lambda']:.2f}, λp={flex_class['flange_lambda_p']:.2f}, λr={flex_class['flange_lambda_r']:.2f}"
        
        for col in range(1, 4):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center_align if col > 1 else left_align
            
            if col == 2:  # Classification column
                if flex_class['flange_class'] == "Compact":
                    cell.fill = compact_fill
                    cell.font = Font(bold=True, color="2E7D32")
                elif flex_class['flange_class'] == "Non-compact":
                    cell.fill = noncompact_fill
                    cell.font = Font(bold=True, color="F57C00")
                else:
                    cell.fill = slender_fill
                    cell.font = Font(bold=True, color="C62828")
        
        # Web classification
        row += 1
        ws_summary[f'A{row}'] = "Web"
        ws_summary[f'B{row}'] = flex_class['web_class']
        ws_summary[f'C{row}'] = f"λ={flex_class['web_lambda']:.2f}, λp={flex_class['web_lambda_p']:.2f}, λr={flex_class['web_lambda_r']:.2f}"
        
        for col in range(1, 4):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center_align if col > 1 else left_align
            
            if col == 2:  # Classification column
                if flex_class['web_class'] == "Compact":
                    cell.fill = compact_fill
                    cell.font = Font(bold=True, color="2E7D32")
                elif flex_class['web_class'] == "Non-compact":
                    cell.fill = noncompact_fill
                    cell.font = Font(bold=True, color="F57C00")
                else:
                    cell.fill = slender_fill
                    cell.font = Font(bold=True, color="C62828")
    
    # ========== COMPRESSION CLASSIFICATION ==========
    if comp_class:
        row += 3
        ws_summary[f'A{row}'] = "COMPRESSION MEMBER CLASSIFICATION (AISC Table B4.1a)"
        ws_summary[f'A{row}'].font = subtitle_font
        ws_summary.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws_summary[f'A{row}'] = "Element"
        ws_summary[f'B{row}'] = "Classification"
        ws_summary[f'C{row}'] = "Details"
        for col in range(1, 4):
            cell = ws_summary.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Overall classification
        row += 1
        ws_summary[f'A{row}'] = "Overall Section"
        ws_summary[f'B{row}'] = comp_class['overall_class']
        
        if comp_class['overall_class'] == "Slender":
            ws_summary[f'C{row}'] = f"Limiting: {comp_class['limiting_element']}"
        else:
            ws_summary[f'C{row}'] = "All elements within limits"
        
        for col in range(1, 4):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center_align if col > 1 else left_align
            
            if col == 2:  # Classification column
                if comp_class['overall_class'] == "Non-slender":
                    cell.fill = compact_fill
                    cell.font = Font(bold=True, color="2E7D32")
                else:
                    cell.fill = slender_fill
                    cell.font = Font(bold=True, color="C62828")
        
        # Flange details
        row += 1
        ws_summary[f'A{row}'] = "Flange"
        ws_summary[f'B{row}'] = "Slender" if comp_class['flange_slender'] else "Non-slender"
        ws_summary[f'C{row}'] = f"λ={comp_class['flange_lambda']:.2f}, λr={comp_class['flange_lambda_r']:.2f}"
        
        for col in range(1, 4):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center_align if col > 1 else left_align
        
        # Web details
        row += 1
        ws_summary[f'A{row}'] = "Web"
        ws_summary[f'B{row}'] = "Slender" if comp_class['web_slender'] else "Non-slender"
        ws_summary[f'C{row}'] = f"λ={comp_class['web_lambda']:.2f}, λr={comp_class['web_lambda_r']:.2f}"
        
        for col in range(1, 4):
            cell = ws_summary.cell(row=row, column=col)
            cell.border = border
            cell.alignment = center_align if col > 1 else left_align
    
    # Auto-size columns for Summary sheet
    try:
        for col_idx in range(1, ws_summary.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, ws_summary.max_row + 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
    except:
        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 30
        ws_summary.column_dimensions['D'].width = 15
    
    # Sheet 2: Analysis Results
    if analysis_results:
        ws_results = wb.create_sheet("Analysis Results")
        
        row = 1
        ws_results['A1'] = "DESIGN ANALYSIS RESULTS"
        ws_results['A1'].font = title_font
        ws_results.merge_cells('A1:D1')
        
        # Flexural Analysis
        if 'flexural' in analysis_results:
            row = 3
            ws_results[f'A{row}'] = "FLEXURAL ANALYSIS (AISC F2)"
            ws_results[f'A{row}'].font = subtitle_font
            ws_results.merge_cells(f'A{row}:B{row}')
            
            row += 1
            ws_results[f'A{row}'] = "Parameter"
            ws_results[f'B{row}'] = "Value"
            for col in range(1, 3):
                cell = ws_results.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            
            flex_data = [
                ['Design Moment (φMn)', f"{analysis_results['flexural']['phi_Mn']:.2f} t·m"],
                ['Nominal Moment (Mn)', f"{analysis_results['flexural']['Mn']:.2f} t·m"],
                ['Plastic Moment (Mp)', f"{analysis_results['flexural']['Mp']:.2f} t·m"],
                ['Critical Length (Lp)', f"{analysis_results['flexural']['Lp']:.3f} m"],
                ['Critical Length (Lr)', f"{analysis_results['flexural']['Lr']:.3f} m"],
                ['Design Case', analysis_results['flexural']['case']],
                ['Design Zone', analysis_results['flexural']['zone']],
                ['Utilization Ratio', f"{analysis_results['flexural']['ratio']:.3f}"],
                ['Status', '✓ ADEQUATE' if analysis_results['flexural']['adequate'] else '✗ INADEQUATE']
            ]
            
            for data_row in flex_data:
                row += 1
                ws_results[f'A{row}'] = data_row[0]
                ws_results[f'B{row}'] = data_row[1]
                for col in range(1, 3):
                    cell = ws_results.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = left_align if col == 1 else center_align
                    
                    # Color code status
                    if col == 2 and data_row[0] == 'Status':
                        if '✓' in str(data_row[1]):
                            cell.fill = compact_fill
                            cell.font = Font(bold=True, color="2E7D32")
                        else:
                            cell.fill = slender_fill
                            cell.font = Font(bold=True, color="C62828")
        
        # Compression Analysis
        if 'compression' in analysis_results:
            row += 3
            ws_results[f'A{row}'] = "COMPRESSION ANALYSIS (AISC E3)"
            ws_results[f'A{row}'].font = subtitle_font
            ws_results.merge_cells(f'A{row}:B{row}')
            
            row += 1
            ws_results[f'A{row}'] = "Parameter"
            ws_results[f'B{row}'] = "Value"
            for col in range(1, 3):
                cell = ws_results.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            
            comp_data = [
                ['Design Strength (φPn)', f"{analysis_results['compression']['phi_Pn']:.2f} tons"],
                ['Nominal Strength (Pn)', f"{analysis_results['compression']['Pn']:.2f} tons"],
                ['Critical Stress (Fcr)', f"{analysis_results['compression']['Fcr']:.1f} kgf/cm²"],
                ['Slenderness Ratio (λc)', f"{analysis_results['compression']['lambda_c']:.1f}"],
                ['Buckling Mode', analysis_results['compression']['mode']],
                ['Utilization Ratio', f"{analysis_results['compression']['ratio']:.3f}"],
                ['Status', '✓ ADEQUATE' if analysis_results['compression']['adequate'] else '✗ INADEQUATE']
            ]
            
            for data_row in comp_data:
                row += 1
                ws_results[f'A{row}'] = data_row[0]
                ws_results[f'B{row}'] = data_row[1]
                for col in range(1, 3):
                    cell = ws_results.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = left_align if col == 1 else center_align
                    
                    # Color code status
                    if col == 2 and data_row[0] == 'Status':
                        if '✓' in str(data_row[1]):
                            cell.fill = compact_fill
                            cell.font = Font(bold=True, color="2E7D32")
                        else:
                            cell.fill = slender_fill
                            cell.font = Font(bold=True, color="C62828")
        
        # Design Parameters Sheet
        ws_params = wb.create_sheet("Design Parameters")
        
        row = 1
        ws_params['A1'] = "DESIGN INPUT PARAMETERS"
        ws_params['A1'].font = title_font
        ws_params.merge_cells('A1:B1')
        
        row = 3
        ws_params[f'A{row}'] = "Parameter"
        ws_params[f'B{row}'] = "Value"
        for col in range(1, 3):
            cell = ws_params.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        if design_params:
            param_data = []
            if 'Mu' in design_params:
                param_data.append(['Applied Moment (Mu)', f"{design_params['Mu']:.2f} t·m"])
            if 'Pu' in design_params:
                param_data.append(['Applied Axial Load (Pu)', f"{design_params['Pu']:.2f} tons"])
            if 'Lb' in design_params:
                param_data.append(['Unbraced Length (Lb)', f"{design_params['Lb']:.2f} m"])
            if 'KL' in design_params:
                param_data.append(['Effective Length (KL)', f"{design_params['KL']:.2f} m"])
            if 'KLx' in design_params:
                param_data.append(['Effective Length X (KLx)', f"{design_params['KLx']:.2f} m"])
            if 'KLy' in design_params:
                param_data.append(['Effective Length Y (KLy)', f"{design_params['KLy']:.2f} m"])
            if 'Cb' in design_params:
                param_data.append(['Moment Gradient Factor (Cb)', f"{design_params['Cb']:.2f}"])
            
            for data_row in param_data:
                row += 1
                ws_params[f'A{row}'] = data_row[0]
                ws_params[f'B{row}'] = data_row[1]
                for col in range(1, 3):
                    cell = ws_params.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = left_align if col == 1 else center_align
        
        # Auto-size columns for all sheets
        for ws in [ws_results, ws_params]:
            try:
                for col_idx in range(1, ws.max_column + 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = 0
                    for row_idx in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    ws.column_dimensions[column_letter].width = adjusted_width
            except:
                ws.column_dimensions['A'].width = 40
                ws.column_dimensions['B'].width = 30
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer
def generate_calculation_report(df, df_mat, section, material, analysis_results, design_params, project_info):
    """
    Generate engineering calculation report with:
    - Project information header
    - Hand calculation style (detailed step-by-step)
    - Summary tables
    - Graphs/Charts
    - formatting
    """
    if not PDF_AVAILABLE:
        return None
    
    buffer = BytesIO()
    
    # Create document
    doc = BaseDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=0.6*inch, 
        leftMargin=0.6*inch,
        topMargin=1.2*inch,
        bottomMargin=0.8*inch,
        title="AISC 360-16 Structural Calculation"
    )
    
    # Define frame
    frame = Frame(
        doc.leftMargin, 
        doc.bottomMargin, 
        doc.width, 
        doc.height - 0.2*inch,
        id='normal',
        topPadding=10,
        bottomPadding=10
    )
    template = PageTemplate(id='main', frames=frame, onPage=lambda c, d: None)
    doc.addPageTemplates([template])
    
    story = []
    styles = getSampleStyleSheet()
    
    # ==================== CUSTOM STYLES ====================
    title_style = ParagraphStyle(
        'CalcTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=rl_colors.HexColor('#1a237e'),
        spaceAfter=12,
        spaceBefore=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        leading=22
    )
    
    heading1_style = ParagraphStyle(
        'CalcHeading1',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=rl_colors.white,
        spaceAfter=8,
        spaceBefore=14,
        fontName='Helvetica-Bold',
        backColor=rl_colors.HexColor('#1a237e'),
        borderPadding=8,
        leading=18
    )
    
    heading2_style = ParagraphStyle(
        'CalcHeading2',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=rl_colors.HexColor('#1565c0'),
        spaceAfter=6,
        spaceBefore=10,
        fontName='Helvetica-Bold',
        borderWidth=0,
        borderColor=rl_colors.HexColor('#1565c0'),
        borderPadding=4,
        leading=16
    )
    
    body_style = ParagraphStyle(
        'CalcBody',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        spaceAfter=4,
        spaceBefore=2,
        alignment=TA_LEFT
    )
    
    equation_style = ParagraphStyle(
        'CalcEquation',
        parent=styles['Code'],
        fontSize=10,
        textColor=rl_colors.HexColor('#0d47a1'),
        backColor=rl_colors.HexColor('#e3f2fd'),
        borderWidth=1,
        borderColor=rl_colors.HexColor('#1976d2'),
        borderPadding=10,
        leftIndent=20,
        rightIndent=20,
        spaceAfter=8,
        spaceBefore=6,
        fontName='Courier',
        leading=14,
        alignment=TA_LEFT
    )
    
    result_style = ParagraphStyle(
        'CalcResult',
        parent=styles['Normal'],
        fontSize=11,
        textColor=rl_colors.HexColor('#1b5e20'),
        backColor=rl_colors.HexColor('#e8f5e9'),
        borderWidth=2,
        borderColor=rl_colors.HexColor('#4caf50'),
        borderPadding=10,
        spaceAfter=10,
        spaceBefore=6,
        fontName='Helvetica-Bold',
        leading=16,
        alignment=TA_CENTER
    )
    
    reference_style = ParagraphStyle(
        'CalcReference',
        parent=styles['Normal'],
        fontSize=9,
        textColor=rl_colors.HexColor('#616161'),
        fontName='Helvetica-Oblique',
        leading=12,
        leftIndent=30
    )
    
    # ==================== PROJECT HEADER TABLE ====================
    story.append(Spacer(1, 0.1*inch))
    
    # Project info table (engineering header)
    header_data = [
        [Paragraph('<b>STRUCTURAL CALCULATION SHEET</b>', 
                  ParagraphStyle('header', fontSize=14, alignment=TA_CENTER, textColor=rl_colors.white)),
         '', '', ''],
        ['Project:', project_info.get('project_name', 'N/A'), 
         'Project No.:', project_info.get('project_no', 'N/A')],
        ['Subject:', f'Steel Design - {section}', 
         'Date:', project_info.get('date', datetime.now().strftime('%Y-%m-%d'))],
        ['Designer:', project_info.get('designer', 'N/A'), 
         'Checker:', project_info.get('checker', 'N/A')],
        ['Reference:', 'AISC 360-16', 
         'Revision:', project_info.get('revision', '0')],
    ]
    
    header_table = Table(header_data, colWidths=[1.2*inch, 2.8*inch, 1.0*inch, 1.5*inch])
    header_table.setStyle(TableStyle([
        # Title row
        ('SPAN', (0, 0), (-1, 0)),
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Data rows
        ('BACKGROUND', (0, 1), (0, -1), rl_colors.HexColor('#e8eaf6')),
        ('BACKGROUND', (2, 1), (2, -1), rl_colors.HexColor('#e8eaf6')),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, rl_colors.HexColor('#1a237e')),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.2*inch))
    
    # ==================== TABLE OF CONTENTS ====================
    story.append(Paragraph("TABLE OF CONTENTS", heading1_style))
    story.append(Spacer(1, 6))
    
    toc_data = [
        ['1.', 'Design Data & Material Properties', ''],
        ['2.', 'Section Properties', ''],
        ['3.', 'Section Classification (AISC Table B4.1)', ''],
        ['4.', 'Flexural Design (AISC Chapter F2)', ''],
        ['5.', 'Compression Design (AISC Chapter E3)', ''],
        ['6.', 'Design Summary & Conclusion', ''],
    ]
    
    toc_table = Table(toc_data, colWidths=[0.4*inch, 5*inch, 1*inch])
    toc_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
    ]))
    story.append(toc_table)
    story.append(Spacer(1, 0.15*inch))
    
    # ==================== 1. DESIGN DATA ====================
    story.append(Paragraph("1. DESIGN DATA & MATERIAL PROPERTIES", heading1_style))
    story.append(Spacer(1, 8))
    
    # Get material properties
    Fy = safe_scalar(df_mat.loc[material, "Yield Point (ksc)"])
    Fu = safe_scalar(df_mat.loc[material, "Tensile Strength (ksc)"])
    E = safe_scalar(df_mat.loc[material, "E"])
    
    story.append(Paragraph("<b>1.1 Material Properties</b>", heading2_style))
    story.append(Paragraph(f"Steel Grade: <b>{material}</b>", body_style))
    story.append(Spacer(1, 4))
    
    mat_table_data = [
        ['Property', 'Symbol', 'Value', 'Unit'],
        ['Yield Strength', 'Fy', f'{Fy:.1f}', 'kgf/cm²'],
        ['Tensile Strength', 'Fu', f'{Fu:.1f}', 'kgf/cm²'],
        ['Modulus of Elasticity', 'E', f'{E:,.0f}', 'kgf/cm²'],
        ['Poisson\'s Ratio', 'ν', '0.30', '-'],
        ['Shear Modulus', 'G', f'{E/(2*1.3):,.0f}', 'kgf/cm²'],
    ]
    
    mat_table = Table(mat_table_data, colWidths=[2.2*inch, 1*inch, 1.5*inch, 1.3*inch])
    mat_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1565c0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#f5f5f5')])
    ]))
    story.append(mat_table)
    story.append(Spacer(1, 10))
    
    # Design Parameters
    story.append(Paragraph("<b>1.2 Design Parameters (Input)</b>", heading2_style))
    
    design_table_data = [['Parameter', 'Symbol', 'Value', 'Unit']]
    
    if 'Mu' in design_params:
        design_table_data.append(['Required Moment Strength', 'Mu', f"{design_params['Mu']:.2f}", 't·m'])
    if 'Pu' in design_params:
        design_table_data.append(['Required Axial Strength', 'Pu', f"{design_params['Pu']:.2f}", 'tons'])
    if 'Lb' in design_params:
        design_table_data.append(['Unbraced Length', 'Lb', f"{design_params['Lb']:.2f}", 'm'])
    if 'KL' in design_params:
        design_table_data.append(['Effective Length', 'KL', f"{design_params['KL']:.2f}", 'm'])
    if 'Cb' in design_params:
        design_table_data.append(['Moment Gradient Factor', 'Cb', f"{design_params['Cb']:.2f}", '-'])
    
    design_table = Table(design_table_data, colWidths=[2.2*inch, 1*inch, 1.5*inch, 1.3*inch])
    design_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#ff6f00')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
    ]))
    story.append(design_table)
    
    # ==================== 2. SECTION PROPERTIES ====================
    story.append(PageBreak())
    story.append(Paragraph("2. SECTION PROPERTIES", heading1_style))
    story.append(Spacer(1, 8))
    
    story.append(Paragraph(f"<b>Selected Section: {section}</b>", heading2_style))
    story.append(Spacer(1, 4))
    
    # Get all section properties
    d = safe_scalar(df.loc[section, 'd [mm]'])
    bf = safe_scalar(df.loc[section, 'bf [mm]'])
    tf = safe_scalar(df.loc[section, 'tf [mm]'])
    tw = safe_scalar(df.loc[section, 'tw [mm]'])
    r = safe_scalar(df.loc[section, 'r [mm]']) if 'r [mm]' in df.columns else 0
    A = safe_scalar(df.loc[section, 'A [cm2]'])
    Ix = safe_scalar(df.loc[section, 'Ix [cm4]'])
    Iy = safe_scalar(df.loc[section, 'Iy [cm4]'])
    rx = safe_scalar(df.loc[section, 'rx [cm]'])
    ry = safe_scalar(df.loc[section, 'ry [cm]'])
    Sx = safe_scalar(df.loc[section, 'Sx [cm3]'])
    Sy = safe_scalar(df.loc[section, 'Sy [cm3]'])
    Zx = safe_scalar(df.loc[section, 'Zx [cm3]'])
    Zy = safe_scalar(df.loc[section, 'Zy [cm3]'])
    J = safe_scalar(df.loc[section, 'j [cm4]']) if 'j [cm4]' in df.columns else 0
    ho = safe_scalar(df.loc[section, 'ho [mm]']) if 'ho [mm]' in df.columns else (d - tf)
    rts = safe_scalar(df.loc[section, 'rts [cm]']) if 'rts [cm]' in df.columns else ry * 1.2
    
    weight_col = 'Unit Weight [kg/m]' if 'Unit Weight [kg/m]' in df.columns else 'w [kg/m]'
    weight = safe_scalar(df.loc[section, weight_col])
    
    # Section properties in two columns
    props_data_1 = [
        ['Geometric Properties', '', ''],
        ['Overall Depth', 'd', f'{d:.0f} mm'],
        ['Flange Width', 'bf', f'{bf:.0f} mm'],
        ['Flange Thickness', 'tf', f'{tf:.1f} mm'],
        ['Web Thickness', 'tw', f'{tw:.1f} mm'],
        ['Fillet Radius', 'r', f'{r:.1f} mm'],
        ['Unit Weight', 'w', f'{weight:.1f} kg/m'],
    ]
    
    props_data_2 = [
        ['Section Properties', '', ''],
        ['Cross-sectional Area', 'A', f'{A:.2f} cm²'],
        ['Moment of Inertia (x)', 'Ix', f'{Ix:,.0f} cm⁴'],
        ['Moment of Inertia (y)', 'Iy', f'{Iy:,.0f} cm⁴'],
        ['Radius of Gyration (x)', 'rx', f'{rx:.2f} cm'],
        ['Radius of Gyration (y)', 'ry', f'{ry:.2f} cm'],
        ['Torsional Constant', 'J', f'{J:.2f} cm⁴'],
    ]
    
    props_data_3 = [
        ['Section Moduli', '', ''],
        ['Elastic Modulus (x)', 'Sx', f'{Sx:,.1f} cm³'],
        ['Elastic Modulus (y)', 'Sy', f'{Sy:,.1f} cm³'],
        ['Plastic Modulus (x)', 'Zx', f'{Zx:,.1f} cm³'],
        ['Plastic Modulus (y)', 'Zy', f'{Zy:,.1f} cm³'],
        ['Distance b/w flanges', 'ho', f'{ho:.1f} mm'],
        ['Effective radius (LTB)', 'rts', f'{rts:.2f} cm'],
    ]
    
    # Create combined table
    def create_props_table(data):
        table = Table(data, colWidths=[2.2*inch, 0.8*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#37474f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('SPAN', (0, 0), (-1, 0)),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#fafafa')])
        ]))
        return table
    
    # Create a combined layout
    combined_data = [[create_props_table(props_data_1), create_props_table(props_data_2)]]
    combined_table = Table(combined_data, colWidths=[3.3*inch, 3.3*inch])
    combined_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(combined_table)
    story.append(Spacer(1, 8))
    story.append(create_props_table(props_data_3))
    
    # Add section diagram
    story.append(Spacer(1, 0.15*inch))
    story.append(Paragraph("<b>Section Cross-Section:</b>", body_style))
    story.append(Spacer(1, 6))
    
    # Generate section diagram
    fig_section = create_detailed_section_diagram(d, bf, tf, tw, section)
    img_buffer_section = BytesIO()
    plt.savefig(img_buffer_section, format='png', dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig_section)
    img_buffer_section.seek(0)
    
    section_img = Image(img_buffer_section, width=4*inch, height=3*inch)
    story.append(section_img)
    
    # ==================== 3. SECTION CLASSIFICATION ====================
    story.append(PageBreak())
    story.append(Paragraph("3. SECTION CLASSIFICATION (AISC Table B4.1)", heading1_style))
    story.append(Spacer(1, 8))
    
    # Get classifications
    flex_class = classify_section_flexure(df, df_mat, section, material)
    comp_class = classify_section_compression(df, df_mat, section, material)
    
    if flex_class:
        story.append(Paragraph("<b>3.1 Flexural Member Classification (Table B4.1b)</b>", heading2_style))
        story.append(Spacer(1, 4))
        
        # Flange Classification - Hand Calculation Style
        story.append(Paragraph("<b>Flange Classification (Case 10: Flanges of I-shaped sections)</b>", body_style))
        story.append(Paragraph("<i>Reference: AISC 360-16, Table B4.1b</i>", reference_style))
        story.append(Spacer(1, 4))
        
        story.append(Paragraph(
            f"<font face='Courier'>λf = bf / (2·tf) = {bf:.1f} / (2 × {tf:.1f}) = <b>{flex_class['flange_lambda']:.2f}</b></font>",
            equation_style
        ))
        
        story.append(Paragraph(
            f"<font face='Courier'>λpf = 0.38·√(E/Fy) = 0.38 × √({E:,.0f}/{Fy:.1f}) = <b>{flex_class['flange_lambda_p']:.2f}</b></font>",
            equation_style
        ))
        
        story.append(Paragraph(
            f"<font face='Courier'>λrf = 1.0·√(E/Fy) = 1.0 × √({E:,.0f}/{Fy:.1f}) = <b>{flex_class['flange_lambda_r']:.2f}</b></font>",
            equation_style
        ))
        
        # Classification check
        if flex_class['flange_lambda'] <= flex_class['flange_lambda_p']:
            check_text = f"λf = {flex_class['flange_lambda']:.2f} ≤ λpf = {flex_class['flange_lambda_p']:.2f}"
            result_text = "FLANGE: COMPACT ✓"
            result_color = '#4caf50'
        elif flex_class['flange_lambda'] <= flex_class['flange_lambda_r']:
            check_text = f"λpf = {flex_class['flange_lambda_p']:.2f} < λf = {flex_class['flange_lambda']:.2f} ≤ λrf = {flex_class['flange_lambda_r']:.2f}"
            result_text = "FLANGE: NON-COMPACT"
            result_color = '#ff9800'
        else:
            check_text = f"λf = {flex_class['flange_lambda']:.2f} > λrf = {flex_class['flange_lambda_r']:.2f}"
            result_text = "FLANGE: SLENDER ✗"
            result_color = '#f44336'
        
        story.append(Paragraph(f"<font face='Courier'>Check: {check_text}</font>", equation_style))
        story.append(Paragraph(
            f"<b>{result_text}</b>",
            ParagraphStyle('FlgResult', parent=result_style, textColor=rl_colors.HexColor(result_color))
        ))
        story.append(Spacer(1, 10))
        
        # Web Classification - Hand Calculation Style
        story.append(Paragraph("<b>Web Classification (Case 15: Webs in flexural compression)</b>", body_style))
        story.append(Paragraph("<i>Reference: AISC 360-16, Table B4.1b</i>", reference_style))
        story.append(Spacer(1, 4))
        
        h_web = ho if 'ho [mm]' in df.columns else (d - 2*tf)
        story.append(Paragraph(
            f"<font face='Courier'>λw = h / tw = {h_web:.1f} / {tw:.1f} = <b>{flex_class['web_lambda']:.2f}</b></font>",
            equation_style
        ))
        
        story.append(Paragraph(
            f"<font face='Courier'>λpw = 3.76·√(E/Fy) = 3.76 × √({E:,.0f}/{Fy:.1f}) = <b>{flex_class['web_lambda_p']:.2f}</b></font>",
            equation_style
        ))
        
        story.append(Paragraph(
            f"<font face='Courier'>λrw = 5.70·√(E/Fy) = 5.70 × √({E:,.0f}/{Fy:.1f}) = <b>{flex_class['web_lambda_r']:.2f}</b></font>",
            equation_style
        ))
        
        # Web classification check
        if flex_class['web_lambda'] <= flex_class['web_lambda_p']:
            check_text = f"λw = {flex_class['web_lambda']:.2f} ≤ λpw = {flex_class['web_lambda_p']:.2f}"
            result_text = "WEB: COMPACT ✓"
            result_color = '#4caf50'
        elif flex_class['web_lambda'] <= flex_class['web_lambda_r']:
            check_text = f"λpw = {flex_class['web_lambda_p']:.2f} < λw = {flex_class['web_lambda']:.2f} ≤ λrw = {flex_class['web_lambda_r']:.2f}"
            result_text = "WEB: NON-COMPACT"
            result_color = '#ff9800'
        else:
            check_text = f"λw = {flex_class['web_lambda']:.2f} > λrw = {flex_class['web_lambda_r']:.2f}"
            result_text = "WEB: SLENDER ✗"
            result_color = '#f44336'
        
        story.append(Paragraph(f"<font face='Courier'>Check: {check_text}</font>", equation_style))
        story.append(Paragraph(
            f"<b>{result_text}</b>",
            ParagraphStyle('WebResult', parent=result_style, textColor=rl_colors.HexColor(result_color))
        ))
    
    # ==================== 4. FLEXURAL DESIGN ====================
    if analysis_results and 'flexural' in analysis_results:
        story.append(PageBreak())
        story.append(Paragraph("4. FLEXURAL DESIGN (AISC Chapter F2)", heading1_style))
        story.append(Spacer(1, 8))
        
        flex = analysis_results['flexural']
        Lb = design_params.get('Lb', 0)
        Cb = design_params.get('Cb', 1.0)
        
        # Step 1: Plastic Moment
        story.append(Paragraph("<b>Step 1: Calculate Plastic Moment Capacity (Mp)</b>", heading2_style))
        story.append(Paragraph("<i>Reference: AISC 360-16, Equation F2-1</i>", reference_style))
        story.append(Spacer(1, 4))
        
        Mp_kgcm = Fy * Zx
        Mp_tm = Mp_kgcm / 100000
        
        story.append(Paragraph(
            f"<font face='Courier'>Mp = Fy × Zx</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Mp = {Fy:.1f} × {Zx:,.1f}</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Mp = {Mp_kgcm:,.0f} kgf·cm = <b>{Mp_tm:.2f} t·m</b></font>",
            equation_style
        ))
        story.append(Spacer(1, 8))
        
        # Step 2: Limiting Lengths
        story.append(Paragraph("<b>Step 2: Calculate Limiting Laterally Unbraced Lengths</b>", heading2_style))
        story.append(Spacer(1, 4))
        
        # Lp calculation
        story.append(Paragraph("<b>Compact Limit Lp (Equation F2-5):</b>", body_style))
        story.append(Paragraph("<i>Reference: AISC 360-16, Equation F2-5</i>", reference_style))
        
        Lp_cm = 1.76 * ry * math.sqrt(E / Fy)
        story.append(Paragraph(
            f"<font face='Courier'>Lp = 1.76 × ry × √(E/Fy)</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Lp = 1.76 × {ry:.2f} × √({E:,.0f}/{Fy:.1f})</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Lp = {Lp_cm:.1f} cm = <b>{flex['Lp']:.3f} m</b></font>",
            equation_style
        ))
        story.append(Spacer(1, 6))
        
        # Lr calculation
        story.append(Paragraph("<b>Inelastic Limit Lr (Equation F2-6):</b>", body_style))
        story.append(Paragraph("<i>Reference: AISC 360-16, Equation F2-6</i>", reference_style))
        
        story.append(Paragraph(
            f"<font face='Courier'>Lr = 1.95 × rts × (E/0.7Fy) × √(Jc/Sxho) × √(1 + √(1 + 6.76(0.7Fy/E)²(Sxho/Jc)²))</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>where: rts = {rts:.2f} cm, J = {J:.2f} cm⁴, ho = {ho:.1f} mm, c = 1.0</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Lr = <b>{flex['Lr']:.3f} m</b></font>",
            equation_style
        ))
        story.append(Spacer(1, 8))
        
        # Step 3: Determine Nominal Moment
        story.append(Paragraph(f"<b>Step 3: Determine Nominal Moment (Lb = {Lb:.2f} m)</b>", heading2_style))
        story.append(Spacer(1, 4))
        
        if Lb <= flex['Lp']:
            story.append(Paragraph(
                f"<font face='Courier'>Lb = {Lb:.2f} m ≤ Lp = {flex['Lp']:.3f} m</font>",
                equation_style
            ))
            story.append(Paragraph("<b>→ YIELDING CONTROLS (F2.1) - LTB does not apply</b>", body_style))
            story.append(Paragraph("<i>Reference: AISC 360-16, Equation F2-1</i>", reference_style))
            story.append(Paragraph(
                f"<font face='Courier'>Mn = Mp = <b>{flex['Mn']:.2f} t·m</b></font>",
                equation_style
            ))
            
        elif Lb <= flex['Lr']:
            story.append(Paragraph(
                f"<font face='Courier'>Lp = {flex['Lp']:.3f} m < Lb = {Lb:.2f} m ≤ Lr = {flex['Lr']:.3f} m</font>",
                equation_style
            ))
            story.append(Paragraph("<b>→ INELASTIC LATERAL-TORSIONAL BUCKLING (F2.2)</b>", body_style))
            story.append(Paragraph("<i>Reference: AISC 360-16, Equation F2-2</i>", reference_style))
            
            Mr = 0.7 * Fy * Sx / 100000
            story.append(Paragraph(
                f"<font face='Courier'>Mn = Cb × [Mp - (Mp - 0.7FySx) × ((Lb-Lp)/(Lr-Lp))] ≤ Mp</font>",
                equation_style
            ))
            story.append(Paragraph(
                f"<font face='Courier'>Mn = {Cb:.2f} × [{Mp_tm:.2f} - ({Mp_tm:.2f} - {Mr:.2f}) × (({Lb:.2f}-{flex['Lp']:.3f})/({flex['Lr']:.3f}-{flex['Lp']:.3f}))]</font>",
                equation_style
            ))
            story.append(Paragraph(
                f"<font face='Courier'>Mn = <b>{flex['Mn']:.2f} t·m</b></font>",
                equation_style
            ))
            
        else:
            story.append(Paragraph(
                f"<font face='Courier'>Lb = {Lb:.2f} m > Lr = {flex['Lr']:.3f} m</font>",
                equation_style
            ))
            story.append(Paragraph("<b>→ ELASTIC LATERAL-TORSIONAL BUCKLING (F2.3)</b>", body_style))
            story.append(Paragraph("<i>Reference: AISC 360-16, Equations F2-3 and F2-4</i>", reference_style))
            
            story.append(Paragraph(
                f"<font face='Courier'>Fcr = (Cb×π²×E)/((Lb/rts)²) × √(1 + 0.078(Jc/Sxho)(Lb/rts)²)</font>",
                equation_style
            ))
            story.append(Paragraph(
                f"<font face='Courier'>Mn = Fcr × Sx ≤ Mp</font>",
                equation_style
            ))
            story.append(Paragraph(
                f"<font face='Courier'>Mn = <b>{flex['Mn']:.2f} t·m</b></font>",
                equation_style
            ))
        
        story.append(Spacer(1, 8))
        
        # Step 4: Design Strength
        story.append(Paragraph("<b>Step 4: Calculate Design Flexural Strength</b>", heading2_style))
        story.append(Paragraph("<i>Reference: AISC 360-16, Section F1</i>", reference_style))
        story.append(Spacer(1, 4))
        
        story.append(Paragraph(
            f"<font face='Courier'>φb = 0.90 (LRFD)</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>φbMn = 0.90 × {flex['Mn']:.2f} = <b>{flex['phi_Mn']:.2f} t·m</b></font>",
            equation_style
        ))
        story.append(Spacer(1, 8))
        
        # Step 5: Adequacy Check
        story.append(Paragraph("<b>Step 5: Adequacy Check</b>", heading2_style))
        story.append(Spacer(1, 4))
        
        Mu = design_params.get('Mu', 0)
        ratio = flex['ratio']
        
        story.append(Paragraph(
            f"<font face='Courier'>Mu = {Mu:.2f} t·m</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>φbMn = {flex['phi_Mn']:.2f} t·m</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Ratio = Mu / φbMn = {Mu:.2f} / {flex['phi_Mn']:.2f} = <b>{ratio:.3f}</b></font>",
            equation_style
        ))
        
        if flex['adequate']:
            story.append(Paragraph(
                f"<font face='Courier'>Check: {ratio:.3f} ≤ 1.0 → <b>OK ✓</b></font>",
                equation_style
            ))
            story.append(Paragraph("<b>FLEXURAL DESIGN: ADEQUATE ✓</b>", result_style))
        else:
            story.append(Paragraph(
                f"<font face='Courier'>Check: {ratio:.3f} > 1.0 → <b>NG ✗</b></font>",
                equation_style
            ))
            story.append(Paragraph(
                "<b>FLEXURAL DESIGN: INADEQUATE ✗</b>",
                ParagraphStyle('FailResult', parent=result_style, 
                              textColor=rl_colors.HexColor('#c62828'),
                              backColor=rl_colors.HexColor('#ffebee'),
                              borderColor=rl_colors.HexColor('#f44336'))
            ))
        
        # Add flexural capacity chart
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph("<b>Flexural Capacity Curve:</b>", body_style))
        story.append(Spacer(1, 6))
        
        fig_flex = create_flexural_capacity_chart(df, df_mat, section, material, Lb, Cb, flex)
        img_buffer_flex = BytesIO()
        plt.savefig(img_buffer_flex, format='png', dpi=150, bbox_inches='tight',
                    facecolor='white', edgecolor='none')
        plt.close(fig_flex)
        img_buffer_flex.seek(0)
        
        flex_img = Image(img_buffer_flex, width=5.5*inch, height=3.5*inch)
        story.append(flex_img)
    
    # ==================== 5. COMPRESSION DESIGN ====================
    if analysis_results and 'compression' in analysis_results:
        story.append(PageBreak())
        story.append(Paragraph("5. COMPRESSION DESIGN (AISC Chapter E3)", heading1_style))
        story.append(Spacer(1, 8))
        
        comp = analysis_results['compression']
        KL = design_params.get('KL', 0)
        
        # Step 1: Slenderness Ratio
        story.append(Paragraph("<b>Step 1: Calculate Slenderness Ratio</b>", heading2_style))
        story.append(Paragraph("<i>Reference: AISC 360-16, Section E2</i>", reference_style))
        story.append(Spacer(1, 4))
        
        KL_cm = KL * 100
        lambda_x = KL_cm / rx
        lambda_y = KL_cm / ry
        lambda_c = max(lambda_x, lambda_y)
        
        story.append(Paragraph(
            f"<font face='Courier'>KL = {KL:.2f} m = {KL_cm:.0f} cm</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>(KL/r)x = {KL_cm:.0f} / {rx:.2f} = <b>{lambda_x:.1f}</b></font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>(KL/r)y = {KL_cm:.0f} / {ry:.2f} = <b>{lambda_y:.1f}</b></font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Governing: (KL/r)max = <b>{lambda_c:.1f}</b> {'(X-axis)' if lambda_x >= lambda_y else '(Y-axis)'}</font>",
            equation_style
        ))
        
        # Check slenderness limit
        if lambda_c <= 200:
            story.append(Paragraph(
                f"<font face='Courier'>Check: {lambda_c:.1f} ≤ 200 → OK (AISC E2)</font>",
                equation_style
            ))
        else:
            story.append(Paragraph(
                f"<font face='Courier'>Check: {lambda_c:.1f} > 200 → Exceeds preferred limit (AISC E2)</font>",
                equation_style
            ))
        story.append(Spacer(1, 8))
        
        # Step 2: Elastic Buckling Stress
        story.append(Paragraph("<b>Step 2: Calculate Elastic Buckling Stress (Fe)</b>", heading2_style))
        story.append(Paragraph("<i>Reference: AISC 360-16, Equation E3-4</i>", reference_style))
        story.append(Spacer(1, 4))
        
        Fe = (math.pi**2 * E) / (lambda_c**2)
        
        story.append(Paragraph(
            f"<font face='Courier'>Fe = π²E / (KL/r)²</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Fe = π² × {E:,.0f} / {lambda_c:.1f}²</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Fe = <b>{Fe:,.1f} kgf/cm²</b></font>",
            equation_style
        ))
        story.append(Spacer(1, 8))
        
        # Step 3: Critical Stress
        story.append(Paragraph("<b>Step 3: Determine Critical Stress (Fcr)</b>", heading2_style))
        story.append(Spacer(1, 4))
        
        lambda_limit = 4.71 * math.sqrt(E / Fy)
        
        story.append(Paragraph(
            f"<font face='Courier'>Limiting Slenderness = 4.71 × √(E/Fy) = 4.71 × √({E:,.0f}/{Fy:.1f}) = {lambda_limit:.1f}</font>",
            equation_style
        ))
        
        if lambda_c <= lambda_limit:
            story.append(Paragraph(
                f"<font face='Courier'>(KL/r) = {lambda_c:.1f} ≤ {lambda_limit:.1f}</font>",
                equation_style
            ))
            story.append(Paragraph("<b>→ INELASTIC BUCKLING CONTROLS (E3-2)</b>", body_style))
            story.append(Paragraph("<i>Reference: AISC 360-16, Equation E3-2</i>", reference_style))
            
            story.append(Paragraph(
                f"<font face='Courier'>Fcr = [0.658^(Fy/Fe)] × Fy</font>",
                equation_style
            ))
            story.append(Paragraph(
                f"<font face='Courier'>Fcr = [0.658^({Fy:.1f}/{Fe:.1f})] × {Fy:.1f}</font>",
                equation_style
            ))
        else:
            story.append(Paragraph(
                f"<font face='Courier'>(KL/r) = {lambda_c:.1f} > {lambda_limit:.1f}</font>",
                equation_style
            ))
            story.append(Paragraph("<b>→ ELASTIC BUCKLING CONTROLS (E3-3)</b>", body_style))
            story.append(Paragraph("<i>Reference: AISC 360-16, Equation E3-3</i>", reference_style))
            
            story.append(Paragraph(
                f"<font face='Courier'>Fcr = 0.877 × Fe</font>",
                equation_style
            ))
            story.append(Paragraph(
                f"<font face='Courier'>Fcr = 0.877 × {Fe:.1f}</font>",
                equation_style
            ))
        
        story.append(Paragraph(
            f"<font face='Courier'>Fcr = <b>{comp['Fcr']:,.1f} kgf/cm²</b></font>",
            equation_style
        ))
        story.append(Spacer(1, 8))
        
        # Step 4: Nominal and Design Strength
        story.append(Paragraph("<b>Step 4: Calculate Nominal and Design Strength</b>", heading2_style))
        story.append(Paragraph("<i>Reference: AISC 360-16, Equation E3-1</i>", reference_style))
        story.append(Spacer(1, 4))
        
        story.append(Paragraph(
            f"<font face='Courier'>Pn = Fcr × Ag</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Pn = {comp['Fcr']:,.1f} × {A:.2f} = {comp['Fcr']*A:,.0f} kgf = <b>{comp['Pn']:.2f} tons</b></font>",
            equation_style
        ))
        story.append(Spacer(1, 4))
        
        story.append(Paragraph(
            f"<font face='Courier'>φc = 0.90 (LRFD)</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>φcPn = 0.90 × {comp['Pn']:.2f} = <b>{comp['phi_Pn']:.2f} tons</b></font>",
            equation_style
        ))
        story.append(Spacer(1, 8))
        
        # Step 5: Adequacy Check
        story.append(Paragraph("<b>Step 5: Adequacy Check</b>", heading2_style))
        story.append(Spacer(1, 4))
        
        Pu = design_params.get('Pu', 0)
        ratio = comp['ratio']
        
        story.append(Paragraph(
            f"<font face='Courier'>Pu = {Pu:.2f} tons</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>φcPn = {comp['phi_Pn']:.2f} tons</font>",
            equation_style
        ))
        story.append(Paragraph(
            f"<font face='Courier'>Ratio = Pu / φcPn = {Pu:.2f} / {comp['phi_Pn']:.2f} = <b>{ratio:.3f}</b></font>",
            equation_style
        ))
        
        if comp['adequate']:
            story.append(Paragraph(
                f"<font face='Courier'>Check: {ratio:.3f} ≤ 1.0 → <b>OK ✓</b></font>",
                equation_style
            ))
            story.append(Paragraph("<b>COMPRESSION DESIGN: ADEQUATE ✓</b>", result_style))
        else:
            story.append(Paragraph(
                f"<font face='Courier'>Check: {ratio:.3f} > 1.0 → <b>NG ✗</b></font>",
                equation_style
            ))
            story.append(Paragraph(
                "<b>COMPRESSION DESIGN: INADEQUATE ✗</b>",
                ParagraphStyle('FailResult', parent=result_style, 
                              textColor=rl_colors.HexColor('#c62828'),
                              backColor=rl_colors.HexColor('#ffebee'),
                              borderColor=rl_colors.HexColor('#f44336'))
            ))
        
        # Add compression capacity chart
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph("<b>Column Capacity Curve:</b>", body_style))
        story.append(Spacer(1, 6))
        
        fig_comp = create_compression_capacity_chart(E, Fy, A, lambda_c, lambda_limit, comp, Pu)
        img_buffer_comp = BytesIO()
        plt.savefig(img_buffer_comp, format='png', dpi=150, bbox_inches='tight',
                    facecolor='white', edgecolor='none')
        plt.close(fig_comp)
        img_buffer_comp.seek(0)
        
        comp_img = Image(img_buffer_comp, width=5.5*inch, height=3.5*inch)
        story.append(comp_img)
    
    # ==================== 6. DESIGN SUMMARY ====================
    story.append(PageBreak())
    story.append(Paragraph("6. DESIGN SUMMARY & CONCLUSION", heading1_style))
    story.append(Spacer(1, 10))
    
    # Summary Table
    summary_data = [
        ['CHECK', 'CAPACITY', 'DEMAND', 'RATIO', 'STATUS'],
    ]
    
    if 'flexural' in analysis_results:
        flex = analysis_results['flexural']
        Mu = design_params.get('Mu', 0)
        status = '✓ PASS' if flex['adequate'] else '✗ FAIL'
        summary_data.append(['Flexural (φbMn)', f"{flex['phi_Mn']:.2f} t·m", 
                           f"{Mu:.2f} t·m", f"{flex['ratio']:.3f}", status])
    
    if 'compression' in analysis_results:
        comp = analysis_results['compression']
        Pu = design_params.get('Pu', 0)
        status = '✓ PASS' if comp['adequate'] else '✗ FAIL'
        summary_data.append(['Compression (φcPn)', f"{comp['phi_Pn']:.2f} tons",
                           f"{Pu:.2f} tons", f"{comp['ratio']:.3f}", status])
    
    summary_table = Table(summary_data, colWidths=[1.6*inch, 1.3*inch, 1.2*inch, 0.9*inch, 1*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, rl_colors.HexColor('#1a237e')),
    ]))
    
    # Color code status column
    for i, row in enumerate(summary_data[1:], start=1):
        if '✓' in row[-1]:
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (-1, i), (-1, i), rl_colors.HexColor('#c8e6c9')),
                ('TEXTCOLOR', (-1, i), (-1, i), rl_colors.HexColor('#2e7d32')),
                ('FONTNAME', (-1, i), (-1, i), 'Helvetica-Bold'),
            ]))
        else:
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (-1, i), (-1, i), rl_colors.HexColor('#ffcdd2')),
                ('TEXTCOLOR', (-1, i), (-1, i), rl_colors.HexColor('#c62828')),
                ('FONTNAME', (-1, i), (-1, i), 'Helvetica-Bold'),
            ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.2*inch))
    
    # Overall Conclusion
    overall_adequate = True
    if 'flexural' in analysis_results:
        overall_adequate = overall_adequate and analysis_results['flexural']['adequate']
    if 'compression' in analysis_results:
        overall_adequate = overall_adequate and analysis_results['compression']['adequate']
    
    if overall_adequate:
        conclusion_style = ParagraphStyle(
            'Conclusion',
            parent=result_style,
            fontSize=14,
            textColor=rl_colors.HexColor('#1b5e20'),
            backColor=rl_colors.HexColor('#c8e6c9'),
            borderColor=rl_colors.HexColor('#4caf50'),
            borderWidth=3
        )
        story.append(Paragraph(
            f"<b>CONCLUSION: SECTION {section} IS ADEQUATE FOR THE DESIGN LOADS ✓</b>",
            conclusion_style
        ))
    else:
        conclusion_style = ParagraphStyle(
            'Conclusion',
            parent=result_style,
            fontSize=14,
            textColor=rl_colors.HexColor('#b71c1c'),
            backColor=rl_colors.HexColor('#ffcdd2'),
            borderColor=rl_colors.HexColor('#f44336'),
            borderWidth=3
        )
        story.append(Paragraph(
            f"<b>CONCLUSION: SECTION {section} IS INADEQUATE - REVISE DESIGN ✗</b>",
            conclusion_style
        ))
    
    # Signature Block
    story.append(Spacer(1, 0.5*inch))
    
    sig_data = [
        ['PREPARED BY:', '', 'CHECKED BY:', ''],
        ['', '', '', ''],
        ['', '', '', ''],
        [f"Name: {project_info.get('designer', '_______________')}", '',
         f"Name: {project_info.get('checker', '_______________')}", ''],
        ['Date: _____________', '', 'Date: _____________', ''],
        ['Signature: _____________', '', 'Signature: _____________', ''],
    ]
    
    sig_table = Table(sig_data, colWidths=[2*inch, 0.7*inch, 2*inch, 0.7*inch])
    sig_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(sig_table)
    
    # Build PDF
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer

# ==================== ENHANCED CSS ====================
st.markdown("""
<style>
    /* Import Font */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    /* Global Styling */
    * {
        font-family: 'Inter', sans-serif;
    }
    
    .main {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
    }
    
    /* Enhanced Tab Styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 12px;
        background: white;
        padding: 1rem;
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.07);
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 60px;
        padding: 0 30px;
        background: #f8f9fa;
        border-radius: 10px;
        font-weight: 600;
        font-size: 15px;
        border: 2px solid transparent;
        transition: all 0.3s ease;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background: #e9ecef;
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.1);
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white !important;
        border-color: #667eea;
        box-shadow: 0 6px 12px rgba(102, 126, 234, 0.4);
    }
    
    /* Headers */
    .main-header {
        font-size: 2.8rem;
        font-weight: 800;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin: 2rem 0;
        letter-spacing: -1px;
    }
    
    .section-header {
        font-size: 1.8rem;
        font-weight: 700;
        color: #2c3e50;
        margin: 2rem 0 1.5rem 0;
        padding-bottom: 0.75rem;
        border-bottom: 3px solid #667eea;
        position: relative;
    }
    
    .section-header::after {
        content: '';
        position: absolute;
        bottom: -3px;
        left: 0;
        width: 100px;
        height: 3px;
        background: linear-gradient(90deg, #667eea, #764ba2);
    }
    
    /* Enhanced Card Designs */
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.07);
        border-left: 5px solid #667eea;
        margin: 1rem 0;
        transition: transform 0.3s ease, box-shadow 0.3s ease;
    }
    
    .metric-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 8px 16px rgba(0,0,0,0.12);
    }
    
    .evaluation-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 2rem;
        border-radius: 15px;
        margin: 1.5rem 0;
        box-shadow: 0 8px 16px rgba(102, 126, 234, 0.3);
    }
    
    .evaluation-card h3 {
        color: white;
        margin-bottom: 1rem;
        font-weight: 700;
    }
    
    .critical-lengths-box {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        box-shadow: 0 6px 12px rgba(245, 87, 108, 0.3);
    }
    
    .design-summary {
        background: white;
        border: 3px solid #4caf50;
        border-radius: 15px;
        padding: 1.5rem;
        margin: 1.5rem 0;
        box-shadow: 0 4px 8px rgba(76, 175, 80, 0.2);
    }
    
    /* Enhanced Status Boxes */
    .info-box {
        background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%);
        border-left: 5px solid #2196f3;
        padding: 1.25rem;
        border-radius: 10px;
        margin: 1rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    
    .success-box {
        background: linear-gradient(135deg, #e8f5e9 0%, #c8e6c9 100%);
        border-left: 5px solid #4caf50;
        padding: 1.25rem;
        border-radius: 10px;
        margin: 1rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    
    .warning-box {
        background: linear-gradient(135deg, #fff3e0 0%, #ffe0b2 100%);
        border-left: 5px solid #ff9800;
        padding: 1.25rem;
        border-radius: 10px;
        margin: 1rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    
    .error-box {
        background: linear-gradient(135deg, #ffebee 0%, #ffcdd2 100%);
        border-left: 5px solid #f44336;
        padding: 1.25rem;
        border-radius: 10px;
        margin: 1rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    
    /* AISC Equation Styling */
    .aisc-equation {
        background: linear-gradient(135deg, #e7f3ff 0%, #d0e8ff 100%);
        border-left: 5px solid #2196f3;
        padding: 1.25rem;
        margin: 1rem 0;
        border-radius: 10px;
        font-family: 'Courier New', monospace;
        font-size: 0.95rem;
        box-shadow: 0 2px 4px rgba(33, 150, 243, 0.1);
    }
    
    /* Enhanced Calculation Note */
    .calculation-note {
        background: #2c3e50;
        color: #ecf0f1;
        border: 2px solid #34495e;
        padding: 1.5rem;
        border-radius: 10px;
        font-family: 'Courier New', monospace;
        font-size: 0.9rem;
        margin: 1rem 0;
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
        max-height: 600px;
        overflow-y: auto;
    }
    
    /* Data Tables */
    .dataframe {
        font-size: 14px !important;
        border-radius: 10px;
        overflow: hidden;
    }
    
    .dataframe thead tr th {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white !important;
        font-weight: 600;
        padding: 15px 12px !important;
        text-align: center;
        font-size: 15px;
    }
    
    .dataframe tbody tr {
        transition: background-color 0.2s ease;
    }
    
    .dataframe tbody tr:hover {
        background-color: #f5f7fa !important;
    }
    
    .dataframe tbody tr td {
        padding: 12px !important;
        border-bottom: 1px solid #e9ecef;
        text-align: center;
    }
    
    /* Enhanced Metrics */
    [data-testid="stMetricValue"] {
        font-size: 2rem;
        font-weight: 700;
        color: #667eea;
    }
    
    [data-testid="stMetricLabel"] {
        font-size: 1rem;
        font-weight: 600;
        color: #2c3e50;
    }
    
    /* Buttons */
    .stButton>button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        padding: 12px 30px;
        border-radius: 10px;
        font-weight: 600;
        font-size: 15px;
        transition: all 0.3s ease;
        box-shadow: 0 4px 8px rgba(102, 126, 234, 0.3);
    }
    
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(102, 126, 234, 0.4);
    }
    
    /* Download Buttons */
    .stDownloadButton>button {
        background: linear-gradient(135deg, #4caf50 0%, #45a049 100%);
        color: white;
        border: none;
        padding: 12px 30px;
        border-radius: 10px;
        font-weight: 600;
        transition: all 0.3s ease;
    }
    
    .stDownloadButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 12px rgba(76, 175, 80, 0.4);
    }
    
    /* Sidebar Styling */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #ffffff 0%, #f8f9fa 100%);
        border-right: 2px solid #e9ecef;
    }
    
    /* Input Fields */
    .stNumberInput>div>div>input,
    .stTextInput>div>div>input,
    .stSelectbox>div>div {
        border-radius: 8px;
        border: 2px solid #e9ecef;
        padding: 10px;
        font-size: 15px;
        transition: border-color 0.3s ease;
    }
    
    .stNumberInput>div>div>input:focus,
    .stTextInput>div>div>input:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
    }
    
    /* Slider Styling */
    .stSlider>div>div>div>div {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background: #f8f9fa;
        border-radius: 10px;
        font-weight: 600;
        font-size: 16px;
        padding: 12px;
    }
    
    /* Scrollbar Styling */
    ::-webkit-scrollbar {
        width: 10px;
        height: 10px;
    }
    
    ::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(135deg, #764ba2 0%, #667eea 100%);
    }
</style>
""", unsafe_allow_html=True)

def create_detailed_section_diagram(d, bf, tf, tw, section_name):
    """Create a detailed I-beam cross-section diagram with dimensions"""
    fig, ax = plt.subplots(figsize=(6, 5))
    
    # Scale for visualization
    scale = 1.0
    d_s, bf_s, tf_s, tw_s = d*scale, bf*scale, tf*scale, tw*scale
    
    # Center
    cx, cy = 0, 0
    
    # Draw I-beam
    # Top flange
    ax.add_patch(patches.Rectangle(
        (cx - bf_s/2, cy + d_s/2 - tf_s), bf_s, tf_s,
        linewidth=2, edgecolor='#1a237e', facecolor='#e8eaf6'
    ))
    
    # Bottom flange
    ax.add_patch(patches.Rectangle(
        (cx - bf_s/2, cy - d_s/2), bf_s, tf_s,
        linewidth=2, edgecolor='#1a237e', facecolor='#e8eaf6'
    ))
    
    # Web
    ax.add_patch(patches.Rectangle(
        (cx - tw_s/2, cy - d_s/2 + tf_s), tw_s, d_s - 2*tf_s,
        linewidth=2, edgecolor='#1a237e', facecolor='#e8eaf6'
    ))
    
    # Dimension lines
    arrow_props = dict(arrowstyle='<->', color='#d32f2f', lw=1.5)
    
    # Overall height (d)
    offset_d = bf_s * 0.25
    ax.annotate('', xy=(bf_s/2 + offset_d, d_s/2),
                xytext=(bf_s/2 + offset_d, -d_s/2),
                arrowprops=arrow_props)
    ax.text(bf_s/2 + offset_d*1.8, 0, f'd = {d:.0f} mm',
            fontsize=10, color='#d32f2f', va='center', fontweight='bold')
    
    # Flange width (bf)
    offset_bf = d_s * 0.15
    ax.annotate('', xy=(-bf_s/2, d_s/2 + offset_bf),
                xytext=(bf_s/2, d_s/2 + offset_bf),
                arrowprops=arrow_props)
    ax.text(0, d_s/2 + offset_bf*2, f'bf = {bf:.0f} mm',
            fontsize=10, color='#d32f2f', ha='center', fontweight='bold')
    
    # Flange thickness (tf)
    ax.annotate('', xy=(-bf_s/2 - offset_d*0.5, d_s/2),
                xytext=(-bf_s/2 - offset_d*0.5, d_s/2 - tf_s),
                arrowprops=dict(arrowstyle='<->', color='#1976d2', lw=1.2))
    ax.text(-bf_s/2 - offset_d*1.5, d_s/2 - tf_s/2, f'tf = {tf:.1f}',
            fontsize=9, color='#1976d2', va='center', ha='right')
    
    # Web thickness (tw)
    ax.annotate('', xy=(-tw_s/2, -d_s/2 + tf_s + offset_bf*0.5),
                xytext=(tw_s/2, -d_s/2 + tf_s + offset_bf*0.5),
                arrowprops=dict(arrowstyle='<->', color='#388e3c', lw=1.2))
    ax.text(0, -d_s/2 + tf_s - offset_bf*0.8, f'tw = {tw:.1f}',
            fontsize=9, color='#388e3c', ha='center')
    
    # Styling
    ax.set_xlim(-bf_s * 1.0, bf_s * 1.3)
    ax.set_ylim(-d_s * 0.75, d_s * 0.85)
    ax.set_aspect('equal')
    ax.axis('off')
    ax.set_title(f'{section_name}\nCross-Section', fontsize=12, fontweight='bold', color='#1a237e')
    
    # Add axis labels
    ax.plot([cx, cx], [-d_s*0.6, d_s*0.6], 'k--', lw=0.5, alpha=0.5)
    ax.plot([-bf_s*0.6, bf_s*0.6], [cy, cy], 'k--', lw=0.5, alpha=0.5)
    ax.text(bf_s*0.65, 0, 'X', fontsize=8, va='center')
    ax.text(0, d_s*0.65, 'Y', fontsize=8, ha='center')
    
    plt.tight_layout()
    return fig


def create_flexural_capacity_chart(df, df_mat, section, material, Lb, Cb, flex_result):
    """Create flexural capacity curve for PDF report"""
    fig, ax = plt.subplots(figsize=(8, 5))
    
    # Generate capacity curve
    Lb_points = np.linspace(0.1, 15, 200)
    Mn_points = []
    
    for lb in Lb_points:
        r = aisc_360_16_f2_flexural_design(df, df_mat, section, material, lb, Cb)
        Mn_points.append(0.9 * r['Mn'] if r else 0)
    
    # Plot capacity curve
    ax.plot(Lb_points, Mn_points, 'b-', linewidth=2.5, label='φbMn Capacity')
    
    # Zone shading
    ax.axvspan(0, flex_result['Lp'], alpha=0.15, color='green', label='Yielding Zone (F2.1)')
    ax.axvspan(flex_result['Lp'], flex_result['Lr'], alpha=0.15, color='orange', label='Inelastic LTB (F2.2)')
    ax.axvspan(flex_result['Lr'], 15, alpha=0.15, color='red', label='Elastic LTB (F2.3)')
    
    # Vertical lines
    ax.axvline(x=flex_result['Lp'], color='green', linestyle='--', linewidth=1.5)
    ax.axvline(x=flex_result['Lr'], color='orange', linestyle='--', linewidth=1.5)
    
    # Add labels
    ax.text(flex_result['Lp'], ax.get_ylim()[1]*0.95, f'Lp={flex_result["Lp"]:.2f}m',
            fontsize=9, color='green', ha='center', fontweight='bold')
    ax.text(flex_result['Lr'], ax.get_ylim()[1]*0.95, f'Lr={flex_result["Lr"]:.2f}m',
            fontsize=9, color='orange', ha='center', fontweight='bold')
    
    # Design point
    ax.plot([Lb], [flex_result['phi_Mn']], 'r*', markersize=15, 
            label=f'Design Point (Lb={Lb:.2f}m)', zorder=5)
    
    # Styling
    ax.set_xlabel('Unbraced Length, Lb (m)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Design Moment, φbMn (t·m)', fontsize=11, fontweight='bold')
    ax.set_title(f'AISC F2: Flexural Capacity Curve - {section}', fontsize=12, fontweight='bold')
    ax.grid(True, alpha=0.3, linestyle=':')
    ax.legend(loc='upper right', fontsize=9, framealpha=0.95)
    ax.set_xlim(0, 15)
    ax.set_ylim(0, None)
    
    plt.tight_layout()
    return fig


def create_compression_capacity_chart(E, Fy, Ag, lambda_c, lambda_limit, comp_result, Pu):
    """Create compression capacity curve for PDF report"""
    fig, ax = plt.subplots(figsize=(8, 5))
    
    # Generate capacity curve
    lambda_points = np.linspace(1, 250, 250)
    Pn_points = []
    
    for lc in lambda_points:
        Fe = (math.pi**2 * E) / (lc**2)
        if lc <= lambda_limit:
            Fcr = (0.658**(Fy/Fe)) * Fy
        else:
            Fcr = 0.877 * Fe
        Pn_points.append(0.9 * Fcr * Ag / 1000.0)
    
    # Plot capacity curve
    ax.plot(lambda_points, Pn_points, 'b-', linewidth=2.5, label='φcPn Capacity')
    
    # Zone shading
    ax.axvspan(0, lambda_limit, alpha=0.15, color='green', label='Inelastic Buckling (E3-2)')
    ax.axvspan(lambda_limit, 250, alpha=0.15, color='orange', label='Elastic Buckling (E3-3)')
    
    # Vertical line at limit
    ax.axvline(x=lambda_limit, color='orange', linestyle='--', linewidth=1.5)
    ax.text(lambda_limit, ax.get_ylim()[1]*0.95 if ax.get_ylim()[1] > 0 else Pn_points[0]*0.95, 
            f'λ limit={lambda_limit:.1f}',
            fontsize=9, color='orange', ha='center', fontweight='bold')
    
    # Design point
    ax.plot([comp_result['lambda_c']], [comp_result['phi_Pn']], 'r*', markersize=15,
            label=f'Design Point (λ={comp_result["lambda_c"]:.1f})', zorder=5)
    
    # Demand line
    if Pu > 0:
        ax.axhline(y=Pu, color='red', linestyle='--', linewidth=1.5, label=f'Pu = {Pu:.1f} tons')
    
    # Styling
    ax.set_xlabel('Slenderness Ratio (KL/r)', fontsize=11, fontweight='bold')
    ax.set_ylabel('Design Strength, φcPn (tons)', fontsize=11, fontweight='bold')
    ax.set_title('AISC E3: Column Capacity Curve', fontsize=12, fontweight='bold')
    ax.grid(True, alpha=0.3, linestyle=':')
    ax.legend(loc='upper right', fontsize=9, framealpha=0.95)
    ax.set_xlim(0, 250)
    ax.set_ylim(0, None)
    
    plt.tight_layout()
    return fig

# ==================== DATA PATHS ====================
file_path = "https://raw.githubusercontent.com/Thana-site/Steel_Design_2003/main/2003-Steel-Beam-DataBase-H-Shape.csv"
file_path_mat = "https://raw.githubusercontent.com/Thana-site/Steel_Design_2003/main/2003-Steel-Beam-DataBase-Material.csv"

# ==================== HELPER FUNCTIONS ====================
@st.cache_data
def load_data():
    """Load steel section and material databases"""
    try:
        df = pd.read_csv(file_path, index_col=0, encoding='ISO-8859-1')
        df_mat = pd.read_csv(file_path_mat, index_col=0, encoding="utf-8")
        return df, df_mat, True
    except Exception as e:
        st.error(f"❌ Error loading data: {e}")
        return pd.DataFrame(), pd.DataFrame(), False

def safe_scalar(value):
    """Safely convert numpy array or other type to scalar float"""
    try:
        if hasattr(value, 'item'):
            return float(value.item())
        elif hasattr(value, '__iter__') and not isinstance(value, str):
            return float(value[0]) if len(value) > 0 else 0.0
        else:
            return float(value)
    except:
        return 0.0

def safe_sqrt(value):
    """Safe square root that ensures non-negative input"""
    val = safe_scalar(value)
    return math.sqrt(abs(val)) if val >= 0 else 0.0

# ==================== AISC 360-16 DESIGN FUNCTIONS ====================
def aisc_360_16_f2_flexural_design(df, df_mat, section, material, Lb_input, Cb=1.0):
    """AISC 360-16 F2 - Lateral-Torsional Buckling Analysis"""
    try:
        Sx = safe_scalar(df.loc[section, "Sx [cm3]"])
        Zx = safe_scalar(df.loc[section, 'Zx [cm3]'])
        ry = safe_scalar(df.loc[section, 'ry [cm]'])
        
        if 'rts [cm]' in df.columns:
            rts = safe_scalar(df.loc[section, 'rts [cm]'])
        else:
            rts = ry * 1.2
        
        J = safe_scalar(df.loc[section, 'j [cm4]']) if 'j [cm4]' in df.columns else 1.0
        
        if 'ho [mm]' in df.columns:
            ho = safe_scalar(df.loc[section, 'ho [mm]']) / 10.0
        else:
            ho = safe_scalar(df.loc[section, 'd [mm]']) / 10.0
        
        Fy = safe_scalar(df_mat.loc[material, "Yield Point (ksc)"])
        E = safe_scalar(df_mat.loc[material, "E"])
        
        Lb = safe_scalar(Lb_input)
        Cb = safe_scalar(Cb)
        
        # AISC 360-16 Equation F2.5
        Lp = 1.76 * ry * safe_sqrt(E / Fy) / 100.0
        
        # AISC 360-16 Equation F2.6
        c = 1.0
        term1 = 1.95 * rts * E / (0.7 * Fy)
        inner_sqrt = J * c / (Sx * ho)
        term2 = safe_sqrt(inner_sqrt)
        
        ratio_term = (0.7 * Fy / E) ** 2
        geometry_term = (Sx * ho / (J * c)) ** 2
        complex_inner = 1.0 + 6.76 * ratio_term * geometry_term
        term3 = safe_sqrt(1.0 + safe_sqrt(complex_inner))
        
        Lr = term1 * term2 * term3 / 100.0
        
        Mp = Fy * Zx
        
        if Lb <= Lp:
            Case = "F2.1 - Yielding"
            Mn = Mp
        elif Lb <= Lr:
            Case = "F2.2 - Inelastic LTB"
            Lb_cm = Lb * 100.0
            Lp_cm = Lp * 100.0  
            Lr_cm = Lr * 100.0
            
            Mp_minus_Mr = Mp - 0.7 * Fy * Sx
            length_ratio = (Lb_cm - Lp_cm) / (Lr_cm - Lp_cm)
            Mn = Cb * (Mp - Mp_minus_Mr * length_ratio)
            Mn = min(Mp, Mn)
        else:
            Case = "F2.3 - Elastic LTB"
            Lb_cm = Lb * 100.0
            Lb_rts_ratio = Lb_cm / rts
            
            term_1 = (Cb * math.pi**2 * E) / (Lb_rts_ratio**2)
            term_2 = 0.078 * (J * c / (Sx * ho)) * (Lb_rts_ratio**2)
            Fcr = term_1 * safe_sqrt(1.0 + term_2)
            
            Mn = Fcr * Sx
            Mn = min(Mp, Mn)
        
        Mn_tm = Mn / 100000.0
        Mp_tm = Mp / 100000.0
        
        return {
            'Mn': Mn_tm,
            'Mp': Mp_tm,
            'Lp': Lp,
            'Lr': Lr, 
            'Case': Case,
            'Cb': Cb
        }
        
    except Exception as e:
        st.error(f"Error in AISC 360-16 F2 calculation: {str(e)}")
        return None

def aisc_360_16_e3_compression_design(df, df_mat, section, material, KLx, KLy):
    """AISC 360-16 E3 - Flexural Buckling Analysis"""
    try:
        Fy = safe_scalar(df_mat.loc[material, "Yield Point (ksc)"])
        E = safe_scalar(df_mat.loc[material, "E"])
        
        Ag = safe_scalar(df.loc[section, 'A [cm2]'])
        rx = safe_scalar(df.loc[section, 'rx [cm]'])
        ry = safe_scalar(df.loc[section, 'ry [cm]'])
        
        KLx_scalar = safe_scalar(KLx) * 100.0
        KLy_scalar = safe_scalar(KLy) * 100.0
        
        lambda_x = KLx_scalar / rx
        lambda_y = KLy_scalar / ry
        lambda_c = max(lambda_x, lambda_y)
        
        Fe = (math.pi**2 * E) / (lambda_c**2)
        
        lambda_limit = 4.71 * safe_sqrt(E / Fy)
        
        if lambda_c <= lambda_limit:
            buckling_mode = "Inelastic"
            exponent = Fy / Fe
            Fcr = (0.658 ** exponent) * Fy
        else:
            buckling_mode = "Elastic"
            Fcr = 0.877 * Fe
        
        Pn = Fcr * Ag / 1000.0
        phi_c = 0.90
        phi_Pn = phi_c * Pn
        
        slenderness_ok = lambda_c <= 200.0
        
        return {
            'Pn': Pn,
            'phi_Pn': phi_Pn,
            'Fcr': Fcr,
            'Fe': Fe,
            'lambda_x': lambda_x,
            'lambda_y': lambda_y,
            'lambda_c': lambda_c,
            'lambda_limit': lambda_limit,
            'buckling_mode': buckling_mode,
            'slenderness_ok': slenderness_ok,
            'phi_c': phi_c
        }
        
    except Exception as e:
        st.error(f"Error in AISC 360-16 E3 compression analysis: {e}")
        return None

def aisc_360_16_h1_interaction(Pu, phi_Pn, Mux, phi_Mnx, Muy, phi_Mny):
    """AISC 360-16 H1 - Combined Forces Analysis"""
    try:
        Pu = safe_scalar(Pu)
        phi_Pn = safe_scalar(phi_Pn) 
        Mux = safe_scalar(Mux)
        phi_Mnx = safe_scalar(phi_Mnx)
        Muy = safe_scalar(Muy)
        phi_Mny = safe_scalar(phi_Mny)
        
        if phi_Pn <= 0.0 or phi_Mnx <= 0.0 or phi_Mny <= 0.0:
            return None
        
        Pr_Pc = Pu / phi_Pn
        Mrx_Mcx = Mux / phi_Mnx
        Mry_Mcy = Muy / phi_Mny
        
        if Pr_Pc >= 0.2:
            interaction_ratio = Pr_Pc + (8.0/9.0) * (Mrx_Mcx + Mry_Mcy)
            equation = "H1-1a"
        else:
            interaction_ratio = Pr_Pc/2.0 + (Mrx_Mcx + Mry_Mcy)
            equation = "H1-1b"
        
        design_ok = interaction_ratio <= 1.0
        safety_margin = (1.0 - interaction_ratio) if design_ok else None
        
        return {
            'interaction_ratio': interaction_ratio,
            'equation': equation,
            'design_ok': design_ok,
            'Pr_Pc': Pr_Pc,
            'Mrx_Mcx': Mrx_Mcx,
            'Mry_Mcy': Mry_Mcy,
            'safety_margin': safety_margin
        }
        
    except Exception as e:
        st.error(f"Error in AISC 360-16 H1 interaction calculation: {e}")
        return None

# ==================== ENHANCED PLOTLY CHART CONFIGURATIONS ====================
def create_enhanced_plotly_config():
    """Standard configuration for all Plotly charts with improved readability"""
    return {
        'displayModeBar': True,
        'displaylogo': False,
        'modeBarButtonsToRemove': ['pan2d', 'lasso2d', 'select2d'],
        'toImageButtonOptions': {
            'format': 'png',
            'filename': 'aisc_chart',
            'height': 800,
            'width': 1200,
            'scale': 2
        }
    }

def get_enhanced_plotly_layout():
    """Standard layout settings for enhanced readability"""
    return {
        'template': 'plotly_white',
        'font': {
            'family': 'Inter, sans-serif',
            'size': 14,
            'color': '#2c3e50'
        },
        'title': {
            'font': {'size': 20, 'color': '#2c3e50', 'family': 'Inter'},
            'x': 0.5,
            'xanchor': 'center'
        },
        'xaxis': {
            'title': {'font': {'size': 16, 'color': '#34495e'}},
            'tickfont': {'size': 13},
            'gridcolor': '#ecf0f1',
            'showgrid': True,
            'zeroline': True,
            'zerolinecolor': '#bdc3c7',
            'showline': True,
            'linecolor': '#bdc3c7'
        },
        'yaxis': {
            'title': {'font': {'size': 16, 'color': '#34495e'}},
            'tickfont': {'size': 13},
            'gridcolor': '#ecf0f1',
            'showgrid': True,
            'zeroline': True,
            'zerolinecolor': '#bdc3c7',
            'showline': True,
            'linecolor': '#bdc3c7'
        },
        'margin': {'l': 80, 'r': 80, 't': 100, 'b': 80},
        'hovermode': 'closest',
        'plot_bgcolor': 'white',
        'paper_bgcolor': 'white'
    }

def generate_pdf_report(df, df_mat, section, material, analysis_results, design_params):
    """Generate PDF report with perfect formatting - NO OVERLAP"""
    if not PDF_AVAILABLE:
        return None
    
    buffer = BytesIO()
    
    # Create document with proper margins
    doc = BaseDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=0.75*inch, 
        leftMargin=0.75*inch,
        topMargin=1.1*inch,  # Increased for header
        bottomMargin=0.9*inch,  # Increased for footer
        title="AISC 360-16 Steel Design Report"
    )
    
    # Define frame for content with proper margins
    frame = Frame(
        doc.leftMargin, 
        doc.bottomMargin, 
        doc.width, 
        doc.height - 0.3*inch,  # Account for header/footer space
        id='normal',
        topPadding=12,
        bottomPadding=12
    )
    template = PageTemplate(id='main', frames=frame, onPage=lambda c, d: None)
    doc.addPageTemplates([template])
    
    story = []
    styles = getSampleStyleSheet()
    
    # ==================== ENHANCED CUSTOM STYLES - NO OVERLAP ====================
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=rl_colors.HexColor('#667eea'),
        spaceAfter=18,
        spaceBefore=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        leading=24
    )
    
    heading1_style = ParagraphStyle(
        'CustomHeading1',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=rl_colors.HexColor('#2c3e50'),
        spaceAfter=10,
        spaceBefore=16,
        fontName='Helvetica-Bold',
        borderWidth=0,  # Remove border to prevent overlap
        backColor=rl_colors.HexColor('#f0f3ff'),
        borderPadding=6,
        leftIndent=0,
        rightIndent=0,
        keepWithNext=True,
        leading=18
    )
    
    heading2_style = ParagraphStyle(
        'CustomHeading2',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=rl_colors.HexColor('#34495e'),
        spaceAfter=8,
        spaceBefore=12,
        fontName='Helvetica-Bold',
        keepWithNext=True,
        leading=16
    )
    
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['Normal'],
        fontSize=9,
        leading=13,
        spaceAfter=4,
        spaceBefore=2,
        alignment=TA_LEFT,
        wordWrap='CJK'  # Better word wrapping
    )
    
    # Improved equation style - no overlap
    equation_style = ParagraphStyle(
        'EquationStyle',
        parent=styles['Code'],
        fontSize=9,
        textColor=rl_colors.HexColor('#1565C0'),
        backColor=rl_colors.HexColor('#E7F3FF'),
        borderWidth=1,
        borderColor=rl_colors.HexColor('#2196F3'),
        borderPadding=8,
        leftIndent=12,
        rightIndent=12,
        spaceAfter=8,
        spaceBefore=8,
        fontName='Courier',
        leading=12,
        borderRadius=5
    )
    
    # ==================== TITLE PAGE ====================
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph("AISC 360-16 STEEL DESIGN", title_style))
    story.append(Paragraph("COMPREHENSIVE CALCULATION REPORT", title_style))
    story.append(Spacer(1, 0.4*inch))
    
    # Report Info with proper table
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    header_data = [
        ['Report Generated:', timestamp],
        ['Section:', section],
        ['Material Grade:', material],
        ['Analysis Type:', 'Comprehensive Design Evaluation']
    ]
    
    header_table = Table(header_data, colWidths=[2.3*inch, 4.2*inch])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), rl_colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (0, -1), rl_colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, rl_colors.grey),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [rl_colors.HexColor('#f5f5f5'), rl_colors.white])
    ]))
    story.append(KeepTogether(header_table))
    story.append(Spacer(1, 0.25*inch))
    
    # Get material properties
    Fy = safe_scalar(df_mat.loc[material, "Yield Point (ksc)"])
    Fu = safe_scalar(df_mat.loc[material, "Tensile Strength (ksc)"])
    E = safe_scalar(df_mat.loc[material, "E"])
    
    # ==================== MATERIAL PROPERTIES ====================
    story.append(Paragraph("1. MATERIAL PROPERTIES", heading1_style))
    story.append(Spacer(1, 6))
    
    mat_data = [
        ['Property', 'Value', 'Unit', 'Description'],
        ['Fy', f'{Fy:.1f}', 'kgf/cm²', 'Yield Strength'],
        ['Fu', f'{Fu:.1f}', 'kgf/cm²', 'Tensile Strength'],
        ['E', f'{E:.0f}', 'kgf/cm²', 'Modulus of Elasticity']
    ]
    
    mat_table = Table(mat_data, colWidths=[1.3*inch, 1.2*inch, 1.3*inch, 2.7*inch])
    mat_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#f9f9f9')])
    ]))
    story.append(KeepTogether(mat_table))
    story.append(Spacer(1, 0.2*inch))
    
    # ==================== SECTION PROPERTIES ====================
    story.append(PageBreak())
    story.append(Paragraph("2. SECTION PROPERTIES", heading1_style))
    story.append(Spacer(1, 6))
    
    # Build properties table with proper column widths
    props_data = [['Property', 'Symbol', 'Value', 'Unit']]
    
    property_list = [
        ('d [mm]', 'd', 'mm'),
        ('bf [mm]', 'bf', 'mm'),
        ('tw [mm]', 'tw', 'mm'),
        ('tf [mm]', 'tf', 'mm'),
        ('A [cm2]', 'A', 'cm²'),
        ('Ix [cm4]', 'Ix', 'cm⁴'),
        ('Iy [cm4]', 'Iy', 'cm⁴'),
        ('rx [cm]', 'rx', 'cm'),
        ('ry [cm]', 'ry', 'cm'),
        ('Sx [cm3]', 'Sx', 'cm³'),
        ('Sy [cm3]', 'Sy', 'cm³'),
        ('Zx [cm3]', 'Zx', 'cm³'),
        ('Zy [cm3]', 'Zy', 'cm³'),
    ]
    
    weight_col = 'Unit Weight [kg/m]' if 'Unit Weight [kg/m]' in df.columns else 'w [kg/m]'
    
    for prop_key, symbol, unit in property_list:
        if prop_key in df.columns:
            value = safe_scalar(df.loc[section, prop_key])
            description = prop_key.split('[')[0].strip()
            formatted_value = f'{value:.2f}' if value < 1000 else f'{value:.1f}'
            props_data.append([description, symbol, formatted_value, unit])
    
    # Create table with word wrap
    props_table = Table(props_data, colWidths=[2.2*inch, 0.9*inch, 1.4*inch, 1*inch])
    props_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#f9f9f9')]),
        ('WORDWRAP', (0, 0), (-1, -1), True)
    ]))
    story.append(KeepTogether(props_table))
    story.append(Spacer(1, 0.15*inch))
    
    # Add slenderness ratios
    bf = safe_scalar(df.loc[section, 'bf [mm]'])
    tf = safe_scalar(df.loc[section, 'tf [mm]'])
    d = safe_scalar(df.loc[section, 'd [mm]'])
    tw = safe_scalar(df.loc[section, 'tw [mm]'])
    h = safe_scalar(df.loc[section, 'ho [mm]']) if 'ho [mm]' in df.columns else (d - 2*tf)
    
    slender_data = [
        ['Parameter', 'Value'],
        ['Flange Slenderness (bf/2tf)', f'{(bf/2.0)/tf:.2f}'],
        ['Web Slenderness (h/tw)', f'{h/tw:.2f}']
    ]
    
    slender_table = Table(slender_data, colWidths=[3.5*inch, 2*inch])
    slender_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#2196f3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
    ]))
    story.append(KeepTogether(slender_table))
    
    # ==================== SECTION CLASSIFICATION ====================
    story.append(PageBreak())
    story.append(Paragraph("3. SECTION CLASSIFICATION", heading1_style))
    story.append(Spacer(1, 8))
    
    flex_class = classify_section_flexure(df, df_mat, section, material)
    comp_class = classify_section_compression(df, df_mat, section, material)
    
    if flex_class:
        story.append(Paragraph("3.1 Flexural Classification (AISC Table B4.1b)", heading2_style))
        story.append(Spacer(1, 6))
        
        # Flange - Use equation box for formulas
        story.append(Paragraph("<b>Flange (Case 10: I-shaped sections)</b>", body_style))
        story.append(Spacer(1, 4))
        
        # Use custom equation box instead of paragraph
        eq_text = f"λp = 0.38√(E/Fy) = 0.38√({E:.0f}/{Fy:.1f}) = {flex_class['flange_lambda_p']:.2f}"
        story.append(EquationBox(eq_text, doc.width))
        story.append(Spacer(1, 4))
        
        eq_text = f"λr = 1.0√(E/Fy) = 1.0√({E:.0f}/{Fy:.1f}) = {flex_class['flange_lambda_r']:.2f}"
        story.append(EquationBox(eq_text, doc.width))
        story.append(Spacer(1, 6))
        
        story.append(Paragraph(
            f"<b>Actual λ = (bf/2)/tf = {flex_class['flange_lambda']:.2f}</b>",
            body_style
        ))
        story.append(Paragraph(
            f"<b>Classification: {flex_class['flange_class']}</b>",
            body_style
        ))
        story.append(Spacer(1, 12))
        
        # Web
        story.append(Paragraph("<b>Web (Case 15: Webs in flexure)</b>", body_style))
        story.append(Spacer(1, 4))
        
        eq_text = f"λpw = 3.76√(E/Fy) = {flex_class['web_lambda_p']:.2f}"
        story.append(EquationBox(eq_text, doc.width))
        story.append(Spacer(1, 4))
        
        eq_text = f"λrw = 5.70√(E/Fy) = {flex_class['web_lambda_r']:.2f}"
        story.append(EquationBox(eq_text, doc.width))
        story.append(Spacer(1, 6))
        
        story.append(Paragraph(
            f"<b>Actual λ = h/tw = {flex_class['web_lambda']:.2f}</b>",
            body_style
        ))
        story.append(Paragraph(
            f"<b>Classification: {flex_class['web_class']}</b>",
            body_style
        ))
        story.append(Spacer(1, 0.15*inch))
    
    if comp_class:
        story.append(Paragraph("3.2 Compression Classification (AISC Table B4.1a)", heading2_style))
        story.append(Spacer(1, 6))
        
        comp_summary = [
            ['Element', 'λ', 'λr', 'Status'],
            ['Flange', f'{comp_class["flange_lambda"]:.2f}', 
             f'{comp_class["flange_lambda_r"]:.2f}', 
             'Slender' if comp_class['flange_slender'] else 'Non-slender'],
            ['Web', f'{comp_class["web_lambda"]:.2f}',
             f'{comp_class["web_lambda_r"]:.2f}',
             'Slender' if comp_class['web_slender'] else 'Non-slender'],
        ]
        
        comp_table = Table(comp_summary, colWidths=[1.5*inch, 1.3*inch, 1.3*inch, 1.4*inch])
        comp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ]))
        story.append(KeepTogether(comp_table))
        
        story.append(Spacer(1, 8))
        story.append(Paragraph(
            f"<b>Overall: {comp_class['overall_class']}</b>",
            body_style
        ))
    
    # ==================== DESIGN ANALYSIS ====================
    if analysis_results:
        story.append(PageBreak())
        story.append(Paragraph("4. DESIGN ANALYSIS", heading1_style))
        story.append(Spacer(1, 10))
        
        # FLEXURAL ANALYSIS
        if 'flexural' in analysis_results:
            story.append(Paragraph("4.1 Flexural Design (AISC Chapter F2)", heading2_style))
            story.append(Spacer(1, 6))
            
            flex = analysis_results['flexural']
            Lb = design_params.get('Lb', 0)
            Cb = design_params.get('Cb', 1.0)
            
            Sx = safe_scalar(df.loc[section, "Sx [cm3]"])
            Zx = safe_scalar(df.loc[section, 'Zx [cm3]'])
            ry = safe_scalar(df.loc[section, 'ry [cm]'])
            
            # Step 1
            story.append(Paragraph("<b>Step 1: Plastic Moment</b>", body_style))
            story.append(Spacer(1, 4))
            
            eq_text = f"Mp = Fy × Zx = {Fy:.1f} × {Zx:.2f} = {flex['Mp']:.2f} t·m"
            story.append(EquationBox(eq_text, doc.width))
            story.append(Spacer(1, 10))
            
            # Step 2
            story.append(Paragraph("<b>Step 2: Limiting Lengths</b>", body_style))
            story.append(Spacer(1, 4))
            
            eq_text = f"Lp = 1.76 × ry × √(E/Fy) = 1.76 × {ry:.2f} × √({E:.0f}/{Fy:.1f}) = {flex['Lp']:.3f} m"
            story.append(EquationBox(eq_text, doc.width))
            story.append(Spacer(1, 4))
            
            eq_text = f"Lr = {flex['Lr']:.3f} m (AISC Eq. F2-6)"
            story.append(EquationBox(eq_text, doc.width))
            story.append(Spacer(1, 10))
            
            # Step 3
            story.append(Paragraph(f"<b>Step 3: Nominal Moment (Lb = {Lb:.2f} m)</b>", body_style))
            story.append(Spacer(1, 4))
            
            if Lb <= flex['Lp']:
                story.append(Paragraph(
                    f"Lb ({Lb:.2f}m) ≤ Lp ({flex['Lp']:.3f}m) → <b>Yielding (F2.1)</b>",
                    body_style
                ))
                eq_text = f"Mn = Mp = {flex['Mn']:.2f} t·m"
            elif Lb <= flex['Lr']:
                story.append(Paragraph(
                    f"Lp < Lb ({Lb:.2f}m) ≤ Lr → <b>Inelastic LTB (F2.2)</b>",
                    body_style
                ))
                eq_text = f"Mn = Cb[Mp - ...] = {flex['Mn']:.2f} t·m"
            else:
                story.append(Paragraph(
                    f"Lb ({Lb:.2f}m) > Lr → <b>Elastic LTB (F2.3)</b>",
                    body_style
                ))
                eq_text = f"Mn = Fcr × Sx = {flex['Mn']:.2f} t·m"
            
            story.append(Spacer(1, 4))
            story.append(EquationBox(eq_text, doc.width))
            story.append(Spacer(1, 10))
            
            # Step 4
            story.append(Paragraph("<b>Step 4: Design Strength</b>", body_style))
            story.append(Spacer(1, 4))
            
            eq_text = f"φMn = 0.90 × {flex['Mn']:.2f} = {flex['phi_Mn']:.2f} t·m"
            story.append(EquationBox(eq_text, doc.width))
            story.append(Spacer(1, 12))
            
            # Summary table
            flex_summary = [
                ['Parameter', 'Value'],
                ['Design Capacity (φMn)', f"{flex['phi_Mn']:.2f} t·m"],
                ['Nominal Moment (Mn)', f"{flex['Mn']:.2f} t·m"],
                ['Plastic Moment (Mp)', f"{flex['Mp']:.2f} t·m"],
                ['Design Case', flex['case']],
                ['Design Zone', flex['zone']],
                ['Utilization', f"{flex['ratio']:.3f}"],
                ['Status', '✓ ADEQUATE' if flex['adequate'] else '✗ INADEQUATE']
            ]
            
            flex_table = Table(flex_summary, colWidths=[3*inch, 2.5*inch])
            flex_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#4caf50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
            ]))
            story.append(KeepTogether(flex_table))
            
            # Add chart
            story.append(Spacer(1, 0.15*inch))
            story.append(Paragraph("<b>Flexural Capacity Curve:</b>", body_style))
            story.append(Spacer(1, 6))
            
            # Generate chart with better formatting
            fig, ax = plt.subplots(figsize=(6.5, 4))
            Lb_points = np.linspace(0.1, 15, 200)
            Mn_points = []
            
            for lb in Lb_points:
                r = aisc_360_16_f2_flexural_design(df, df_mat, section, material, lb, Cb)
                Mn_points.append(0.9 * r['Mn'] if r else 0)
            
            ax.plot(Lb_points, Mn_points, 'b-', linewidth=2.5, label='φM$_n$')
            ax.axvline(x=flex['Lp'], color='g', linestyle='--', linewidth=1.5, 
                      label=f'L$_p$ = {flex["Lp"]:.2f}m')
            ax.axvline(x=flex['Lr'], color='orange', linestyle='--', linewidth=1.5,
                      label=f'L$_r$ = {flex["Lr"]:.2f}m')
            ax.plot([Lb], [flex['phi_Mn']], 'r*', markersize=15, 
                   label='Design Point', zorder=5)
            
            ax.set_xlabel('Unbraced Length, L$_b$ (m)', fontsize=10, fontweight='bold')
            ax.set_ylabel('Design Moment, φM$_n$ (t·m)', fontsize=10, fontweight='bold')
            ax.set_title(f'Flexural Capacity - {section}', fontsize=11, fontweight='bold')
            ax.grid(True, alpha=0.3, linestyle=':', linewidth=0.7)
            ax.legend(loc='best', framealpha=0.9, fontsize=8)
            ax.tick_params(labelsize=8)
            
            # Save with proper margins
            img_buffer = BytesIO()
            plt.tight_layout(pad=0.5)
            plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight')
            plt.close()
            img_buffer.seek(0)
            
            img = Image(img_buffer, width=5.5*inch, height=3.4*inch)
            story.append(img)
        
        # COMPRESSION ANALYSIS
        if 'compression' in analysis_results:
            story.append(PageBreak())
            story.append(Paragraph("4.2 Compression Design (AISC Chapter E3)", heading2_style))
            story.append(Spacer(1, 6))
            
            comp = analysis_results['compression']
            KL = design_params.get('KL', 0)
            
            Ag = safe_scalar(df.loc[section, 'A [cm2]'])
            rx = safe_scalar(df.loc[section, 'rx [cm]'])
            ry = safe_scalar(df.loc[section, 'ry [cm]'])
            
            # Step 1
            story.append(Paragraph("<b>Step 1: Slenderness Ratio</b>", body_style))
            story.append(Spacer(1, 4))
            
            KL_cm = KL * 100
            lambda_x = KL_cm / rx
            lambda_y = KL_cm / ry
            lambda_c = max(lambda_x, lambda_y)
            
            eq_text = f"λc = max(KL/rx, KL/ry) = max({lambda_x:.1f}, {lambda_y:.1f}) = {lambda_c:.1f}"
            story.append(EquationBox(eq_text, doc.width))
            story.append(Spacer(1, 10))
            
            # Step 2
            story.append(Paragraph("<b>Step 2: Elastic Buckling Stress</b>", body_style))
            story.append(Spacer(1, 4))
            
            Fe = (math.pi**2 * E) / (lambda_c**2)
            eq_text = f"Fe = π²E/(λc)² = π² × {E:.0f} / {lambda_c:.1f}² = {Fe:.1f} kgf/cm²"
            story.append(EquationBox(eq_text, doc.width))
            story.append(Spacer(1, 10))
            
            # Step 3
            story.append(Paragraph("<b>Step 3: Critical Stress</b>", body_style))
            story.append(Spacer(1, 4))
            
            lambda_limit = 4.71 * safe_sqrt(E / Fy)
            
            if lambda_c <= lambda_limit:
                story.append(Paragraph(
                    f"λc ({lambda_c:.1f}) ≤ {lambda_limit:.1f} → <b>Inelastic (E3-2)</b>",
                    body_style
                ))
                exponent = Fy / Fe
                Fcr = (0.658 ** exponent) * Fy
                eq_text = f"Fcr = [0.658^(Fy/Fe)] × Fy = {Fcr:.1f} kgf/cm²"
            else:
                story.append(Paragraph(
                    f"λc ({lambda_c:.1f}) > {lambda_limit:.1f} → <b>Elastic (E3-3)</b>",
                    body_style
                ))
                Fcr = 0.877 * Fe
                eq_text = f"Fcr = 0.877 × Fe = {Fcr:.1f} kgf/cm²"
            
            story.append(Spacer(1, 4))
            story.append(EquationBox(eq_text, doc.width))
            story.append(Spacer(1, 10))
            
            # Step 4
            story.append(Paragraph("<b>Step 4: Design Strength</b>", body_style))
            story.append(Spacer(1, 4))
            
            Pn = Fcr * Ag / 1000
            eq_text = f"Pn = Fcr × Ag = {Fcr:.1f} × {Ag:.2f} / 1000 = {Pn:.2f} tons"
            story.append(EquationBox(eq_text, doc.width))
            story.append(Spacer(1, 4))
            
            eq_text = f"φPn = 0.90 × {Pn:.2f} = {comp['phi_Pn']:.2f} tons"
            story.append(EquationBox(eq_text, doc.width))
            story.append(Spacer(1, 12))
            
            # Summary table
            comp_summary = [
                ['Parameter', 'Value'],
                ['Design Strength (φPn)', f"{comp['phi_Pn']:.2f} tons"],
                ['Nominal Strength (Pn)', f"{comp['Pn']:.2f} tons"],
                ['Critical Stress (Fcr)', f"{comp['Fcr']:.1f} kgf/cm²"],
                ['Slenderness (λc)', f"{comp['lambda_c']:.1f}"],
                ['Buckling Mode', comp['mode']],
                ['Utilization', f"{comp['ratio']:.3f}"],
                ['Status', '✓ ADEQUATE' if comp['adequate'] else '✗ INADEQUATE']
            ]
            
            comp_table = Table(comp_summary, colWidths=[3*inch, 2.5*inch])
            comp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#2196f3')),
                ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.whitesmoke),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.grey),
            ]))
            story.append(KeepTogether(comp_table))
            
            # Add chart
            story.append(Spacer(1, 0.15*inch))
            story.append(Paragraph("<b>Column Capacity Curve:</b>", body_style))
            story.append(Spacer(1, 6))
            
            fig, ax = plt.subplots(figsize=(6.5, 4))
            lambda_points = np.linspace(1, 250, 250)
            Pn_points = []
            
            for lc in lambda_points:
                Fe_temp = (math.pi**2 * E) / (lc**2)
                if lc <= lambda_limit:
                    Fcr_temp = (0.658**(Fy/Fe_temp)) * Fy
                else:
                    Fcr_temp = 0.877 * Fe_temp
                Pn_points.append(0.9 * Fcr_temp * Ag / 1000.0)
            
            ax.plot(lambda_points, Pn_points, 'b-', linewidth=2.5, label='φP$_n$')
            ax.axvline(x=lambda_limit, color='orange', linestyle='--', linewidth=1.5,
                      label=f'λ limit = {lambda_limit:.1f}')
            ax.plot([comp['lambda_c']], [comp['phi_Pn']], 'r*', markersize=15,
                   label='Design Point', zorder=5)
            
            if 'Pu' in design_params and design_params['Pu'] > 0:
                ax.axhline(y=design_params['Pu'], color='g', linestyle='--',
                          linewidth=1.5, label=f'P$_u$ = {design_params["Pu"]:.1f} tons')
            
            ax.set_xlabel('Slenderness Ratio (KL/r)', fontsize=10, fontweight='bold')
            ax.set_ylabel('Design Strength, φP$_n$ (tons)', fontsize=10, fontweight='bold')
            ax.set_title(f'Column Capacity - {section}', fontsize=11, fontweight='bold')
            ax.grid(True, alpha=0.3, linestyle=':', linewidth=0.7)
            ax.legend(loc='best', framealpha=0.9, fontsize=8)
            ax.tick_params(labelsize=8)
            
            img_buffer = BytesIO()
            plt.tight_layout(pad=0.5)
            plt.savefig(img_buffer, format='png', dpi=120, bbox_inches='tight')
            plt.close()
            img_buffer.seek(0)
            
            img = Image(img_buffer, width=5.5*inch, height=3.4*inch)
            story.append(img)
    
    # Footer
    story.append(PageBreak())
    story.append(Spacer(1, 2*inch))
    footer_text = """
    <para align=center>
    <b>AISC 360-16 Steel Design v7.0</b><br/>
    All calculations comply with AISC 360-16 specifications.<br/>
    <br/>
    © 2024 - Structural Engineering Tool
    </para>
    """
    story.append(Paragraph(footer_text, body_style))
    
    # Build PDF with custom canvas
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer


# ==================== ENHANCED EXCEL EXPORT WITH DETAILED CALCULATIONS ====================
def generate_enhanced_excel_report(df, df_mat, section, material, analysis_results, design_params):
    """Generate comprehensive Excel calculation report with detailed AISC equations and charts"""
    if not EXCEL_AVAILABLE:
        return None
    
    buffer = BytesIO()
    wb = Workbook()
    
    # ==================== DEFINE ENHANCED STYLES ====================
    # Title styles
    title_font = Font(bold=True, size=20, color="667EEA")
    title_fill = PatternFill(start_color="F0F3FF", end_color="F0F3FF", fill_type="solid")
    
    # Header styles
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Subheader styles
    subheader_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Section header styles
    section_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    section_font = Font(bold=True, size=14, color="2c3e50")
    
    # Equation styles
    equation_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    equation_font = Font(italic=True, size=10, color="1565C0")
    
    # Calculation styles
    calc_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Result styles
    result_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    result_font = Font(bold=True, color="2E7D32", size=11)
    
    # Status styles
    adequate_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    adequate_font = Font(bold=True, color="2E7D32")
    inadequate_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    inadequate_font = Font(bold=True, color="C62828")
    
    # Classification styles
    compact_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    compact_font = Font(bold=True, color="2E7D32")
    noncompact_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    noncompact_font = Font(bold=True, color="F57C00")
    slender_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    slender_font = Font(bold=True, color="C62828")
    
    # Alignment
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    # ==================== SHEET 1: COVER & SUMMARY ====================
    ws_cover = wb.active
    ws_cover.title = "Cover & Summary"
    
    # Title
    ws_cover['A1'] = "AISC 360-16 STEEL DESIGN"
    ws_cover['A1'].font = title_font
    ws_cover['A1'].fill = title_fill
    ws_cover['A1'].alignment = center_align
    ws_cover.merge_cells('A1:F1')
    ws_cover.row_dimensions[1].height = 30
    
    ws_cover['A2'] = "COMPREHENSIVE CALCULATION REPORT"
    ws_cover['A2'].font = Font(bold=True, size=16, color="764ba2")
    ws_cover['A2'].fill = title_fill
    ws_cover['A2'].alignment = center_align
    ws_cover.merge_cells('A2:F2')
    ws_cover.row_dimensions[2].height = 25
    
    # Report information
    row = 4
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    info_data = [
        ['Report Generated:', timestamp],
        ['Section:', section],
        ['Material Grade:', material],
        ['Analysis Type:', 'Comprehensive Design Evaluation'],
    ]
    
    for label, value in info_data:
        ws_cover[f'A{row}'] = label
        ws_cover[f'A{row}'].font = Font(bold=True, size=11)
        ws_cover[f'A{row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        ws_cover[f'A{row}'].border = thin_border
        ws_cover[f'B{row}'] = value
        ws_cover[f'B{row}'].border = thin_border
        ws_cover.merge_cells(f'B{row}:D{row}')
        row += 1
    
    # Material Properties Summary
    row += 2
    ws_cover[f'A{row}'] = "MATERIAL PROPERTIES"
    ws_cover[f'A{row}'].font = section_font
    ws_cover[f'A{row}'].fill = section_fill
    ws_cover[f'A{row}'].alignment = center_align
    ws_cover.merge_cells(f'A{row}:D{row}')
    
    row += 1
    mat_headers = ['Property', 'Value', 'Unit', 'Description']
    for col, header in enumerate(mat_headers, start=1):
        cell = ws_cover.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    Fy = safe_scalar(df_mat.loc[material, "Yield Point (ksc)"])
    Fu = safe_scalar(df_mat.loc[material, "Tensile Strength (ksc)"])
    E = safe_scalar(df_mat.loc[material, "E"])
    
    mat_data = [
        ['Fy', f'{Fy:.1f}', 'kgf/cm²', 'Yield Strength'],
        ['Fu', f'{Fu:.1f}', 'kgf/cm²', 'Tensile Strength'],
        ['E', f'{E:.0f}', 'kgf/cm²', 'Modulus of Elasticity']
    ]
    
    for mat_row in mat_data:
        row += 1
        for col, value in enumerate(mat_row, start=1):
            cell = ws_cover.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_align if col > 1 else left_align
    
    # Design Parameters
    if design_params:
        row += 2
        ws_cover[f'A{row}'] = "DESIGN PARAMETERS"
        ws_cover[f'A{row}'].font = section_font
        ws_cover[f'A{row}'].fill = section_fill
        ws_cover[f'A{row}'].alignment = center_align
        ws_cover.merge_cells(f'A{row}:C{row}')
        
        row += 1
        param_headers = ['Parameter', 'Value', 'Unit']
        for col, header in enumerate(param_headers, start=1):
            cell = ws_cover.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        param_data = []
        if 'Mu' in design_params:
            param_data.append(['Applied Moment (Mu)', f"{design_params['Mu']:.2f}", 't·m'])
        if 'Pu' in design_params:
            param_data.append(['Applied Axial Load (Pu)', f"{design_params['Pu']:.2f}", 'tons'])
        if 'Lb' in design_params:
            param_data.append(['Unbraced Length (Lb)', f"{design_params['Lb']:.2f}", 'm'])
        if 'KL' in design_params:
            param_data.append(['Effective Length (KL)', f"{design_params['KL']:.2f}", 'm'])
        if 'Cb' in design_params:
            param_data.append(['Moment Gradient Factor (Cb)', f"{design_params['Cb']:.2f}", ''])
        
        for data_row in param_data:
            row += 1
            for col, value in enumerate(data_row, start=1):
                cell = ws_cover.cell(row=row, column=col)
                cell.value = value
                cell.border = thin_border
                cell.alignment = center_align if col > 1 else left_align
    
    # Overall Status
    if analysis_results:
        row += 2
        ws_cover[f'A{row}'] = "DESIGN STATUS"
        ws_cover[f'A{row}'].font = section_font
        ws_cover[f'A{row}'].fill = section_fill
        ws_cover[f'A{row}'].alignment = center_align
        ws_cover.merge_cells(f'A{row}:D{row}')
        
        row += 1
        overall_adequate = True
        status_data = []
        
        if 'flexural' in analysis_results:
            flex = analysis_results['flexural']
            status_data.append(['Flexural Check', f"{flex['ratio']:.3f}", 
                              '✓ PASS' if flex['adequate'] else '✗ FAIL'])
            overall_adequate = overall_adequate and flex['adequate']
        
        if 'compression' in analysis_results:
            comp = analysis_results['compression']
            status_data.append(['Compression Check', f"{comp['ratio']:.3f}",
                              '✓ PASS' if comp['adequate'] else '✗ FAIL'])
            overall_adequate = overall_adequate and comp['adequate']
        
        if 'interaction' in analysis_results:
            inter = analysis_results['interaction']
            status_data.append(['Interaction Check', f"{inter['interaction_ratio']:.3f}",
                              '✓ PASS' if inter['design_ok'] else '✗ FAIL'])
            overall_adequate = overall_adequate and inter['design_ok']
        
        for check_name, ratio, status in status_data:
            row += 1
            ws_cover[f'A{row}'] = check_name
            ws_cover[f'A{row}'].border = thin_border
            ws_cover[f'A{row}'].alignment = left_align
            
            ws_cover[f'B{row}'] = ratio
            ws_cover[f'B{row}'].border = thin_border
            ws_cover[f'B{row}'].alignment = center_align
            
            ws_cover[f'C{row}'] = status
            ws_cover[f'C{row}'].border = thin_border
            ws_cover[f'C{row}'].alignment = center_align
            if '✓' in status:
                ws_cover[f'C{row}'].fill = adequate_fill
                ws_cover[f'C{row}'].font = adequate_font
            else:
                ws_cover[f'C{row}'].fill = inadequate_fill
                ws_cover[f'C{row}'].font = inadequate_font
        
        row += 1
        ws_cover[f'A{row}'] = "OVERALL STATUS"
        ws_cover[f'A{row}'].font = Font(bold=True, size=12)
        ws_cover[f'A{row}'].border = thick_border
        ws_cover[f'A{row}'].alignment = center_align
        ws_cover.merge_cells(f'A{row}:B{row}')
        
        ws_cover[f'C{row}'] = '✓ ADEQUATE' if overall_adequate else '✗ INADEQUATE'
        ws_cover[f'C{row}'].border = thick_border
        ws_cover[f'C{row}'].alignment = center_align
        ws_cover[f'C{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        if overall_adequate:
            ws_cover[f'C{row}'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        else:
            ws_cover[f'C{row}'].fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
    
    # Auto-size columns
    for col in ['A', 'B', 'C', 'D']:
        ws_cover.column_dimensions[col].width = 25
    
    # ==================== SHEET 2: COMPLETE SECTION PROPERTIES ====================
    ws_props = wb.create_sheet("Section Properties")
    
    row = 1
    ws_props['A1'] = "COMPLETE SECTION PROPERTIES"
    ws_props['A1'].font = title_font
    ws_props['A1'].fill = title_fill
    ws_props['A1'].alignment = center_align
    ws_props.merge_cells('A1:E1')
    ws_props.row_dimensions[1].height = 25
    
    row = 3
    ws_props[f'A{row}'] = "Geometric Properties"
    ws_props[f'A{row}'].font = section_font
    ws_props[f'A{row}'].fill = section_fill
    ws_props.merge_cells(f'A{row}:E{row}')
    
    row += 1
    prop_headers = ['Property Description', 'Symbol', 'Value', 'Unit', 'AISC Reference']
    for col, header in enumerate(prop_headers, start=1):
        cell = ws_props.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Comprehensive property list with AISC references
    property_definitions = [
        ('Unit Weight [kg/m]', 'w', 'Weight per unit length', 'Table 1-1'),
        ('d [mm]', 'd', 'Overall depth', 'Table 1-1'),
        ('bf [mm]', 'bf', 'Flange width', 'Table 1-1'),
        ('tw [mm]', 'tw', 'Web thickness', 'Table 1-1'),
        ('tf [mm]', 'tf', 'Flange thickness', 'Table 1-1'),
        ('r [mm]', 'r', 'Fillet radius', 'Table 1-1'),
        ('A [cm2]', 'A', 'Cross-sectional area', 'Table 1-1'),
        ('Ix [cm4]', 'Ix', 'Moment of inertia, X-axis', 'Table 1-1'),
        ('Iy [cm4]', 'Iy', 'Moment of inertia, Y-axis', 'Table 1-1'),
        ('rx [cm]', 'rx', 'Radius of gyration, X-axis', 'Table 1-1'),
        ('ry [cm]', 'ry', 'Radius of gyration, Y-axis', 'Table 1-1'),
        ('Sx [cm3]', 'Sx', 'Elastic section modulus, X-axis', 'Table 1-1'),
        ('Sy [cm3]', 'Sy', 'Elastic section modulus, Y-axis', 'Table 1-1'),
        ('Zx [cm3]', 'Zx', 'Plastic section modulus, X-axis', 'Table 1-1'),
        ('Zy [cm3]', 'Zy', 'Plastic section modulus, Y-axis', 'Table 1-1'),
        ('ho [mm]', 'ho', 'Distance between flange centroids', 'Table 1-1'),
        ('j [cm4]', 'J', 'Torsional constant', 'Table 1-1'),
        ('cw [10^6 cm6]', 'Cw', 'Warping constant', 'Table 1-1'),
        ('rts [cm6]', 'rts', 'Effective radius for LTB', 'Eq. F2-7'),
    ]
    
    weight_col = 'Unit Weight [kg/m]' if 'Unit Weight [kg/m]' in df.columns else 'w [kg/m]'
    
    for prop_key, symbol, description, reference in property_definitions:
        # Handle weight column name variation
        check_key = weight_col if prop_key == 'Unit Weight [kg/m]' else prop_key
        
        if check_key in df.columns:
            row += 1
            value = safe_scalar(df.loc[section, check_key])
            unit = check_key.split('[')[1].replace(']', '') if '[' in check_key else ''
            
            ws_props[f'A{row}'] = description
            ws_props[f'B{row}'] = symbol
            ws_props[f'C{row}'] = f'{value:.3f}' if value < 100 else f'{value:.2f}'
            ws_props[f'D{row}'] = unit
            ws_props[f'E{row}'] = reference
            
            for col in range(1, 6):
                cell = ws_props.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_align if col > 1 else left_align
    
    # Add calculated slenderness ratios
    row += 2
    ws_props[f'A{row}'] = "Calculated Slenderness Ratios"
    ws_props[f'A{row}'].font = section_font
    ws_props[f'A{row}'].fill = section_fill
    ws_props.merge_cells(f'A{row}:E{row}')
    
    row += 1
    for col, header in enumerate(prop_headers, start=1):
        cell = ws_props.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    bf = safe_scalar(df.loc[section, 'bf [mm]'])
    tf = safe_scalar(df.loc[section, 'tf [mm]'])
    d = safe_scalar(df.loc[section, 'd [mm]'])
    tw = safe_scalar(df.loc[section, 'tw [mm]'])
    
    if 'ho [mm]' in df.columns:
        h = safe_scalar(df.loc[section, 'ho [mm]'])
    else:
        h = d - 2 * tf
    
    flange_slenderness = (bf / 2.0) / tf
    web_slenderness = h / tw
    
    slenderness_data = [
        ['Flange slenderness', 'bf/2tf', f'{flange_slenderness:.2f}', '', 'Table B4.1'],
        ['Web slenderness', 'h/tw', f'{web_slenderness:.2f}', '', 'Table B4.1']
    ]
    
    for data_row in slenderness_data:
        row += 1
        for col, value in enumerate(data_row, start=1):
            cell = ws_props.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_align if col > 1 else left_align
    
    # Auto-size columns
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_props.column_dimensions[col].width = 25
    
    # ==================== SHEET 3: SECTION CLASSIFICATION ====================
    ws_class = wb.create_sheet("Section Classification")
    
    row = 1
    ws_class['A1'] = "SECTION CLASSIFICATION PER AISC 360-16"
    ws_class['A1'].font = title_font
    ws_class['A1'].fill = title_fill
    ws_class['A1'].alignment = center_align
    ws_class.merge_cells('A1:F1')
    ws_class.row_dimensions[1].height = 25
    
    # Get classifications
    flex_class = classify_section_flexure(df, df_mat, section, material)
    comp_class = classify_section_compression(df, df_mat, section, material)
    
    if flex_class:
        row = 3
        ws_class[f'A{row}'] = "FLEXURAL MEMBER CLASSIFICATION (Table B4.1b)"
        ws_class[f'A{row}'].font = section_font
        ws_class[f'A{row}'].fill = section_fill
        ws_class.merge_cells(f'A{row}:F{row}')
        
        # Flange Classification
        row += 2
        ws_class[f'A{row}'] = "FLANGE CLASSIFICATION"
        ws_class[f'A{row}'].font = Font(bold=True, size=11, color="2c3e50")
        ws_class.merge_cells(f'A{row}:F{row}')
        
        row += 1
        ws_class[f'A{row}'] = "Case 10: Flanges of I-shaped sections in flexure"
        ws_class[f'A{row}'].font = Font(italic=True, size=10)
        ws_class.merge_cells(f'A{row}:F{row}')
        
        row += 2
        flange_data = [
            ['Parameter', 'Equation', 'Value', ''],
            ['λ (Actual)', f'(bf/2)/tf = ({bf:.1f}/2)/{tf:.1f}', f'{flex_class["flange_lambda"]:.2f}', ''],
            ['λp (Compact limit)', f'0.38√(E/Fy) = 0.38√({E:.0f}/{Fy:.1f})', f'{flex_class["flange_lambda_p"]:.2f}', 'Eq. B4-1a'],
            ['λr (Non-compact limit)', f'1.0√(E/Fy) = 1.0√({E:.0f}/{Fy:.1f})', f'{flex_class["flange_lambda_r"]:.2f}', 'Eq. B4-1b'],
        ]
        
        for data_row in flange_data:
            row += 1
            ws_class[f'A{row}'] = data_row[0]
            ws_class[f'A{row}'].font = Font(bold=True) if 'Parameter' in data_row[0] else Font()
            ws_class[f'A{row}'].border = thin_border
            ws_class[f'A{row}'].alignment = left_align
            
            ws_class.merge_cells(f'B{row}:C{row}')
            ws_class[f'B{row}'] = data_row[1]
            ws_class[f'B{row}'].border = thin_border
            ws_class[f'B{row}'].alignment = left_align
            if 'Eq.' in data_row[0] or 'λp' in data_row[0] or 'λr' in data_row[0]:
                ws_class[f'B{row}'].fill = equation_fill
            
            ws_class[f'D{row}'] = data_row[2]
            ws_class[f'D{row}'].border = thin_border
            ws_class[f'D{row}'].alignment = center_align
            
            ws_class[f'E{row}'] = data_row[3]
            ws_class[f'E{row}'].border = thin_border
            ws_class[f'E{row}'].alignment = center_align
            ws_class[f'E{row}'].font = Font(italic=True, size=9)
        
        row += 1
        ws_class[f'A{row}'] = "CLASSIFICATION:"
        ws_class[f'A{row}'].font = Font(bold=True, size=11)
        ws_class[f'A{row}'].border = thin_border
        
        ws_class.merge_cells(f'B{row}:D{row}')
        ws_class[f'B{row}'] = flex_class['flange_class'].upper()
        ws_class[f'B{row}'].border = thin_border
        ws_class[f'B{row}'].alignment = center_align
        
        if flex_class['flange_class'] == 'Compact':
            ws_class[f'B{row}'].fill = compact_fill
            ws_class[f'B{row}'].font = compact_font
        elif flex_class['flange_class'] == 'Non-compact':
            ws_class[f'B{row}'].fill = noncompact_fill
            ws_class[f'B{row}'].font = noncompact_font
        else:
            ws_class[f'B{row}'].fill = slender_fill
            ws_class[f'B{row}'].font = slender_font
        
        # Web Classification
        row += 3
        ws_class[f'A{row}'] = "WEB CLASSIFICATION"
        ws_class[f'A{row}'].font = Font(bold=True, size=11, color="2c3e50")
        ws_class.merge_cells(f'A{row}:F{row}')
        
        row += 1
        ws_class[f'A{row}'] = "Case 15: Webs of doubly symmetric I-shaped members in flexure"
        ws_class[f'A{row}'].font = Font(italic=True, size=10)
        ws_class.merge_cells(f'A{row}:F{row}')
        
        row += 2
        web_data = [
            ['Parameter', 'Equation', 'Value', ''],
            ['λ (Actual)', f'h/tw = {h:.1f}/{tw:.1f}', f'{flex_class["web_lambda"]:.2f}', ''],
            ['λpw (Compact limit)', f'3.76√(E/Fy) = 3.76√({E:.0f}/{Fy:.1f})', f'{flex_class["web_lambda_p"]:.2f}', 'Eq. B4-1a'],
            ['λrw (Non-compact limit)', f'5.70√(E/Fy) = 5.70√({E:.0f}/{Fy:.1f})', f'{flex_class["web_lambda_r"]:.2f}', 'Eq. B4-1b'],
        ]
        
        for data_row in web_data:
            row += 1
            ws_class[f'A{row}'] = data_row[0]
            ws_class[f'A{row}'].font = Font(bold=True) if 'Parameter' in data_row[0] else Font()
            ws_class[f'A{row}'].border = thin_border
            ws_class[f'A{row}'].alignment = left_align
            
            ws_class.merge_cells(f'B{row}:C{row}')
            ws_class[f'B{row}'] = data_row[1]
            ws_class[f'B{row}'].border = thin_border
            ws_class[f'B{row}'].alignment = left_align
            if 'Eq.' in data_row[0] or 'λp' in data_row[0] or 'λr' in data_row[0]:
                ws_class[f'B{row}'].fill = equation_fill
            
            ws_class[f'D{row}'] = data_row[2]
            ws_class[f'D{row}'].border = thin_border
            ws_class[f'D{row}'].alignment = center_align
            
            ws_class[f'E{row}'] = data_row[3]
            ws_class[f'E{row}'].border = thin_border
            ws_class[f'E{row}'].alignment = center_align
            ws_class[f'E{row}'].font = Font(italic=True, size=9)
        
        row += 1
        ws_class[f'A{row}'] = "CLASSIFICATION:"
        ws_class[f'A{row}'].font = Font(bold=True, size=11)
        ws_class[f'A{row}'].border = thin_border
        
        ws_class.merge_cells(f'B{row}:D{row}')
        ws_class[f'B{row}'] = flex_class['web_class'].upper()
        ws_class[f'B{row}'].border = thin_border
        ws_class[f'B{row}'].alignment = center_align
        
        if flex_class['web_class'] == 'Compact':
            ws_class[f'B{row}'].fill = compact_fill
            ws_class[f'B{row}'].font = compact_font
        elif flex_class['web_class'] == 'Non-compact':
            ws_class[f'B{row}'].fill = noncompact_fill
            ws_class[f'B{row}'].font = noncompact_font
        else:
            ws_class[f'B{row}'].fill = slender_fill
            ws_class[f'B{row}'].font = slender_font
    
    if comp_class:
        row += 4
        ws_class[f'A{row}'] = "COMPRESSION MEMBER CLASSIFICATION (Table B4.1a)"
        ws_class[f'A{row}'].font = section_font
        ws_class[f'A{row}'].fill = section_fill
        ws_class.merge_cells(f'A{row}:F{row}')
        
        row += 2
        comp_data = [
            ['Element', 'Case', 'λ (Actual)', 'λr (Limit)', 'Status'],
            ['Flange', 'Case 1: Flanges of I-shaped sections', f'{comp_class["flange_lambda"]:.2f}', 
             f'{comp_class["flange_lambda_r"]:.2f}', 'Slender' if comp_class['flange_slender'] else 'Non-slender'],
            ['Web', 'Case 5: Webs of doubly symmetric I-sections', f'{comp_class["web_lambda"]:.2f}',
             f'{comp_class["web_lambda_r"]:.2f}', 'Slender' if comp_class['web_slender'] else 'Non-slender'],
        ]
        
        for data_row in comp_data:
            row += 1
            for col, value in enumerate(data_row, start=1):
                cell = ws_class.cell(row=row, column=col)
                cell.value = value
                cell.border = thin_border
                cell.alignment = center_align if col > 2 else left_align
                
                if row > (len(comp_data) + row - len(comp_data)) and col == 1:
                    cell.font = Font(bold=True)
                elif 'Parameter' in str(value) or 'Element' in str(value):
                    cell.font = header_font
                    cell.fill = header_fill
                elif col == 5 and 'Slender' in str(value):
                    cell.fill = slender_fill
                    cell.font = slender_font
                elif col == 5 and 'Non-slender' in str(value):
                    cell.fill = compact_fill
                    cell.font = compact_font
        
        row += 1
        ws_class[f'A{row}'] = "OVERALL CLASSIFICATION:"
        ws_class[f'A{row}'].font = Font(bold=True, size=11)
        ws_class[f'A{row}'].border = thick_border
        ws_class.merge_cells(f'A{row}:C{row}')
        
        ws_class.merge_cells(f'D{row}:E{row}')
        ws_class[f'D{row}'] = comp_class['overall_class'].upper()
        ws_class[f'D{row}'].border = thick_border
        ws_class[f'D{row}'].alignment = center_align
        
        if comp_class['overall_class'] == 'Non-slender':
            ws_class[f'D{row}'].fill = compact_fill
            ws_class[f'D{row}'].font = compact_font
        else:
            ws_class[f'D{row}'].fill = slender_fill
            ws_class[f'D{row}'].font = slender_font
        
        if comp_class['overall_class'] == 'Slender':
            row += 1
            ws_class[f'A{row}'] = f"Limiting Element: {comp_class['limiting_element']}"
            ws_class[f'A{row}'].font = Font(italic=True, color="C62828")
            ws_class.merge_cells(f'A{row}:E{row}')
    
    # Auto-size columns
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_class.column_dimensions[col].width = 20
    ws_class.column_dimensions['B'].width = 35
    
    # ==================== SHEET 4: FLEXURAL ANALYSIS ====================
    if analysis_results and 'flexural' in analysis_results:
        ws_flex = wb.create_sheet("Flexural Analysis")
        
        row = 1
        ws_flex['A1'] = "AISC 360-16 CHAPTER F2: FLEXURAL DESIGN"
        ws_flex['A1'].font = title_font
        ws_flex['A1'].fill = title_fill
        ws_flex['A1'].alignment = center_align
        ws_flex.merge_cells('A1:E1')
        ws_flex.row_dimensions[1].height = 25
        
        flex = analysis_results['flexural']
        Lb = design_params.get('Lb', 0)
        Cb = design_params.get('Cb', 1.0)
        
        Sx = safe_scalar(df.loc[section, "Sx [cm3]"])
        Zx = safe_scalar(df.loc[section, 'Zx [cm3]'])
        ry = safe_scalar(df.loc[section, 'ry [cm]'])
        rts = safe_scalar(df.loc[section, 'rts [cm]']) if 'rts [cm]' in df.columns else ry * 1.2
        J = safe_scalar(df.loc[section, 'j [cm4]']) if 'j [cm4]' in df.columns else 1.0
        ho = safe_scalar(df.loc[section, 'ho [mm]']) / 10.0 if 'ho [mm]' in df.columns else d / 10.0
        
        # Step 1: Calculate Plastic Moment
        row = 3
        ws_flex[f'A{row}'] = "STEP 1: CALCULATE PLASTIC MOMENT Mp"
        ws_flex[f'A{row}'].font = section_font
        ws_flex[f'A{row}'].fill = section_fill
        ws_flex.merge_cells(f'A{row}:E{row}')
        
        row += 2
        ws_flex[f'A{row}'] = "AISC Equation F2-1:"
        ws_flex[f'A{row}'].font = Font(bold=True, italic=True)
        ws_flex[f'A{row}'].fill = equation_fill
        ws_flex.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws_flex[f'A{row}'] = "Mp = Fy × Zx"
        ws_flex.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws_flex[f'A{row}'] = f"Mp = {Fy:.1f} × {Zx:.2f}"
        ws_flex.merge_cells(f'A{row}:B{row}')
        ws_flex[f'C{row}'] = f"{flex['Mp']*100000:.0f} kgf·cm"
        ws_flex[f'D{row}'] = "="
        ws_flex[f'D{row}'].alignment = center_align
        ws_flex[f'E{row}'] = f"{flex['Mp']:.2f} t·m"
        ws_flex[f'E{row}'].fill = result_fill
        ws_flex[f'E{row}'].font = result_font
        
        # Step 2: Limiting Lengths
        row += 3
        ws_flex[f'A{row}'] = "STEP 2: CALCULATE LIMITING LATERALLY UNBRACED LENGTHS"
        ws_flex[f'A{row}'].font = section_font
        ws_flex[f'A{row}'].fill = section_fill
        ws_flex.merge_cells(f'A{row}:E{row}')
        
        row += 2
        ws_flex[f'A{row}'] = "Compact Limit Lp (AISC Equation F2-5):"
        ws_flex[f'A{row}'].font = Font(bold=True, italic=True)
        ws_flex[f'A{row}'].fill = equation_fill
        ws_flex.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws_flex[f'A{row}'] = "Lp = 1.76 × ry × √(E/Fy)"
        ws_flex.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws_flex[f'A{row}'] = f"Lp = 1.76 × {ry:.2f} × √({E:.0f}/{Fy:.1f})"
        ws_flex.merge_cells(f'A{row}:B{row}')
        ws_flex[f'C{row}'] = f"{flex['Lp']*100:.2f} cm"
        ws_flex[f'D{row}'] = "="
        ws_flex[f'D{row}'].alignment = center_align
        ws_flex[f'E{row}'] = f"{flex['Lp']:.3f} m"
        ws_flex[f'E{row}'].fill = result_fill
        ws_flex[f'E{row}'].font = result_font
        
        row += 2
        ws_flex[f'A{row}'] = "Inelastic Limit Lr (AISC Equation F2-6):"
        ws_flex[f'A{row}'].font = Font(bold=True, italic=True)
        ws_flex[f'A{row}'].fill = equation_fill
        ws_flex.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws_flex[f'A{row}'] = "Lr = 1.95 × rts × (E/(0.7Fy)) × √(Jc/(Sx×ho)) × √(1 + √(1 + 6.76×((0.7Fy)/E)²×((Sx×ho)/(Jc))²))"
        ws_flex[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws_flex.merge_cells(f'A{row}:E{row}')
        ws_flex.row_dimensions[row].height = 30
        
        row += 1
        ws_flex[f'A{row}'] = "Where:"
        row += 1
        ws_flex[f'A{row}'] = f"  rts = {rts:.2f} cm"
        row += 1
        ws_flex[f'A{row}'] = f"  J = {J:.2f} cm⁴"
        row += 1
        ws_flex[f'A{row}'] = f"  ho = {ho:.2f} cm"
        row += 1
        ws_flex[f'A{row}'] = f"  c = 1.0 (for doubly symmetric I-shapes)"
        
        row += 1
        ws_flex[f'A{row}'] = "Result:"
        ws_flex.merge_cells(f'A{row}:D{row}')
        ws_flex[f'E{row}'] = f"{flex['Lr']:.3f} m"
        ws_flex[f'E{row}'].fill = result_fill
        ws_flex[f'E{row}'].font = result_font
        
        # Step 3: Determine Nominal Moment
        row += 3
        ws_flex[f'A{row}'] = f"STEP 3: DETERMINE NOMINAL MOMENT Mn (Lb = {Lb:.2f} m)"
        ws_flex[f'A{row}'].font = section_font
        ws_flex[f'A{row}'].fill = section_fill
        ws_flex.merge_cells(f'A{row}:E{row}')
        
        row += 2
        if Lb <= flex['Lp']:
            ws_flex[f'A{row}'] = f"Lb ({Lb:.2f}m) ≤ Lp ({flex['Lp']:.3f}m)"
            ws_flex[f'A{row}'].font = Font(bold=True, size=11)
            ws_flex.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_flex[f'A{row}'] = "AISC Equation F2-1 applies (Yielding - Limit State of Lateral-Torsional Buckling does not apply)"
            ws_flex[f'A{row}'].fill = equation_fill
            ws_flex.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_flex[f'A{row}'] = "Mn = Mp"
            ws_flex.merge_cells(f'A{row}:D{row}')
            ws_flex[f'E{row}'] = f"{flex['Mn']:.2f} t·m"
            ws_flex[f'E{row}'].fill = result_fill
            ws_flex[f'E{row}'].font = result_font
            
        elif Lb <= flex['Lr']:
            ws_flex[f'A{row}'] = f"Lp ({flex['Lp']:.3f}m) < Lb ({Lb:.2f}m) ≤ Lr ({flex['Lr']:.3f}m)"
            ws_flex[f'A{row}'].font = Font(bold=True, size=11)
            ws_flex.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_flex[f'A{row}'] = "AISC Equation F2-2 applies (Inelastic Lateral-Torsional Buckling)"
            ws_flex[f'A{row}'].fill = equation_fill
            ws_flex.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_flex[f'A{row}'] = "Mn = Cb[Mp - (Mp - 0.7FySx)((Lb-Lp)/(Lr-Lp))] ≤ Mp"
            ws_flex[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws_flex.merge_cells(f'A{row}:E{row}')
            
            row += 1
            Mr = 0.7 * Fy * Sx / 100000
            ws_flex[f'A{row}'] = f"Mn = {Cb:.2f}[{flex['Mp']:.2f} - ({flex['Mp']:.2f} - {Mr:.2f}) × (({Lb:.2f}-{flex['Lp']:.3f})/({flex['Lr']:.3f}-{flex['Lp']:.3f}))]"
            ws_flex.merge_cells(f'A{row}:D{row}')
            ws_flex[f'E{row}'] = f"{flex['Mn']:.2f} t·m"
            ws_flex[f'E{row}'].fill = result_fill
            ws_flex[f'E{row}'].font = result_font
            
        else:
            ws_flex[f'A{row}'] = f"Lb ({Lb:.2f}m) > Lr ({flex['Lr']:.3f}m)"
            ws_flex[f'A{row}'].font = Font(bold=True, size=11)
            ws_flex.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_flex[f'A{row}'] = "AISC Equation F2-3 applies (Elastic Lateral-Torsional Buckling)"
            ws_flex[f'A{row}'].fill = equation_fill
            ws_flex.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_flex[f'A{row}'] = "Fcr = (Cbπ²E)/((Lb/rts)²) × √(1 + 0.078(Jc/(Sxho))×(Lb/rts)²)"
            ws_flex[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws_flex.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_flex[f'A{row}'] = "Mn = Fcr × Sx ≤ Mp"
            ws_flex.merge_cells(f'A{row}:D{row}')
            ws_flex[f'E{row}'] = f"{flex['Mn']:.2f} t·m"
            ws_flex[f'E{row}'].fill = result_fill
            ws_flex[f'E{row}'].font = result_font
        
        # Step 4: Design Strength
        row += 3
        ws_flex[f'A{row}'] = "STEP 4: CALCULATE DESIGN STRENGTH"
        ws_flex[f'A{row}'].font = section_font
        ws_flex[f'A{row}'].fill = section_fill
        ws_flex.merge_cells(f'A{row}:E{row}')
        
        row += 2
        ws_flex[f'A{row}'] = "Resistance Factor (AISC Section F1):"
        ws_flex[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_flex[f'A{row}'] = "φb = 0.90"
        ws_flex[f'A{row}'].fill = equation_fill
        
        row += 2
        ws_flex[f'A{row}'] = "φMn = φb × Mn"
        ws_flex.merge_cells(f'A{row}:B{row}')
        ws_flex[f'C{row}'] = f"= 0.90 × {flex['Mn']:.2f}"
        ws_flex[f'D{row}'] = "="
        ws_flex[f'D{row}'].alignment = center_align
        ws_flex[f'E{row}'] = f"{flex['phi_Mn']:.2f} t·m"
        ws_flex[f'E{row}'].fill = result_fill
        ws_flex[f'E{row}'].font = result_font
        ws_flex[f'E{row}'].border = thick_border
        
        # Summary Table
        row += 3
        ws_flex[f'A{row}'] = "FLEXURAL DESIGN SUMMARY"
        ws_flex[f'A{row}'].font = section_font
        ws_flex[f'A{row}'].fill = section_fill
        ws_flex.merge_cells(f'A{row}:C{row}')
        
        row += 1
        summary_data = [
            ['Parameter', 'Value', 'Status'],
            ['Design Moment Capacity (φMn)', f"{flex['phi_Mn']:.2f} t·m", ''],
            ['Nominal Moment (Mn)', f"{flex['Mn']:.2f} t·m", ''],
            ['Plastic Moment (Mp)', f"{flex['Mp']:.2f} t·m", ''],
            ['Limiting Length Lp', f"{flex['Lp']:.3f} m", ''],
            ['Limiting Length Lr', f"{flex['Lr']:.3f} m", ''],
            ['Design Case', flex['case'], ''],
            ['Design Zone', flex['zone'], ''],
            ['Utilization Ratio', f"{flex['ratio']:.3f}", '✓ PASS' if flex['adequate'] else '✗ FAIL'],
        ]
        
        for data_row in summary_data:
            row += 1
            for col, value in enumerate(data_row, start=1):
                cell = ws_flex.cell(row=row, column=col)
                cell.value = value
                cell.border = thin_border
                cell.alignment = center_align if col > 1 else left_align
                
                if 'Parameter' in str(value):
                    cell.font = header_font
                    cell.fill = header_fill
                elif col == 3 and '✓' in str(value):
                    cell.fill = adequate_fill
                    cell.font = adequate_font
                elif col == 3 and '✗' in str(value):
                    cell.fill = inadequate_fill
                    cell.font = inadequate_font
        
        # Auto-size columns
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_flex.column_dimensions[col].width = 25
    
    # ==================== SHEET 5: COMPRESSION ANALYSIS ====================
    if analysis_results and 'compression' in analysis_results:
        ws_comp = wb.create_sheet("Compression Analysis")
        
        row = 1
        ws_comp['A1'] = "AISC 360-16 CHAPTER E3: COMPRESSION DESIGN"
        ws_comp['A1'].font = title_font
        ws_comp['A1'].fill = title_fill
        ws_comp['A1'].alignment = center_align
        ws_comp.merge_cells('A1:E1')
        ws_comp.row_dimensions[1].height = 25
        
        comp = analysis_results['compression']
        KL = design_params.get('KL', 0)
        
        Ag = safe_scalar(df.loc[section, 'A [cm2]'])
        rx = safe_scalar(df.loc[section, 'rx [cm]'])
        ry = safe_scalar(df.loc[section, 'ry [cm]'])
        
        # Step 1: Slenderness Ratio
        row = 3
        ws_comp[f'A{row}'] = "STEP 1: CALCULATE SLENDERNESS RATIO"
        ws_comp[f'A{row}'].font = section_font
        ws_comp[f'A{row}'].fill = section_fill
        ws_comp.merge_cells(f'A{row}:E{row}')
        
        row += 2
        KL_cm = KL * 100
        lambda_x = KL_cm / rx
        lambda_y = KL_cm / ry
        lambda_c = max(lambda_x, lambda_y)
        
        ws_comp[f'A{row}'] = f"KL/rx = {KL_cm:.1f} / {rx:.2f}"
        ws_comp.merge_cells(f'A{row}:D{row}')
        ws_comp[f'E{row}'] = f"{lambda_x:.1f}"
        ws_comp[f'E{row}'].fill = calc_fill
        
        row += 1
        ws_comp[f'A{row}'] = f"KL/ry = {KL_cm:.1f} / {ry:.2f}"
        ws_comp.merge_cells(f'A{row}:D{row}')
        ws_comp[f'E{row}'] = f"{lambda_y:.1f}"
        ws_comp[f'E{row}'].fill = calc_fill
        
        row += 1
        ws_comp[f'A{row}'] = "λc = max(KL/rx, KL/ry)"
        ws_comp.merge_cells(f'A{row}:D{row}')
        ws_comp[f'E{row}'] = f"{lambda_c:.1f}"
        ws_comp[f'E{row}'].fill = result_fill
        ws_comp[f'E{row}'].font = result_font
        
        # Step 2: Elastic Buckling Stress
        row += 3
        ws_comp[f'A{row}'] = "STEP 2: CALCULATE ELASTIC BUCKLING STRESS Fe"
        ws_comp[f'A{row}'].font = section_font
        ws_comp[f'A{row}'].fill = section_fill
        ws_comp.merge_cells(f'A{row}:E{row}')
        
        row += 2
        ws_comp[f'A{row}'] = "AISC Equation E3-4:"
        ws_comp[f'A{row}'].font = Font(bold=True, italic=True)
        ws_comp[f'A{row}'].fill = equation_fill
        ws_comp.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws_comp[f'A{row}'] = "Fe = π²E / (KL/r)²"
        ws_comp.merge_cells(f'A{row}:E{row}')
        
        row += 1
        Fe = (math.pi**2 * E) / (lambda_c**2)
        ws_comp[f'A{row}'] = f"Fe = π² × {E:.0f} / {lambda_c:.1f}²"
        ws_comp.merge_cells(f'A{row}:D{row}')
        ws_comp[f'E{row}'] = f"{Fe:.1f} kgf/cm²"
        ws_comp[f'E{row}'].fill = result_fill
        ws_comp[f'E{row}'].font = result_font
        
        # Step 3: Determine Critical Stress
        row += 3
        ws_comp[f'A{row}'] = "STEP 3: DETERMINE CRITICAL STRESS Fcr"
        ws_comp[f'A{row}'].font = section_font
        ws_comp[f'A{row}'].fill = section_fill
        ws_comp.merge_cells(f'A{row}:E{row}')
        
        row += 2
        lambda_limit = 4.71 * safe_sqrt(E / Fy)
        ws_comp[f'A{row}'] = "Limiting Slenderness:"
        ws_comp[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_comp[f'A{row}'] = f"4.71√(E/Fy) = 4.71√({E:.0f}/{Fy:.1f})"
        ws_comp.merge_cells(f'A{row}:D{row}')
        ws_comp[f'E{row}'] = f"{lambda_limit:.1f}"
        ws_comp[f'E{row}'].fill = calc_fill
        
        row += 2
        if lambda_c <= lambda_limit:
            ws_comp[f'A{row}'] = f"λc ({lambda_c:.1f}) ≤ {lambda_limit:.1f}"
            ws_comp[f'A{row}'].font = Font(bold=True, size=11)
            ws_comp.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_comp[f'A{row}'] = "INELASTIC BUCKLING - AISC Equation E3-2:"
            ws_comp[f'A{row}'].fill = equation_fill
            ws_comp.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_comp[f'A{row}'] = "Fcr = [0.658^(Fy/Fe)] × Fy"
            ws_comp.merge_cells(f'A{row}:E{row}')
            
            row += 1
            exponent = Fy / Fe
            Fcr = (0.658 ** exponent) * Fy
            ws_comp[f'A{row}'] = f"Fcr = [0.658^({Fy:.1f}/{Fe:.1f})] × {Fy:.1f}"
            ws_comp.merge_cells(f'A{row}:D{row}')
            ws_comp[f'E{row}'] = f"{Fcr:.1f} kgf/cm²"
            ws_comp[f'E{row}'].fill = result_fill
            ws_comp[f'E{row}'].font = result_font
        else:
            ws_comp[f'A{row}'] = f"λc ({lambda_c:.1f}) > {lambda_limit:.1f}"
            ws_comp[f'A{row}'].font = Font(bold=True, size=11)
            ws_comp.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_comp[f'A{row}'] = "ELASTIC BUCKLING - AISC Equation E3-3:"
            ws_comp[f'A{row}'].fill = equation_fill
            ws_comp.merge_cells(f'A{row}:E{row}')
            
            row += 1
            ws_comp[f'A{row}'] = "Fcr = 0.877 × Fe"
            ws_comp.merge_cells(f'A{row}:E{row}')
            
            row += 1
            Fcr = 0.877 * Fe
            ws_comp[f'A{row}'] = f"Fcr = 0.877 × {Fe:.1f}"
            ws_comp.merge_cells(f'A{row}:D{row}')
            ws_comp[f'E{row}'] = f"{Fcr:.1f} kgf/cm²"
            ws_comp[f'E{row}'].fill = result_fill
            ws_comp[f'E{row}'].font = result_font
        
        # Step 4: Calculate Strength
        row += 3
        ws_comp[f'A{row}'] = "STEP 4: CALCULATE NOMINAL AND DESIGN STRENGTH"
        ws_comp[f'A{row}'].font = section_font
        ws_comp[f'A{row}'].fill = section_fill
        ws_comp.merge_cells(f'A{row}:E{row}')
        
        row += 2
        ws_comp[f'A{row}'] = "AISC Equation E3-1:"
        ws_comp[f'A{row}'].font = Font(bold=True, italic=True)
        ws_comp[f'A{row}'].fill = equation_fill
        ws_comp.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws_comp[f'A{row}'] = "Pn = Fcr × Ag"
        ws_comp.merge_cells(f'A{row}:E{row}')
        
        row += 1
        Pn = Fcr * Ag / 1000
        ws_comp[f'A{row}'] = f"Pn = {Fcr:.1f} × {Ag:.2f} / 1000"
        ws_comp.merge_cells(f'A{row}:D{row}')
        ws_comp[f'E{row}'] = f"{Pn:.2f} tons"
        ws_comp[f'E{row}'].fill = result_fill
        ws_comp[f'E{row}'].font = result_font
        
        row += 2
        ws_comp[f'A{row}'] = "Resistance Factor (AISC Section E1):"
        ws_comp[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws_comp[f'A{row}'] = "φc = 0.90"
        ws_comp[f'A{row}'].fill = equation_fill
        
        row += 2
        ws_comp[f'A{row}'] = "φPn = φc × Pn"
        ws_comp.merge_cells(f'A{row}:B{row}')
        ws_comp[f'C{row}'] = f"= 0.90 × {Pn:.2f}"
        ws_comp[f'D{row}'] = "="
        ws_comp[f'D{row}'].alignment = center_align
        ws_comp[f'E{row}'] = f"{comp['phi_Pn']:.2f} tons"
        ws_comp[f'E{row}'].fill = result_fill
        ws_comp[f'E{row}'].font = result_font
        ws_comp[f'E{row}'].border = thick_border
        
        # Summary Table
        row += 3
        ws_comp[f'A{row}'] = "COMPRESSION DESIGN SUMMARY"
        ws_comp[f'A{row}'].font = section_font
        ws_comp[f'A{row}'].fill = section_fill
        ws_comp.merge_cells(f'A{row}:C{row}')
        
        row += 1
        summary_data = [
            ['Parameter', 'Value', 'Status'],
            ['Design Strength (φPn)', f"{comp['phi_Pn']:.2f} tons", ''],
            ['Nominal Strength (Pn)', f"{comp['Pn']:.2f} tons", ''],
            ['Critical Stress (Fcr)', f"{comp['Fcr']:.1f} kgf/cm²", ''],
            ['Elastic Buckling Stress (Fe)', f"{Fe:.1f} kgf/cm²", ''],
            ['Slenderness Ratio (λc)', f"{comp['lambda_c']:.1f}", ''],
            ['Buckling Mode', comp['mode'], ''],
            ['Utilization Ratio', f"{comp['ratio']:.3f}", '✓ PASS' if comp['adequate'] else '✗ FAIL'],
        ]
        
        for data_row in summary_data:
            row += 1
            for col, value in enumerate(data_row, start=1):
                cell = ws_comp.cell(row=row, column=col)
                cell.value = value
                cell.border = thin_border
                cell.alignment = center_align if col > 1 else left_align
                
                if 'Parameter' in str(value):
                    cell.font = header_font
                    cell.fill = header_fill
                elif col == 3 and '✓' in str(value):
                    cell.fill = adequate_fill
                    cell.font = adequate_font
                elif col == 3 and '✗' in str(value):
                    cell.fill = inadequate_fill
                    cell.font = inadequate_font
        
        # Auto-size columns
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_comp.column_dimensions[col].width = 25
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==================== CSS STYLES ====================

def get_tab_styles():
    """Return CSS styles for the Streamlit tab"""
    return """
    <style>
        .report-preview {
            border: 1px solid #ddd;
            border-radius: 8px;
            padding: 15px;
            background: #fafafa;
            max-height: 600px;
            overflow-y: auto;
        }
        .config-section {
            background: #f8f9fa;
            border-radius: 8px;
            padding: 15px;
            margin-bottom: 15px;
        }
        .member-card {
            border: 1px solid #e0e0e0;
            border-radius: 5px;
            padding: 10px;
            margin: 5px 0;
            background: white;
        }
        .status-pass {
            color: #28a745;
            font-weight: bold;
        }
        .status-fail {
            color: #dc3545;
            font-weight: bold;
        }
        .summary-metric {
            text-align: center;
            padding: 10px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-radius: 8px;
            margin: 5px;
        }
        .summary-metric h3 {
            margin: 0;
            font-size: 24px;
        }
        .summary-metric p {
            margin: 5px 0 0 0;
            font-size: 12px;
            opacity: 0.9;
        }
    </style>
    """


# ==================== HELPER FUNCTIONS ====================

def get_section_properties_from_df(df_sections, df_materials, section_name, material_name):
    """
    Extract section properties from dataframes
    
    Parameters:
    - df_sections: DataFrame with section properties
    - df_materials: DataFrame with material properties
    - section_name: Name/index of section
    - material_name: Name/index of material
    
    Returns dict with section properties
    """
    try:
        sec = df_sections.loc[section_name]
        mat = df_materials.loc[material_name]
        
        # Map column names (adjust based on your actual column names)
        props = {
            'Ag': float(sec.get('A [cm2]', sec.get('Ag', 0))),
            'Ix': float(sec.get('Ix [cm4]', sec.get('Ix', 0))),
            'Iy': float(sec.get('Iy [cm4]', sec.get('Iy', 0))),
            'Sx': float(sec.get('Sx [cm3]', sec.get('Sx', 0))),
            'Zx': float(sec.get('Zx [cm3]', sec.get('Zx', 0))),
            'Sy': float(sec.get('Sy [cm3]', sec.get('Sy', 0))),
            'Zy': float(sec.get('Zy [cm3]', sec.get('Zy', 0))),
            'rx': float(sec.get('rx [cm]', sec.get('rx', 0))),
            'ry': float(sec.get('ry [cm]', sec.get('ry', 0))),
            'J': float(sec.get('J [cm4]', sec.get('J', 1))),
            'd': float(sec.get('d [mm]', sec.get('d', 0))) / 10,  # Convert mm to cm
            'bf': float(sec.get('bf [mm]', sec.get('bf', 0))) / 10,
            'tf': float(sec.get('tf [mm]', sec.get('tf', 0))) / 10,
            'tw': float(sec.get('tw [mm]', sec.get('tw', 0))) / 10,
            'Fy': float(mat.get('Yield Point (ksc)', mat.get('Fy', 2500))) / 10.197,  # ksc to MPa
            'Fu': float(mat.get('Tensile Strength (ksc)', mat.get('Fu', 4000))) / 10.197,
        }
        
        # Calculate derived properties if needed
        if props['d'] > 0 and props['tf'] > 0:
            props['ho'] = props['d'] - props['tf']
        else:
            props['ho'] = props['d'] * 0.9
        
        # Effective area for tension (assume 85% if not specified)
        props['Ae'] = props['Ag'] * 0.85
        
        # rts approximation
        props['rts'] = props['ry'] * 1.1 if props['ry'] > 0 else 1
        
        # Cw approximation
        if props['Iy'] > 0 and props['d'] > 0:
            props['Cw'] = props['Iy'] * (props['d'] - props['tf'])**2 / 4
        else:
            props['Cw'] = 1
        
        return props
        
    except Exception as e:
        st.error(f"Error extracting section properties: {e}")
        return None


def get_html_download_link(html_content, filename="report.html"):
    """Generate a download link for HTML content"""
    b64 = base64.b64encode(html_content.encode()).decode()
    href = f'<a href="data:text/html;base64,{b64}" download="{filename}" class="download-btn">📥 Download HTML Report</a>'
    return href


# ==================== MAIN TAB FUNCTION ====================

def render_design_report_tab(df_sections, df_materials, loaded_data=None, member_groups=None, analysis_results=None):
    """
    Render the Steel Design Report tab in Streamlit
    
    Parameters:
    - df_sections: DataFrame with available sections
    - df_materials: DataFrame with available materials
    - loaded_data: DataFrame with load data (optional, from Tab 5)
    - member_groups: Dict with member configurations (optional, from Tab 5)
    - analysis_results: Dict with analysis results (optional, from Tab 5)
    """
    
    st.markdown(get_tab_styles(), unsafe_allow_html=True)
    st.markdown("## 📄 Steel Design Report Generator")
    st.markdown("Generate equation-based design reports per AISC 360-16")
    
    # Initialize session state
    if 'report_members' not in st.session_state:
        st.session_state.report_members = []
    if 'generated_report' not in st.session_state:
        st.session_state.generated_report = None
    
    # Create sub-tabs
    subtab1, subtab2, subtab3 = st.tabs([
        "📝 Configure Members",
        "⚙️ Generate Report", 
        "📊 Preview & Export"
    ])
    
    # ==================== SUB-TAB 1: CONFIGURE MEMBERS ====================
    with subtab1:
        st.markdown("### Member Configuration")
        
        # Option to import from existing analysis
        if analysis_results and len(analysis_results) > 0:
            st.info(f"💡 Found {len(analysis_results)} analyzed members from Design Check tab")
            
            if st.button("📥 Import Analyzed Members", type="primary"):
                imported_members = []
                
                for member_no, data in analysis_results.items():
                    config = data.get('config', {})
                    results_df = data.get('results', pd.DataFrame())
                    
                    # Get section properties
                    section_name = config.get('section', list(df_sections.index)[0])
                    material_name = config.get('material', list(df_materials.index)[0])
                    
                    section_props = get_section_properties_from_df(
                        df_sections, df_materials, section_name, material_name
                    )
                    
                    if section_props is None:
                        continue
                    
                    # Convert loads
                    loads = []
                    for _, row in results_df.iterrows():
                        # Convert from tons/t·m to kN/kN·m
                        Pu_kN = float(row.get('Pu (tons)', 0)) * 9.81
                        Mu_kNm = float(row.get('Mu (t·m)', 0)) * 9.81
                        
                        loads.append({
                            'LC': int(row.get('LC', 1)),
                            'Pu': Pu_kN,
                            'Mux': Mu_kNm,
                            'Muy': 0
                        })
                    
                    # Determine member type
                    member_type_raw = config.get('member_type', 'Beam-Column (Compression)')
                    if 'Tension Member' in member_type_raw:
                        member_type = 'Tension Member'
                    elif 'Beam-Column' in member_type_raw:
                        member_type = 'Beam-Column'
                    elif 'Column' in member_type_raw:
                        member_type = 'Column'
                    else:
                        member_type = 'Beam'
                    
                    # Create member data
                    member = create_member_data(
                        member_no=member_no,
                        section_name=section_name,
                        section_props=section_props,
                        member_type=member_type,
                        length=config.get('Lb', 3.0),
                        K=1.0,
                        KL=config.get('KL', 3.0),
                        Lb=config.get('Lb', 3.0),
                        loads=loads,
                        classification='Compact'  # Assume compact for now
                    )
                    
                    imported_members.append(member)
                
                st.session_state.report_members = imported_members
                st.success(f"✅ Imported {len(imported_members)} members")
                st.rerun()
        
        st.markdown("---")
        
        # Manual member entry
        st.markdown("### ➕ Add Member Manually")
        
        col1, col2 = st.columns(2)
        
        with col1:
            new_member_no = st.text_input("Member No.", value="M-001", key="new_member_no")
            new_section = st.selectbox("Section", list(df_sections.index), key="new_section")
            new_material = st.selectbox("Material", list(df_materials.index), key="new_material")
            new_type = st.selectbox("Member Type", 
                                   ["Beam", "Column", "Beam-Column", "Tension Member"],
                                   key="new_type")
        
        with col2:
            new_length = st.number_input("Length (m)", 0.1, 30.0, 4.0, 0.1, key="new_length")
            new_K = st.number_input("K Factor", 0.5, 2.0, 1.0, 0.05, key="new_K")
            new_KL = st.number_input("KL (m)", 0.1, 30.0, 4.0, 0.1, key="new_KL")
            new_Lb = st.number_input("Lb (m)", 0.1, 30.0, 2.0, 0.1, key="new_Lb")
        
        # Load entry
        st.markdown("#### Load Combinations")
        
        load_input_method = st.radio("Load Input Method:", 
                                     ["Manual Entry", "Paste from Excel/CSV"],
                                     horizontal=True, key="load_input_method")
        
        if load_input_method == "Manual Entry":
            num_loads = st.number_input("Number of Load Combinations", 1, 20, 3, key="num_loads")
            
            loads = []
            cols = st.columns(4)
            cols[0].markdown("**LC**")
            cols[1].markdown("**Pu (kN)**")
            cols[2].markdown("**Mux (kN·m)**")
            cols[3].markdown("**Muy (kN·m)**")
            
            for i in range(int(num_loads)):
                cols = st.columns(4)
                lc = cols[0].number_input(f"LC{i+1}", value=i+1, key=f"lc_{i}", label_visibility="collapsed")
                pu = cols[1].number_input(f"Pu{i+1}", value=0.0, key=f"pu_{i}", label_visibility="collapsed")
                mux = cols[2].number_input(f"Mux{i+1}", value=0.0, key=f"mux_{i}", label_visibility="collapsed")
                muy = cols[3].number_input(f"Muy{i+1}", value=0.0, key=f"muy_{i}", label_visibility="collapsed")
                
                loads.append({'LC': int(lc), 'Pu': pu, 'Mux': mux, 'Muy': muy})
        else:
            st.markdown("Paste data with columns: LC, Pu, Mux, Muy")
            load_text = st.text_area("Paste Load Data:", height=150, key="load_paste")
            
            loads = []
            if load_text:
                try:
                    lines = load_text.strip().split('\n')
                    for line in lines:
                        parts = line.replace(',', '\t').split('\t')
                        if len(parts) >= 3:
                            loads.append({
                                'LC': int(float(parts[0])),
                                'Pu': float(parts[1]),
                                'Mux': float(parts[2]),
                                'Muy': float(parts[3]) if len(parts) > 3 else 0
                            })
                    st.success(f"Parsed {len(loads)} load combinations")
                except Exception as e:
                    st.error(f"Error parsing: {e}")
        
        # Add member button
        if st.button("➕ Add Member to Report", type="primary", key="add_member"):
            if len(loads) == 0:
                st.warning("Please enter at least one load combination")
            else:
                section_props = get_section_properties_from_df(
                    df_sections, df_materials, new_section, new_material
                )
                
                if section_props:
                    member = create_member_data(
                        member_no=new_member_no,
                        section_name=new_section,
                        section_props=section_props,
                        member_type=new_type,
                        length=new_length,
                        K=new_K,
                        KL=new_KL,
                        Lb=new_Lb,
                        loads=loads,
                        classification='Compact'
                    )
                    
                    st.session_state.report_members.append(member)
                    st.success(f"✅ Added member {new_member_no}")
                    st.rerun()
        
        # Display configured members
        st.markdown("---")
        st.markdown("### 📋 Configured Members")
        
        if len(st.session_state.report_members) == 0:
            st.info("No members configured yet. Add members above or import from analysis.")
        else:
            for i, member in enumerate(st.session_state.report_members):
                # Get governing ratio
                interactions = member.get('interaction_results', [])
                if interactions:
                    max_ratio = max(inter.get('interaction_ratio', 0) for inter in interactions)
                    status = "✓ PASS" if max_ratio <= 1.0 else "✗ FAIL"
                    status_class = "status-pass" if max_ratio <= 1.0 else "status-fail"
                else:
                    max_ratio = 0
                    status = "—"
                    status_class = ""
                
                col1, col2, col3, col4 = st.columns([2, 2, 1, 1])
                
                with col1:
                    st.markdown(f"**{member['member_no']}** | {member['section_name']}")
                
                with col2:
                    st.markdown(f"{member['member_type']} | {len(member.get('loads', []))} LCs")
                
                with col3:
                    st.markdown(f"<span class='{status_class}'>{status}</span> ({max_ratio:.3f})", 
                               unsafe_allow_html=True)
                
                with col4:
                    if st.button("🗑️", key=f"del_member_{i}"):
                        st.session_state.report_members.pop(i)
                        st.rerun()
            
            if st.button("🗑️ Clear All Members", type="secondary"):
                st.session_state.report_members = []
                st.rerun()
    
    # ==================== SUB-TAB 2: GENERATE REPORT ====================
    with subtab2:
        st.markdown("### Report Settings")
        
        col1, col2 = st.columns(2)
        
        with col1:
            project_name = st.text_input("Project Name", value="Steel Structure Project", key="proj_name")
            engineer_name = st.text_input("Engineer Name", value="", key="eng_name")
        
        with col2:
            report_date = st.date_input("Report Date", value=datetime.now(), key="report_date")
            include_pm_diagram = st.checkbox("Include P-M Interaction Diagram", value=False, key="include_pm")
        
        st.markdown("---")
        
        # Summary of members
        st.markdown("### Members to Include")
        
        if len(st.session_state.report_members) == 0:
            st.warning("⚠️ No members configured. Go to 'Configure Members' tab.")
        else:
            # Summary metrics
            total = len(st.session_state.report_members)
            passing = sum(1 for m in st.session_state.report_members 
                         if max((i.get('interaction_ratio', 0) for i in m.get('interaction_results', [{}])), default=0) <= 1.0)
            failing = total - passing
            
            cols = st.columns(4)
            
            with cols[0]:
                st.markdown(f"""
                <div class="summary-metric">
                    <h3>{total}</h3>
                    <p>Total Members</p>
                </div>
                """, unsafe_allow_html=True)
            
            with cols[1]:
                st.markdown(f"""
                <div class="summary-metric" style="background: linear-gradient(135deg, #28a745 0%, #218838 100%);">
                    <h3>{passing}</h3>
                    <p>Passing</p>
                </div>
                """, unsafe_allow_html=True)
            
            with cols[2]:
                st.markdown(f"""
                <div class="summary-metric" style="background: linear-gradient(135deg, #dc3545 0%, #c82333 100%);">
                    <h3>{failing}</h3>
                    <p>Failing</p>
                </div>
                """, unsafe_allow_html=True)
            
            with cols[3]:
                max_ratio_all = max(
                    max((i.get('interaction_ratio', 0) for i in m.get('interaction_results', [{}])), default=0)
                    for m in st.session_state.report_members
                ) if st.session_state.report_members else 0
                
                st.markdown(f"""
                <div class="summary-metric" style="background: linear-gradient(135deg, #ffc107 0%, #e0a800 100%);">
                    <h3>{max_ratio_all:.3f}</h3>
                    <p>Max Ratio</p>
                </div>
                """, unsafe_allow_html=True)
            
            # Member list
            st.markdown("#### Member Summary")
            
            summary_data = []
            for m in st.session_state.report_members:
                interactions = m.get('interaction_results', [])
                max_ratio = max((i.get('interaction_ratio', 0) for i in interactions), default=0)
                gov_lc = "—"
                if interactions:
                    max_idx = max(range(len(interactions)), 
                                 key=lambda i: interactions[i].get('interaction_ratio', 0))
                    gov_lc = interactions[max_idx].get('LC', '—')
                
                summary_data.append({
                    'Member': m['member_no'],
                    'Section': m['section_name'],
                    'Type': m['member_type'],
                    '# LCs': len(m.get('loads', [])),
                    'Gov. LC': gov_lc,
                    'Max Ratio': max_ratio,
                    'Status': '✓ OK' if max_ratio <= 1.0 else '✗ NG'
                })
            
            df_summary = pd.DataFrame(summary_data)
            st.dataframe(df_summary, use_container_width=True)
            
            # Generate button
            st.markdown("---")
            
            if st.button("🚀 Generate Report", type="primary", key="generate_report"):
                with st.spinner("Generating report..."):
                    # Create report generator
                    generator = SteelDesignReportGenerator(
                        project_info={
                            'name': project_name,
                            'engineer': engineer_name,
                            'date': report_date.strftime("%Y-%m-%d")
                        }
                    )
                    
                    # Add members
                    for member in st.session_state.report_members:
                        generator.add_member(member)
                    
                    # Generate HTML
                    html_report = generator.generate_full_report()
                    st.session_state.generated_report = html_report
                    
                st.success("✅ Report generated successfully!")
                st.info("Go to 'Preview & Export' tab to view and download")
    
    # ==================== SUB-TAB 3: PREVIEW & EXPORT ====================
    with subtab3:
        st.markdown("### Report Preview & Export")
        
        if st.session_state.generated_report is None:
            st.warning("⚠️ No report generated yet. Go to 'Generate Report' tab.")
        else:
            # Export buttons
            col1, col2, col3 = st.columns(3)
            
            with col1:
                # HTML download
                b64 = base64.b64encode(st.session_state.generated_report.encode()).decode()
                filename = f"Steel_Design_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
                
                st.download_button(
                    label="📥 Download HTML",
                    data=st.session_state.generated_report,
                    file_name=filename,
                    mime="text/html",
                    key="download_html"
                )
            
            with col2:
                st.info("💡 Open HTML in browser, then Print to PDF")
            
            with col3:
                if st.button("🔄 Regenerate", key="regenerate"):
                    st.session_state.generated_report = None
                    st.rerun()
            
            st.markdown("---")
            
            # Preview
            st.markdown("### 📄 Report Preview")
            
            # Use iframe for preview
            st.components.v1.html(
                st.session_state.generated_report,
                height=800,
                scrolling=True
            )


# ==================== STANDALONE TAB CODE ====================
# Copy this section into your main Streamlit app

TAB_CODE = '''
# ==================== TAB: DESIGN REPORT ====================
# Add this to your main Streamlit application

with tab_report:  # Change to your tab variable name
    st.markdown('<h2 class="section-header">📄 Steel Design Report Generator</h2>', unsafe_allow_html=True)
    
    # Import the report tab renderer
    from steel_design_report_streamlit import render_design_report_tab
    
    # Get analysis results from Tab 5 if available
    loaded_data = st.session_state.get('loaded_data', None)
    member_groups = st.session_state.get('member_groups', {})
    analysis_results = st.session_state.get('analysis_results_tab5', {})
    
    # Render the report tab
    render_design_report_tab(
        df_sections=df,          # Your sections DataFrame
        df_materials=df_mat,     # Your materials DataFrame
        loaded_data=loaded_data,
        member_groups=member_groups,
        analysis_results=analysis_results
    )
'''

if __name__ == "__main__":
    # Demo/test mode - page config already set at top of file
    # Note: st.set_page_config() is already called at line ~20
    
    # Create sample DataFrames for testing
    df_sections = pd.DataFrame({
        'A [cm2]': [41.8, 34.2, 78.5],
        'Ix [cm4]': [8247, 3040, 15600],
        'Iy [cm4]': [533, 310, 1200],
        'Sx [cm3]': [467, 278, 820],
        'Zx [cm3]': [524, 316, 920],
        'rx [cm]': [14.0, 9.42, 12.5],
        'ry [cm]': [3.57, 3.00, 4.20],
        'J [cm4]': [16.4, 8.5, 35.0],
        'd [mm]': [350, 200, 400],
        'bf [mm]': [127, 134, 180],
        'tf [mm]': [10.7, 8.4, 13.5],
        'tw [mm]': [6.1, 5.8, 8.6]
    }, index=['W14x22', 'W8x18', 'W16x36'])
    
    df_materials = pd.DataFrame({
        'Yield Point (ksc)': [3515, 2530],  # 345 MPa, 248 MPa
        'Tensile Strength (ksc)': [4588, 4078]  # 450 MPa, 400 MPa
    }, index=['ASTM A992 Gr.50', 'ASTM A36'])
    
    # This demo code will only run if this file is executed directly
    # When imported as part of the main app, this section is skipped
    st.info("ℹ️ Running in demo mode with sample data")
    # render_design_report_tab(df_sections, df_materials)  # Uncomment if needed

# ==================== EVALUATION FUNCTION ====================
def evaluate_section_design(df, df_mat, section, material, design_loads, design_lengths):
    """Comprehensive section evaluation"""
    try:
        weight_col = 'Unit Weight [kg/m]' if 'Unit Weight [kg/m]' in df.columns else 'w [kg/m]'
        weight = safe_scalar(df.loc[section, weight_col])
        
        flex_result = aisc_360_16_f2_flexural_design(df, df_mat, section, material, design_lengths['Lb'])
        comp_result = aisc_360_16_e3_compression_design(df, df_mat, section, material, 
                                                       design_lengths['KLx'], design_lengths['KLy'])
        
        if flex_result and comp_result:
            phi_Mn = 0.9 * flex_result['Mn']
            phi_Pn = comp_result['phi_Pn']
            
            moment_ratio = design_loads['Mu'] / phi_Mn if phi_Mn > 0 else 999
            axial_ratio = design_loads['Pu'] / phi_Pn if phi_Pn > 0 else 999
            
            flexural_adequate = moment_ratio <= 1.0
            compression_adequate = axial_ratio <= 1.0
            overall_adequate = flexural_adequate and compression_adequate
            
            Lp = flex_result['Lp']
            Lr = flex_result['Lr'] 
            current_Lb = design_lengths['Lb']
            
            if current_Lb <= Lp:
                flexural_zone = "Yielding (F2.1)"
            elif current_Lb <= Lr:
                flexural_zone = "Inelastic LTB (F2.2)"
            else:
                flexural_zone = "Elastic LTB (F2.3)"
            
            return {
                'section': section,
                'material': material,
                'weight': weight,
                'flexural': {
                    'Mn': flex_result['Mn'],
                    'phi_Mn': phi_Mn,
                    'Mp': flex_result['Mp'],
                    'Lp': Lp,
                    'Lr': Lr,
                    'case': flex_result['Case'],
                    'zone': flexural_zone,
                    'ratio': moment_ratio,
                    'adequate': flexural_adequate,
                },
                'compression': {
                    'Pn': comp_result['Pn'],
                    'phi_Pn': phi_Pn,
                    'Fcr': comp_result['Fcr'],
                    'lambda_c': comp_result['lambda_c'],
                    'mode': comp_result['buckling_mode'],
                    'ratio': axial_ratio,
                    'adequate': compression_adequate,
                },
                'design_check': {
                    'overall_adequate': overall_adequate,
                    'moment_utilization': moment_ratio,
                    'axial_utilization': axial_ratio,
                }
            }
    except Exception as e:
        st.error(f"Error in section evaluation: {e}")
        return None

# ==================== LOAD DATA ====================
@st.cache_data
def load_data():
    """Load steel section and material databases"""
    try:
        file_path = "https://raw.githubusercontent.com/Thana-site/Steel_Design_2003/main/2003-Steel-Beam-DataBase-H-Shape.csv"
        file_path_mat = "https://raw.githubusercontent.com/Thana-site/Steel_Design_2003/main/2003-Steel-Beam-DataBase-Material.csv"
        
        df = pd.read_csv(file_path, index_col=0, encoding='ISO-8859-1')
        df_mat = pd.read_csv(file_path_mat, index_col=0, encoding="utf-8")
        return df, df_mat, True
    except Exception as e:
        return pd.DataFrame(), pd.DataFrame(), False

# Load data immediately
df, df_mat, data_loaded = load_data()

# Stop if data failed to load
if not data_loaded:
    st.error("❌ Failed to load data. Please check your internet connection.")
    st.stop()

# ==================== NOW SAFE TO INITIALIZE SESSION STATE ====================
if 'selected_section' not in st.session_state:
    st.session_state.selected_section = list(df.index)[0] if len(df.index) > 0 else None

if 'selected_material' not in st.session_state:
    st.session_state.selected_material = list(df_mat.index)[0] if len(df_mat.index) > 0 else None

if 'selected_sections' not in st.session_state:
    st.session_state.selected_sections = []

if 'calculation_report' not in st.session_state:
    st.session_state.calculation_report = ""

if 'evaluation_results' not in st.session_state:
    st.session_state.evaluation_results = None

if 'project_info' not in st.session_state:
    st.session_state.project_info = {
        'project_name': '',
        'project_no': '',
        'designer': '',
        'checker': '',
        'date': datetime.now().strftime('%Y-%m-%d'),
        'revision': '0'
    }

if 'report_members' not in st.session_state:
    st.session_state.report_members = []

if 'generated_report' not in st.session_state:
    st.session_state.generated_report = None

if 'loaded_data' not in st.session_state:
    st.session_state.loaded_data = None

if 'member_groups' not in st.session_state:
    st.session_state.member_groups = {}

if 'analysis_results_tab5' not in st.session_state:
    st.session_state.analysis_results_tab5 = {}

# ==================== LIBRARY STATUS WARNINGS ====================
if not PDF_AVAILABLE:
    st.sidebar.warning("⚠️ PDF export unavailable. Install: `pip install reportlab`")
if not EXCEL_AVAILABLE:
    st.sidebar.warning("⚠️ Excel export unavailable. Install: `pip install openpyxl`")

# ==================== MAIN HEADER ====================
st.markdown('<h1 class="main-header">AISC 360-16 Steel Design v7.0</h1>', unsafe_allow_html=True)
st.markdown('<p style="text-align: center; color: #7f8c8d; font-size: 1.1rem; font-weight: 500;">UI/UX | Advanced Export Capabilities | Enhanced Visualizations</p>', unsafe_allow_html=True)

# ==================== SIDEBAR WITH CUSTOM HTML DROPDOWNS ====================
with st.sidebar:
    st.markdown("### 🔧 Design Configuration")
    st.markdown("---")

    # ========== MATERIAL SELECTION (CUSTOM HTML) ==========
    material_list = [str(x).strip() for x in df_mat.index.tolist()]

    # Initialize
    if 'selected_material' not in st.session_state or not st.session_state.selected_material:
        st.session_state.selected_material = material_list[0]

    # Clean and validate
    st.session_state.selected_material = str(st.session_state.selected_material).strip()
    if st.session_state.selected_material not in material_list:
        st.session_state.selected_material = material_list[0]

    # Create options HTML
    material_options = ""
    for mat in material_list:
        selected_attr = 'selected="selected"' if mat == st.session_state.selected_material else ''
        material_options += f'<option value="{mat}" {selected_attr}>{mat}</option>\n'

    # Custom HTML dropdown for material
    st.markdown("**⚙️ Steel Grade:**")
    material_html = f"""
    <script src="https://cdn.jsdelivr.net/npm/streamlit-component-lib@1.0.0/dist/streamlit-component-lib.js"></script>
    <select id="material_select" style="width: 100%; padding: 8px; font-size: 14px; border: 1px solid #ccc; border-radius: 4px;">
        {material_options}
    </select>
    <script>
        var select = document.getElementById('material_select');
        select.value = '{st.session_state.selected_material}';

        select.addEventListener('change', function() {{
            window.parent.Streamlit.setComponentValue(this.value);
        }});
    </script>
    """
    selected_material_new = components.html(material_html, height=50)

    # Only update and rerun if value actually changed
    if selected_material_new and selected_material_new != st.session_state.selected_material:
        st.session_state.selected_material = selected_material_new
        st.rerun()

    selected_material = st.session_state.selected_material

    # Material Properties Display
    if selected_material:
        Fy = safe_scalar(df_mat.loc[selected_material, "Yield Point (ksc)"])
        Fu = safe_scalar(df_mat.loc[selected_material, "Tensile Strength (ksc)"])
        E = safe_scalar(df_mat.loc[selected_material, "E"])

        st.markdown(f"""
        <div class="info-box">
        <b>📋 Material: {selected_material}</b><br>
        • Fy = {Fy:.1f} kgf/cm²<br>
        • Fu = {Fu:.1f} kgf/cm²<br>
        • E = {E:,.0f} kgf/cm²
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 📐 Section Selection")

    # ========== SECTION SELECTION (CUSTOM HTML) ==========
    section_list = [str(x).strip() for x in df.index.tolist()]

    # Initialize
    if 'selected_section' not in st.session_state or not st.session_state.selected_section:
        st.session_state.selected_section = section_list[0]

    # Clean and validate
    st.session_state.selected_section = str(st.session_state.selected_section).strip()
    if st.session_state.selected_section not in section_list:
        st.session_state.selected_section = section_list[0]

    # Create options HTML
    section_options = ""
    for sec in section_list:
        selected_attr = 'selected="selected"' if sec == st.session_state.selected_section else ''
        section_options += f'<option value="{sec}" {selected_attr}>{sec}</option>\n'

    # Custom HTML dropdown for section
    st.markdown("**🔩 Select Section:**")
    section_html = f"""
    <script src="https://cdn.jsdelivr.net/npm/streamlit-component-lib@1.0.0/dist/streamlit-component-lib.js"></script>
    <select id="section_select" style="width: 100%; padding: 8px; font-size: 14px; border: 1px solid #ccc; border-radius: 4px;">
        {section_options}
    </select>
    <script>
        var select = document.getElementById('section_select');
        select.value = '{st.session_state.selected_section}';

        select.addEventListener('change', function() {{
            window.parent.Streamlit.setComponentValue(this.value);
        }});
    </script>
    """
    selected_section_new = components.html(section_html, height=50)

    # Only update and rerun if value actually changed
    if selected_section_new and selected_section_new != st.session_state.selected_section:
        st.session_state.selected_section = selected_section_new
        st.rerun()

    selected_section = st.session_state.selected_section

    # Section Properties Display
    if selected_section:
        weight_col = 'Unit Weight [kg/m]' if 'Unit Weight [kg/m]' in df.columns else 'w [kg/m]'

        # Get section properties
        weight = safe_scalar(df.loc[selected_section, weight_col])
        d = safe_scalar(df.loc[selected_section, 'd [mm]'])
        bf = safe_scalar(df.loc[selected_section, 'bf [mm]'])
        tf = safe_scalar(df.loc[selected_section, 'tf [mm]'])
        tw = safe_scalar(df.loc[selected_section, 'tw [mm]'])
        Zx = safe_scalar(df.loc[selected_section, 'Zx [cm3]'])
        Ix = safe_scalar(df.loc[selected_section, 'Ix [cm4]'])
        A = safe_scalar(df.loc[selected_section, 'A [cm2]'])

        st.markdown(f"""
        <div class="success-box">
        <h4 style="margin:0; color:#2E7D32;">✅ {selected_section}</h4>
        <hr style="margin:8px 0; border-color:#4caf50;">
        <b>Dimensions:</b><br>
        • d = {d:.0f} mm<br>
        • bf = {bf:.0f} mm<br>
        • tf = {tf:.1f} mm<br>
        • tw = {tw:.1f} mm<br>
        <b>Properties:</b><br>
        • A = {A:.1f} cm²<br>
        • Ix = {Ix:,.0f} cm⁴<br>
        • Zx = {Zx:,.0f} cm³<br>
        • Weight = {weight:.1f} kg/m
        </div>
        """, unsafe_allow_html=True)
    
    # ========== PROJECT INFORMATION ==========
    st.markdown("---")
    st.markdown("### 📁 Project Information")
    
    with st.expander("📝 Edit Project Info", expanded=False):
        if 'project_info' not in st.session_state:
            st.session_state.project_info = {
                'project_name': '',
                'project_no': '',
                'designer': '',
                'checker': '',
                'date': datetime.now().strftime('%Y-%m-%d'),
                'revision': '0'
            }
        
        st.session_state.project_info['project_name'] = st.text_input(
            "Project Name:", 
            value=st.session_state.project_info['project_name'],
            key="proj_name"
        )
        st.session_state.project_info['project_no'] = st.text_input(
            "Project No.:", 
            value=st.session_state.project_info['project_no'],
            key="proj_no"
        )
        st.session_state.project_info['designer'] = st.text_input(
            "Designer:", 
            value=st.session_state.project_info['designer'],
            key="designer"
        )
        st.session_state.project_info['checker'] = st.text_input(
            "Checker:", 
            value=st.session_state.project_info['checker'],
            key="checker"
        )
        st.session_state.project_info['date'] = st.text_input(
            "Date:", 
            value=st.session_state.project_info['date'],
            key="date"
        )
        st.session_state.project_info['revision'] = st.text_input(
            "Revision:", 
            value=st.session_state.project_info['revision'],
            key="revision"
        )
    
    # Show current project info summary
    if st.session_state.project_info['project_name']:
        st.markdown(f"""
        <div class="metric-card">
        <b>📋 Current Project:</b><br>
        {st.session_state.project_info['project_name']}<br>
        <small>Rev. {st.session_state.project_info['revision']} | {st.session_state.project_info['date']}</small>
        </div>
        """, unsafe_allow_html=True)

# Create local variables for compatibility with tab code
selected_material = st.session_state.selected_material
selected_section = st.session_state.selected_section

# ==================== ENHANCED TABS WITH TAB 5 ====================
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Design Analysis",
    "📈 Section Comparison", 
    "📋 Design Evaluation & Export",
    "📚 Documentation",
    "📦 Load Import & Member Check"
])

# ==================== TAB 1: DESIGN ANALYSIS ====================
with tab1:
    st.markdown('<h2 class="section-header">📊 Comprehensive Design Analysis</h2>', unsafe_allow_html=True)
    
    if st.session_state.selected_section and selected_material:
        section = st.session_state.selected_section
        
        # Analysis Type Selection
        analysis_type = st.radio(
            "Select Analysis Type:",
            ["Flexural Design (F2)", "Column Design (E3)", "Beam-Column (H1)"],
            horizontal=True
        )
        
        st.markdown("---")
        
        if analysis_type == "Flexural Design (F2)":
            col1, col2 = st.columns([1, 2])
            
            with col1:
                st.markdown("### Input Parameters")
                Lb = st.slider("Unbraced Length Lb (m):", 0.1, 20.0, 3.0, 0.1)
                Cb = st.number_input("Cb Factor:", 1.0, 2.3, 1.0, 0.1)
                
                result = aisc_360_16_f2_flexural_design(df, df_mat, section, selected_material, Lb, Cb)
                
                if result:
                    st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                    st.metric("φMn", f"{0.9*result['Mn']:.2f} t·m")
                    st.metric("Design Case", result['Case'])
                    st.markdown('</div>', unsafe_allow_html=True)
                    
                    st.markdown(f"""
                    <div class="critical-lengths-box">
                    <b>Critical Lengths:</b><br>
                    Lp = {result['Lp']:.3f} m<br>
                    Lr = {result['Lr']:.3f} m
                    </div>
                    """, unsafe_allow_html=True)
            
            with col2:
                if result:
                    # Generate capacity curve
                    Lb_points = np.linspace(0.1, 15, 200)
                    Mn_points = []
                    
                    for lb in Lb_points:
                        r = aisc_360_16_f2_flexural_design(df, df_mat, section, selected_material, lb, Cb)
                        Mn_points.append(0.9 * r['Mn'] if r else 0)
                    
                    fig = go.Figure()
                    
                    fig.add_trace(go.Scatter(
                        x=Lb_points, y=Mn_points,
                        mode='lines',
                        name='φMn',
                        line=dict(color='#667eea', width=3),
                        hovertemplate='Lb: %{x:.2f}m<br>φMn: %{y:.2f} t·m<extra></extra>'
                    ))
                    
                    fig.add_trace(go.Scatter(
                        x=[Lb], y=[0.9*result['Mn']],
                        mode='markers',
                        name='Design Point',
                        marker=dict(color='#f44336', size=15, symbol='star')
                    ))
                    
                    fig.add_vline(x=result['Lp'], line_dash="dash", line_color='#4caf50', line_width=2,
                                annotation_text=f"Lp={result['Lp']:.2f}m")
                    fig.add_vline(x=result['Lr'], line_dash="dash", line_color='#ff9800', line_width=2,
                                annotation_text=f"Lr={result['Lr']:.2f}m")
                    
                    layout = get_enhanced_plotly_layout()
                    layout['title'] = f"AISC F2: Flexural Analysis - {section}"
                    layout['xaxis']['title'] = "Unbraced Length, Lb (m)"
                    layout['yaxis']['title'] = "Design Moment, φMn (t·m)"
                    layout['height'] = 600
                    
                    fig.update_layout(layout)
                    st.plotly_chart(fig, use_container_width=True, config=create_enhanced_plotly_config())
        
        elif analysis_type == "Column Design (E3)":
            col1, col2 = st.columns([1, 2])
            
            with col1:
                st.markdown("### Input Parameters")
                KL = st.slider("Effective Length KL (m):", 0.1, 10.0, 3.0, 0.1)
                Pu = st.number_input("Applied Load Pu (tons):", 0.0, 500.0, 100.0, 10.0)
                
                comp_result = aisc_360_16_e3_compression_design(df, df_mat, section, selected_material, KL, KL)
                
                if comp_result:
                    st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                    st.metric("φPn", f"{comp_result['phi_Pn']:.2f} tons")
                    st.metric("Buckling Mode", comp_result['buckling_mode'])
                    ratio = Pu / comp_result['phi_Pn']
                    st.metric("Utilization", f"{ratio:.3f}", 
                             delta="✓ PASS" if ratio <= 1.0 else "✗ FAIL")
                    st.markdown('</div>', unsafe_allow_html=True)
            
            with col2:
                if comp_result:
                    # Generate capacity curve
                    lambda_points = np.linspace(1, 250, 250)
                    Pn_points = []
                    
                    Fy = safe_scalar(df_mat.loc[selected_material, "Yield Point (ksc)"])
                    E = safe_scalar(df_mat.loc[selected_material, "E"])
                    Ag = safe_scalar(df.loc[section, 'A [cm2]'])
                    lambda_limit = 4.71 * safe_sqrt(E / Fy)
                    
                    for lc in lambda_points:
                        Fe = (math.pi**2 * E) / (lc**2)
                        if lc <= lambda_limit:
                            Fcr = (0.658**(Fy/Fe)) * Fy
                        else:
                            Fcr = 0.877 * Fe
                        Pn_points.append(0.9 * Fcr * Ag / 1000.0)
                    
                    fig = go.Figure()
                    
                    fig.add_trace(go.Scatter(
                        x=lambda_points, y=Pn_points,
                        mode='lines',
                        name='φPn',
                        line=dict(color='#2196f3', width=3)
                    ))
                    
                    fig.add_trace(go.Scatter(
                        x=[comp_result['lambda_c']], y=[comp_result['phi_Pn']],
                        mode='markers',
                        name='Design Point',
                        marker=dict(color='#f44336', size=15, symbol='star')
                    ))
                    
                    fig.add_vline(x=lambda_limit, line_dash="dash", line_color='#ff9800', line_width=2,
                                annotation_text=f"λ limit={lambda_limit:.1f}")
                    
                    if Pu > 0:
                        fig.add_hline(y=Pu, line_dash="dash", line_color='#4caf50', line_width=2,
                                    annotation_text=f"Pu={Pu:.1f} tons")
                    
                    layout = get_enhanced_plotly_layout()
                    layout['title'] = f"AISC E3: Column Capacity - {section}"
                    layout['xaxis']['title'] = "Slenderness Ratio (KL/r)"
                    layout['yaxis']['title'] = "Design Strength, φPn (tons)"
                    layout['height'] = 600
                    
                    fig.update_layout(layout)
                    st.plotly_chart(fig, use_container_width=True, config=create_enhanced_plotly_config())
        
        else:  # Beam-Column H1
            col1, col2 = st.columns([1, 2])
            
            with col1:
                st.markdown("### Input Parameters")
                Pu_bc = st.number_input("Axial Load Pu (tons):", 0.0, 200.0, 50.0, 5.0, key="h1_pu")
                Mux = st.number_input("Moment Mux (t·m):", 0.0, 100.0, 30.0, 5.0, key="h1_mux")
                KL_bc = st.slider("Effective Length KL (m):", 0.1, 10.0, 3.0, 0.1, key="h1_kl")
                Lb_bc = st.slider("Unbraced Length Lb (m):", 0.1, 10.0, 3.0, 0.1, key="h1_lb")
                
                comp_result = aisc_360_16_e3_compression_design(df, df_mat, section, selected_material, KL_bc, KL_bc)
                flex_result = aisc_360_16_f2_flexural_design(df, df_mat, section, selected_material, Lb_bc)
                
                if comp_result and flex_result:
                    phi_Pn = comp_result['phi_Pn']
                    phi_Mnx = 0.9 * flex_result['Mn']
                    
                    Zy = safe_scalar(df.loc[section, 'Zy [cm3]'])
                    Fy = safe_scalar(df_mat.loc[selected_material, "Yield Point (ksc)"])
                    phi_Mny = 0.9 * 0.9 * Fy * Zy / 100000.0
                    
                    interaction_result = aisc_360_16_h1_interaction(Pu_bc, phi_Pn, Mux, phi_Mnx, 0, phi_Mny)
                    
                    if interaction_result:
                        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                        st.metric("Unity Check", f"{interaction_result['interaction_ratio']:.3f}")
                        st.metric("Equation", interaction_result['equation'])
                        status = "✓ PASS" if interaction_result['design_ok'] else "✗ FAIL"
                        st.metric("Status", status)
                        st.markdown('</div>', unsafe_allow_html=True)
            
            with col2:
                if comp_result and flex_result and interaction_result:
                    # P-M Interaction Diagram
                    M_ratios = np.linspace(0, 1.2, 100)
                    P_ratios_h1a = []
                    P_ratios_h1b = []
                    
                    for m in M_ratios:
                        # H1-1a (Pr/Pc >= 0.2)
                        p_a = 1.0 - (9.0/8.0) * m
                        P_ratios_h1a.append(max(0, p_a))
                        
                        # H1-1b (Pr/Pc < 0.2)
                        p_b = 2.0 * (1.0 - m)
                        P_ratios_h1b.append(max(0, min(0.2, p_b)))
                    
                    fig = go.Figure()
                    
                    fig.add_trace(go.Scatter(
                        x=M_ratios, y=P_ratios_h1a,
                        mode='lines',
                        name='H1-1a (Pr/Pc ≥ 0.2)',
                        line=dict(color='#2196f3', width=3),
                        fill='tozeroy',
                        fillcolor='rgba(33, 150, 243, 0.1)'
                    ))
                    
                    fig.add_trace(go.Scatter(
                        x=M_ratios, y=P_ratios_h1b,
                        mode='lines',
                        name='H1-1b (Pr/Pc < 0.2)',
                        line=dict(color='#4caf50', width=3),
                        fill='tozeroy',
                        fillcolor='rgba(76, 175, 80, 0.1)'
                    ))
                    
                    M_combined = interaction_result['Mrx_Mcx']
                    P_ratio = interaction_result['Pr_Pc']
                    
                    fig.add_trace(go.Scatter(
                        x=[M_combined], y=[P_ratio],
                        mode='markers',
                        name='Design Point',
                        marker=dict(color='#f44336', size=20, symbol='star')
                    ))
                    
                    layout = get_enhanced_plotly_layout()
                    layout['title'] = f"AISC H1: P-M Interaction - {section}"
                    layout['xaxis']['title'] = "Moment Ratio (Mr/Mc)"
                    layout['yaxis']['title'] = "Axial Ratio (Pr/Pc)"
                    layout['height'] = 600
                    
                    fig.update_layout(layout)
                    st.plotly_chart(fig, use_container_width=True, config=create_enhanced_plotly_config())
    else:
        st.warning("⚠️ Please select a section from the sidebar")

# ==================== TAB 2: SECTION COMPARISON ====================
with tab2:
    st.markdown('<h2 class="section-header">📈 Advanced Section Comparison</h2>', unsafe_allow_html=True)
    
    # Selection interface
    col1, col2, col3 = st.columns(3)
    
    with col1:
        sections_to_compare = st.multiselect(
            "Select Sections to Compare:",
            list(df.index),
            default=list(df.index)[:5] if len(df.index) >= 5 else list(df.index)
        )
    
    with col2:
        compare_Lb = st.slider("Lb for Comparison (m):", 0.1, 15.0, 3.0, 0.5)
    
    with col3:
        compare_KL = st.slider("KL for Comparison (m):", 0.1, 10.0, 3.0, 0.5)
    
    if sections_to_compare and selected_material:
        comparison_data = []
        
        for sec in sections_to_compare:
            try:
                weight_col = 'Unit Weight [kg/m]' if 'Unit Weight [kg/m]' in df.columns else 'w [kg/m]'
                weight = safe_scalar(df.loc[sec, weight_col])
                
                flex_result = aisc_360_16_f2_flexural_design(df, df_mat, sec, selected_material, compare_Lb)
                comp_result = aisc_360_16_e3_compression_design(df, df_mat, sec, selected_material, compare_KL, compare_KL)
                
                if flex_result and comp_result:
                    comparison_data.append({
                        'Section': sec,
                        'Weight (kg/m)': weight,
                        'φMn (t·m)': 0.9 * flex_result['Mn'],
                        'φPn (tons)': comp_result['phi_Pn'],
                        'Lp (m)': flex_result['Lp'],
                        'Lr (m)': flex_result['Lr'],
                        'Moment Efficiency': (0.9 * flex_result['Mn']) / weight,
                        'Compression Efficiency': comp_result['phi_Pn'] / weight
                    })
            except:
                continue
        
        if comparison_data:
            df_comparison = pd.DataFrame(comparison_data)
            
            # Enhanced Comparison Chart
            fig = make_subplots(
                rows=2, cols=2,
                subplot_titles=('Moment Capacity', 'Compression Capacity',
                               'Moment Efficiency', 'Weight Comparison'),
                specs=[[{"type": "bar"}, {"type": "bar"}],
                       [{"type": "bar"}, {"type": "bar"}]]
            )
            
            # Moment capacity
            fig.add_trace(go.Bar(
                x=df_comparison['Section'],
                y=df_comparison['φMn (t·m)'],
                marker_color='#667eea',
                name='φMn',
                showlegend=False
            ), row=1, col=1)
            
            # Compression capacity
            fig.add_trace(go.Bar(
                x=df_comparison['Section'],
                y=df_comparison['φPn (tons)'],
                marker_color='#2196f3',
                name='φPn',
                showlegend=False
            ), row=1, col=2)
            
            # Moment efficiency
            fig.add_trace(go.Bar(
                x=df_comparison['Section'],
                y=df_comparison['Moment Efficiency'],
                marker_color='#4caf50',
                name='Moment Eff.',
                showlegend=False
            ), row=2, col=1)
            
            # Weight
            fig.add_trace(go.Bar(
                x=df_comparison['Section'],
                y=df_comparison['Weight (kg/m)'],
                marker_color='#ff9800',
                name='Weight',
                showlegend=False
            ), row=2, col=2)
            
            layout = get_enhanced_plotly_layout()
            layout['height'] = 800
            layout['title'] = "AISC 360-16 Multi-Section Comparison"
            fig.update_layout(layout)
            
            st.plotly_chart(fig, use_container_width=True, config=create_enhanced_plotly_config())
            
            # Enhanced Comparison Table
            st.markdown("### 📊 Detailed Comparison Table")
            
            # Style the dataframe
            styled_df = df_comparison.style.format({
                'Weight (kg/m)': '{:.1f}',
                'φMn (t·m)': '{:.2f}',
                'φPn (tons)': '{:.2f}',
                'Lp (m)': '{:.3f}',
                'Lr (m)': '{:.3f}',
                'Moment Efficiency': '{:.3f}',
                'Compression Efficiency': '{:.3f}'
            }).background_gradient(cmap='Blues', subset=['φMn (t·m)', 'φPn (tons)'])
            
            st.dataframe(styled_df, use_container_width=True, height=400)
    else:
        st.warning("⚠️ Please select sections to compare")

# ==================== TAB 3: DESIGN EVALUATION & EXPORT ====================
with tab3:
    st.markdown('<h2 class="section-header">📋 Design Evaluation & Export</h2>', unsafe_allow_html=True)
    
    if st.session_state.selected_section and selected_material:
        section = st.session_state.selected_section
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### Design Parameters")
            Mu_eval = st.number_input("Design Moment Mu (t·m):", 0.0, 200.0, 50.0, 5.0, key="eval_mu")
            Pu_eval = st.number_input("Axial Load Pu (tons):", 0.0, 500.0, 100.0, 10.0, key="eval_pu")
        
        with col2:
            st.markdown("### Design Lengths")
            Lb_eval = st.number_input("Unbraced Length Lb (m):", 0.1, 20.0, 3.0, 0.1, key="eval_lb")
            KL_eval = st.number_input("Effective Length KL (m):", 0.1, 20.0, 3.0, 0.1, key="eval_kl")
        
        if st.button("🔍 Perform Comprehensive Evaluation", type="primary"):
            design_loads = {'Mu': Mu_eval, 'Pu': Pu_eval}
            design_lengths = {'Lb': Lb_eval, 'KLx': KL_eval, 'KLy': KL_eval}
            
            evaluation = evaluate_section_design(df, df_mat, section, selected_material, design_loads, design_lengths)
            
            if evaluation:
                st.session_state.evaluation_results = evaluation
                
                # Overall Status
                if evaluation['design_check']['overall_adequate']:
                    st.markdown(f"""
                    <div class="success-box">
                    <h3>✅ DESIGN STATUS: ADEQUATE</h3>
                    <p><b>Moment Utilization:</b> {evaluation['flexural']['ratio']:.3f}</p>
                    <p><b>Axial Utilization:</b> {evaluation['compression']['ratio']:.3f}</p>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div class="error-box">
                    <h3>❌ DESIGN STATUS: INADEQUATE</h3>
                    <p><b>Moment Utilization:</b> {evaluation['flexural']['ratio']:.3f}</p>
                    <p><b>Axial Utilization:</b> {evaluation['compression']['ratio']:.3f}</p>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Detailed Results
                col_r1, col_r2 = st.columns(2)
                
                with col_r1:
                    st.markdown('<div class="design-summary">', unsafe_allow_html=True)
                    st.markdown("#### Flexural Analysis (F2)")
                    st.metric("φMn", f"{evaluation['flexural']['phi_Mn']:.2f} t·m")
                    st.metric("Utilization", f"{evaluation['flexural']['ratio']:.3f}")
                    st.write(f"**Zone:** {evaluation['flexural']['zone']}")
                    st.write(f"**Lp:** {evaluation['flexural']['Lp']:.3f} m")
                    st.write(f"**Lr:** {evaluation['flexural']['Lr']:.3f} m")
                    st.markdown('</div>', unsafe_allow_html=True)
                
                with col_r2:
                    st.markdown('<div class="design-summary">', unsafe_allow_html=True)
                    st.markdown("#### Compression Analysis (E3)")
                    st.metric("φPn", f"{evaluation['compression']['phi_Pn']:.2f} tons")
                    st.metric("Utilization", f"{evaluation['compression']['ratio']:.3f}")
                    st.write(f"**Mode:** {evaluation['compression']['mode']}")
                    st.write(f"**λc:** {evaluation['compression']['lambda_c']:.1f}")
                    st.write(f"**Fcr:** {evaluation['compression']['Fcr']:.1f} kgf/cm²")
                    st.markdown('</div>', unsafe_allow_html=True)
        
        # Export Section
        col_export1, col_export2 = st.columns(2)
        
        with col_export1:
            if PDF_AVAILABLE:
                if st.button("📄 Generate Calculation Report", type="primary"):
                    design_params = {
                        'Mu': Mu_eval, 'Pu': Pu_eval,
                        'Lb': Lb_eval, 'KL': KL_eval,
                        'Cb': 1.0
                    }
                    
                    # Get project info from session state
                    project_info = st.session_state.get('project_info', {
                        'project_name': 'N/A',
                        'project_no': 'N/A',
                        'designer': 'N/A',
                        'checker': 'N/A',
                        'date': datetime.now().strftime('%Y-%m-%d'),
                        'revision': '0'
                    })
                    
                    with st.spinner('Generating calculation report...'):
                        pdf_buffer = generate_calculation_report(
                            df, df_mat, section, selected_material, 
                            st.session_state.evaluation_results, design_params, project_info
                        )
                    
                    if pdf_buffer:
                        # Create filename with project info
                        proj_name = project_info.get('project_name', 'Project').replace(' ', '_')[:20]
                        filename = f"Calc_{proj_name}_{section}_{datetime.now().strftime('%Y%m%d')}.pdf"
                        
                        st.download_button(
                            label="📥 Download Calculation Report (PDF)",
                            data=pdf_buffer,
                            file_name=filename,
                            mime="application/pdf"
                        )
                        st.success("✅ calculation report generated!")
            else:
                st.warning("⚠️ PDF export requires reportlab library")
                st.code("pip install reportlab")
        
        # Excel Export
        with col_export2:
            if EXCEL_AVAILABLE:
                if st.button("📊 Generate Enhanced Excel Report", type="primary"):
                    design_params = {
                        'Mu': Mu_eval, 'Pu': Pu_eval,
                        'Lb': Lb_eval, 'KL': KL_eval,
                        'Cb': 1.0
                    }
                    
                    with st.spinner('Generating enhanced Excel report...'):
                        excel_buffer = generate_enhanced_excel_report(
                            df, df_mat, section, selected_material, 
                            st.session_state.evaluation_results, design_params
                        )
                    
                    if excel_buffer:
                        st.download_button(
                            label="📥 Download Enhanced Excel Report",
                            data=excel_buffer,
                            file_name=f"AISC_Enhanced_Report_{section}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success("✅ Enhanced Excel report with detailed calculations generated successfully!")
            else:
                st.warning("⚠️ Excel export requires openpyxl library")
                st.code("pip install openpyxl")
    else:
        st.info("ℹ️ Select a section and material first, then perform evaluation to enable export functionality")

# ==================== TAB 4: DOCUMENTATION ====================
with tab4:
    st.markdown('<h2 class="section-header">📚 User Guide & AISC References</h2>', unsafe_allow_html=True)
    
    with st.expander("🎯 Application Features", expanded=True):
        st.markdown("""
        ### UI/UX Enhancements v7.0
        
        **Visual Improvements:**
        - 🎨 Modern gradient color scheme with typography
        - 📊 Enhanced data tables with improved readability and alignment
        - 📈 Optimized charts with proper margins and label placement
        - 🎯 Clean, intuitive layout with logical information flow
        
        **Export Capabilities:**
        - 📄 **PDF Export:** Comprehensive calculation reports with formatting
        - 📊 **Excel Export:** Detailed analysis with formatted tables and styling
        - 📋 Both formats include input parameters, step-by-step calculations, and results
        
        **Analysis Tools:**
        - Flexural design per AISC 360-16 Chapter F2
        - Column design per AISC 360-16 Chapter E3
        - Beam-column interaction per AISC 360-16 Chapter H1
        - Multi-section comparison with efficiency metrics
        """)
    
    with st.expander("📖 AISC 360-16 References"):
        st.markdown("""
        ### Design Specifications
        
        **Chapter F2 - Lateral-Torsional Buckling:**
        - F2.1: Yielding (Lb ≤ Lp)
        - F2.2: Inelastic LTB (Lp < Lb ≤ Lr)
        - F2.3: Elastic LTB (Lb > Lr)
        
        **Chapter E3 - Flexural Buckling:**
        - E3.2(a): Inelastic buckling (λc ≤ 4.71√(E/Fy))
        - E3.2(b): Elastic buckling (λc > 4.71√(E/Fy))
        
        **Chapter H1 - Combined Forces:**
        - H1-1a: Pr/Pc ≥ 0.2
        - H1-1b: Pr/Pc < 0.2
        """)
    
    with st.expander("🚀 Quick Start Guide"):
        st.markdown("""
        ### Getting Started
        
        1. **Select Material:** Choose steel grade from sidebar
        2. **Select Section:** Pick steel section from dropdown
        3. **Choose Analysis:** Select design type in main tabs
        4. **Input Parameters:** Enter loads and lengths
        5. **View Results:** Analyze charts and metrics
        6. **Export Report:** Generate PDF or Excel documentation
        
        ### Export Instructions
        
        **PDF Reports:**
        - formatting with AISC compliance
        - Includes material properties, section properties, and analysis results
        - Download directly from the browser
        
        **Excel Reports:**
        - Formatted spreadsheets with multiple sheets
        - Color-coded tables for easy reading
        - Can be edited and customized after export
        """)
# ==================== ENHANCED TAB 5: LOAD IMPORT & MEMBER-BASED DESIGN CHECK ====================
# Version: 2.0 - Improved Member Grouping, Multiple Member Types, Batch Analysis
# Features: CSV/Excel Import, Member Groups, Section Assignment, Comprehensive Design Checks

"""
INTEGRATION INSTRUCTIONS:
Replace the entire Tab 5 section in your main AISC application with the code below.
The code should replace everything between:
    with tab5:
        ...
    (until the next tab or end of tabs)

This enhanced Tab 5 includes:
1. Sub-Tab 5.1: Data Import - Upload and preview load data
2. Sub-Tab 5.2: Member Groups - Configure sections/materials for each member
3. Sub-Tab 5.3: Design Check - Run analysis for all load combinations
4. Sub-Tab 5.4: Summary Report - Export results and visualizations
"""

# ==================== TAB 5: LOAD IMPORT & MEMBER CHECK (ENHANCED) ====================
with tab5:
    st.markdown('<h2 class="section-header">📦 Load Data Import & Member-Based Design Check</h2>', unsafe_allow_html=True)
    
    # Initialize session state for Tab 5
    if 'loaded_data' not in st.session_state:
        st.session_state.loaded_data = None
    if 'member_groups' not in st.session_state:
        st.session_state.member_groups = {}
    if 'analysis_results_tab5' not in st.session_state:
        st.session_state.analysis_results_tab5 = {}
    
    # Create sub-tabs for organization
    subtab1, subtab2, subtab3, subtab4 = st.tabs([
        "📁 Data Import",
        "👥 Member Groups",
        "🔍 Design Check",
        "📊 Summary Report"
    ])
    
    # ==================== SUB-TAB 1: DATA IMPORT ====================
    with subtab1:
        st.markdown("### 📁 Upload Load Analysis Results")
        
        col_upload1, col_upload2 = st.columns([2, 1])
        
        with col_upload1:
            uploaded_file = st.file_uploader(
                "Upload CSV or Excel file with load data",
                type=['csv', 'xlsx', 'xls'],
                help="File should contain columns: Member No., Load Combination, Mu, Pu",
                key="tab5_file_uploader"
            )
        
        with col_upload2:
            st.markdown("""
            <div class="info-box">
            <b>📋 Required Columns:</b><br>
            • Member No. (ID)<br>
            • Load Combination<br>
            • Mu (t·m) – Moment<br>
            • Pu (tons) – Axial<br>
            <small>+Pu = Compression<br>-Pu = Tension</small>
            </div>
            """, unsafe_allow_html=True)
        
        # Download template
        template_data = """Member No.,Load Combination,Mu,Pu
B101,1,45.5,120.3
B101,2,52.8,135.7
B101,3,38.2,-98.5
B101,4,61.3,145.2
B101,5,55.0,128.0
B102,1,32.1,85.4
B102,2,28.9,-92.1
B102,3,41.5,78.3
B102,4,35.7,88.9
B102,5,39.2,81.2
C201,1,12.5,250.3
C201,2,15.8,285.7
C201,3,8.2,198.5
C201,4,18.3,315.2
C201,5,14.0,268.0
T301,1,0.0,-45.5
T301,2,0.0,-52.8
T301,3,0.0,-38.2
T301,4,0.0,-61.3
T301,5,0.0,-55.0"""
        
        st.download_button(
            label="📥 Download CSV Template",
            data=template_data,
            file_name="load_data_template.csv",
            mime="text/csv",
            key="tab5_template_download"
        )
        
        # Process uploaded file
        if uploaded_file is not None:
            try:
                # Read file based on type
                if uploaded_file.name.endswith('.csv'):
                    df_loads = pd.read_csv(uploaded_file)
                else:
                    df_loads = pd.read_excel(uploaded_file)
                
                # Validate required columns
                required_cols = ['Member No.', 'Load Combination', 'Mu', 'Pu']
                missing_cols = [col for col in required_cols if col not in df_loads.columns]
                
                if missing_cols:
                    st.error(f"❌ Missing required columns: {', '.join(missing_cols)}")
                    st.info("💡 Please ensure your file has columns: Member No., Load Combination, Mu, Pu")
                else:
                    # Clean and validate data
                    df_loads = df_loads.dropna(subset=required_cols)
                    df_loads['Member No.'] = df_loads['Member No.'].astype(str).str.strip()
                    df_loads['Load Combination'] = pd.to_numeric(df_loads['Load Combination'], errors='coerce').astype(int)
                    df_loads['Mu'] = pd.to_numeric(df_loads['Mu'], errors='coerce')
                    df_loads['Pu'] = pd.to_numeric(df_loads['Pu'], errors='coerce')
                    
                    # Store in session state
                    st.session_state.loaded_data = df_loads
                    
                    # Success message
                    st.success(f"✅ Successfully loaded {len(df_loads)} load cases for {df_loads['Member No.'].nunique()} members")
                    
                    # Summary statistics
                    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                    with col_stat1:
                        st.metric("Total Load Cases", len(df_loads))
                    with col_stat2:
                        st.metric("Unique Members", df_loads['Member No.'].nunique())
                    with col_stat3:
                        compression_count = len(df_loads[df_loads['Pu'] > 0])
                        st.metric("Compression Cases", compression_count)
                    with col_stat4:
                        tension_count = len(df_loads[df_loads['Pu'] < 0])
                        st.metric("Tension Cases", tension_count)
                    
                    # Preview data
                    with st.expander("👁️ Preview Raw Load Data", expanded=True):
                        st.dataframe(df_loads, use_container_width=True, height=300)
                    
                    # Member summary
                    with st.expander("📊 Member Summary", expanded=True):
                        member_summary = df_loads.groupby('Member No.').agg({
                            'Load Combination': 'count',
                            'Mu': ['min', 'max'],
                            'Pu': ['min', 'max']
                        }).round(2)
                        member_summary.columns = ['# LCs', 'Mu_min', 'Mu_max', 'Pu_min', 'Pu_max']
                        member_summary = member_summary.reset_index()
                        
                        # Add load type indicator
                        member_summary['Load Type'] = member_summary.apply(
                            lambda x: '🔴 Compression' if x['Pu_min'] >= 0 else ('🔵 Tension' if x['Pu_max'] <= 0 else '🟡 Mixed'),
                            axis=1
                        )
                        
                        st.dataframe(member_summary, use_container_width=True)
                    
                    # Auto-create member groups button
                    if st.button("🔄 Auto-Create Member Groups", type="primary", key="auto_create_groups"):
                        members = sorted(df_loads['Member No.'].unique())
                        for member in members:
                            if member not in st.session_state.member_groups:
                                # Determine default member type based on load pattern
                                member_data = df_loads[df_loads['Member No.'] == member]
                                has_compression = (member_data['Pu'] > 0).any()
                                has_tension = (member_data['Pu'] < 0).any()
                                has_moment = (member_data['Mu'].abs() > 0.1).any()
                                
                                if not has_moment and has_tension and not has_compression:
                                    default_type = "Tension Member"
                                elif has_compression and has_moment:
                                    default_type = "Beam-Column (Compression)"
                                elif has_tension and has_moment:
                                    default_type = "Beam-Column (Tension)"
                                else:
                                    default_type = "Beam (Flexure Only)"
                                
                                st.session_state.member_groups[member] = {
                                    'section': list(df.index)[0] if len(df.index) > 0 else None,
                                    'material': list(df_mat.index)[0] if len(df_mat.index) > 0 else None,
                                    'member_type': default_type,
                                    'Lb': 3.0,
                                    'KL': 3.0,
                                    'Cb': 1.0
                                }
                        st.success(f"✅ Created {len(members)} member groups. Go to 'Member Groups' tab to configure.")
                        st.rerun()
                    
            except Exception as e:
                st.error(f"❌ Error processing file: {str(e)}")
                st.info("💡 Please ensure your file matches the required format")
        
        else:
            st.info("📁 Upload a load data file to begin member-based design checks")
            
            # Show example format
            with st.expander("📖 Example Input Format", expanded=True):
                example_data = {
                    'Member No.': ['B101', 'B101', 'B101', 'C201', 'C201', 'T301', 'T301'],
                    'Load Combination': [1, 2, 3, 1, 2, 1, 2],
                    'Mu': [45.5, 52.8, 38.2, 12.5, 15.8, 0.0, 0.0],
                    'Pu': [120.3, 135.7, -98.5, 250.3, 285.7, -45.5, -52.8]
                }
                st.dataframe(pd.DataFrame(example_data), use_container_width=True)
                
                st.markdown("""
                <div class="info-box">
                <b>📋 Sign Convention:</b><br>
                • <b>+Pu (Positive):</b> Compression<br>
                • <b>-Pu (Negative):</b> Tension<br>
                • <b>Mu:</b> Absolute value of bending moment
                </div>
                """, unsafe_allow_html=True)
    
    # ==================== SUB-TAB 2: MEMBER GROUPS ====================
    with subtab2:
        st.markdown("### 👥 Member Groups Configuration")
        
        if st.session_state.loaded_data is None:
            st.warning("⚠️ Please upload load data first in the 'Data Import' tab")
        else:
            df_loads = st.session_state.loaded_data
            members = sorted(df_loads['Member No.'].unique())
            
            # Batch configuration
            st.markdown("#### 🔧 Batch Configuration")
            with st.expander("Apply Settings to Multiple Members", expanded=False):
                col_batch1, col_batch2 = st.columns(2)
                
                with col_batch1:
                    selected_members_batch = st.multiselect(
                        "Select Members:",
                        members,
                        default=[],
                        key="batch_member_select"
                    )
                    
                    batch_section = st.selectbox(
                        "Section:",
                        list(df.index),
                        index=0,
                        key="batch_section"
                    )
                    
                    batch_material = st.selectbox(
                        "Material:",
                        list(df_mat.index),
                        index=0,
                        key="batch_material"
                    )
                
                with col_batch2:
                    batch_type = st.selectbox(
                        "Member Type:",
                        ["Beam-Column (Compression)", "Beam-Column (Tension)", 
                         "Beam (Flexure Only)", "Tension Member"],
                        key="batch_type"
                    )
                    
                    batch_Lb = st.number_input("Lb (m):", 0.1, 20.0, 3.0, 0.1, key="batch_lb")
                    batch_KL = st.number_input("KL (m):", 0.1, 20.0, 3.0, 0.1, key="batch_kl")
                    batch_Cb = st.number_input("Cb:", 1.0, 3.0, 1.0, 0.1, key="batch_cb")
                
                if st.button("✅ Apply to Selected Members", key="apply_batch"):
                    for member in selected_members_batch:
                        st.session_state.member_groups[member] = {
                            'section': batch_section,
                            'material': batch_material,
                            'member_type': batch_type,
                            'Lb': batch_Lb,
                            'KL': batch_KL,
                            'Cb': batch_Cb
                        }
                    st.success(f"✅ Applied settings to {len(selected_members_batch)} members")
                    st.rerun()
            
            st.markdown("---")
            
            # Individual member configuration
            st.markdown("#### 📝 Individual Member Editor")
            
            if len(st.session_state.member_groups) == 0:
                st.info("💡 No member groups configured yet. Click 'Auto-Create Member Groups' in the Data Import tab or add members below.")
            
            # Member editor
            selected_member_edit = st.selectbox(
                "Select Member to Edit:",
                members,
                key="member_edit_select"
            )
            
            if selected_member_edit:
                # Get current config or create default
                current_config = st.session_state.member_groups.get(selected_member_edit, {
                    'section': list(df.index)[0],
                    'material': list(df_mat.index)[0],
                    'member_type': "Beam-Column (Compression)",
                    'Lb': 3.0,
                    'KL': 3.0,
                    'Cb': 1.0
                })
                
                col_edit1, col_edit2, col_edit3 = st.columns(3)
                
                with col_edit1:
                    edit_section = st.selectbox(
                        "Section:",
                        list(df.index),
                        index=list(df.index).index(current_config['section']) if current_config['section'] in df.index else 0,
                        key="edit_section"
                    )
                    
                    edit_material = st.selectbox(
                        "Material:",
                        list(df_mat.index),
                        index=list(df_mat.index).index(current_config['material']) if current_config['material'] in df_mat.index else 0,
                        key="edit_material"
                    )
                
                with col_edit2:
                    type_options = ["Beam-Column (Compression)", "Beam-Column (Tension)", 
                                   "Beam (Flexure Only)", "Tension Member"]
                    edit_type = st.selectbox(
                        "Member Type:",
                        type_options,
                        index=type_options.index(current_config['member_type']) if current_config['member_type'] in type_options else 0,
                        key="edit_type"
                    )
                    
                    edit_Lb = st.number_input("Lb (m):", 0.1, 20.0, current_config['Lb'], 0.1, key="edit_lb")
                
                with col_edit3:
                    edit_KL = st.number_input("KL (m):", 0.1, 20.0, current_config['KL'], 0.1, key="edit_kl")
                    edit_Cb = st.number_input("Cb:", 1.0, 3.0, current_config['Cb'], 0.1, key="edit_cb")
                
                if st.button("💾 Save Member Configuration", key="save_member_config"):
                    st.session_state.member_groups[selected_member_edit] = {
                        'section': edit_section,
                        'material': edit_material,
                        'member_type': edit_type,
                        'Lb': edit_Lb,
                        'KL': edit_KL,
                        'Cb': edit_Cb
                    }
                    st.success(f"✅ Saved configuration for Member {selected_member_edit}")
                
                # Show member load data
                with st.expander(f"📊 Load Data for Member {selected_member_edit}", expanded=True):
                    member_data = df_loads[df_loads['Member No.'] == selected_member_edit]
                    st.dataframe(member_data, use_container_width=True)
            
            # Member groups summary table
            st.markdown("---")
            st.markdown("#### 📋 Member Groups Summary")
            
            if len(st.session_state.member_groups) > 0:
                summary_data = []
                for member, config in st.session_state.member_groups.items():
                    member_loads = df_loads[df_loads['Member No.'] == member]
                    summary_data.append({
                        'Member': member,
                        'Section': config['section'],
                        'Material': config['material'],
                        'Type': config['member_type'].split('(')[0].strip(),
                        'Lb (m)': config['Lb'],
                        'KL (m)': config['KL'],
                        '# LCs': len(member_loads),
                        'Max |Mu|': member_loads['Mu'].abs().max(),
                        'Max Pu': member_loads['Pu'].max(),
                        'Min Pu': member_loads['Pu'].min()
                    })
                
                df_summary = pd.DataFrame(summary_data)
                st.dataframe(df_summary.style.format({
                    'Lb (m)': '{:.2f}',
                    'KL (m)': '{:.2f}',
                    'Max |Mu|': '{:.2f}',
                    'Max Pu': '{:.2f}',
                    'Min Pu': '{:.2f}'
                }), use_container_width=True, height=400)
            else:
                st.info("No member groups configured yet.")
    
    # ==================== SUB-TAB 3: DESIGN CHECK ====================
    with subtab3:
        st.markdown("### 🔍 Design Check Analysis")
        
        if st.session_state.loaded_data is None:
            st.warning("⚠️ Please upload load data first in the 'Data Import' tab")
        elif len(st.session_state.member_groups) == 0:
            st.warning("⚠️ Please configure member groups first in the 'Member Groups' tab")
        else:
            df_loads = st.session_state.loaded_data
            
            # Analysis options
            col_opt1, col_opt2 = st.columns(2)
            
            with col_opt1:
                analysis_scope = st.radio(
                    "Analysis Scope:",
                    ["Single Member", "All Configured Members"],
                    horizontal=True
                )
            
            with col_opt2:
                if analysis_scope == "Single Member":
                    configured_members = list(st.session_state.member_groups.keys())
                    if configured_members:
                        selected_analysis_member = st.selectbox(
                            "Select Member:",
                            configured_members,
                            key="analysis_member_select"
                        )
                    else:
                        st.warning("No members configured")
                        selected_analysis_member = None
            
            # Run Analysis Button
            if st.button("🚀 Run Design Analysis", type="primary", key="run_analysis"):
                
                # Determine which members to analyze
                if analysis_scope == "Single Member":
                    members_to_analyze = [selected_analysis_member] if selected_analysis_member else []
                else:
                    members_to_analyze = list(st.session_state.member_groups.keys())
                
                if len(members_to_analyze) == 0:
                    st.warning("⚠️ No members to analyze")
                else:
                    # Progress bar
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    all_results = {}
                    
                    for i, member in enumerate(members_to_analyze):
                        status_text.text(f"Analyzing Member {member}... ({i+1}/{len(members_to_analyze)})")
                        
                        config = st.session_state.member_groups[member]
                        member_loads = df_loads[df_loads['Member No.'] == member].copy()
                        
                        section = config['section']
                        material = config['material']
                        member_type = config['member_type']
                        Lb = config['Lb']
                        KL = config['KL']
                        Cb = config['Cb']
                        
                        results = []
                        
                        for idx, row in member_loads.iterrows():
                            lc = int(row['Load Combination'])
                            Mu = float(row['Mu'])
                            Pu = float(row['Pu'])
                            
                            try:
                                # Get material and section properties
                                Fy = safe_scalar(df_mat.loc[material, "Yield Point (ksc)"])
                                Ag = safe_scalar(df.loc[section, 'A [cm2]'])
                                Zx = safe_scalar(df.loc[section, 'Zx [cm3]'])
                                Zy = safe_scalar(df.loc[section, 'Zy [cm3]'])
                                
                                # Initialize results
                                phi_Pn = None
                                phi_Mn = None
                                ratio = None
                                status = None
                                equation = None
                                mode = None
                                
                                # ==================== TENSION MEMBER ====================
                                if member_type == "Tension Member":
                                    # Tension capacity: φPn = φ × Fy × Ag
                                    phi_Pn = 0.9 * Fy * Ag / 1000.0  # Convert to tons
                                    phi_Mn = 0.0
                                    
                                    if Pu >= 0:
                                        # Compression force on tension member - not allowed
                                        ratio = 999.0
                                        status = "⚠️ COMPRESSION"
                                        equation = "N/A"
                                        mode = "Invalid Load"
                                    else:
                                        # Tension check
                                        ratio = abs(Pu) / phi_Pn
                                        status = "✓ OK" if ratio <= 1.0 else "✗ NG"
                                        equation = "Pu/φPn"
                                        mode = "Tension Yielding"
                                
                                # ==================== BEAM (FLEXURE ONLY) ====================
                                elif member_type == "Beam (Flexure Only)":
                                    flex_result = aisc_360_16_f2_flexural_design(
                                        df, df_mat, section, material, Lb, Cb
                                    )
                                    
                                    if flex_result:
                                        phi_Mn = 0.9 * flex_result['Mn']
                                        phi_Pn = 0.0
                                        
                                        ratio = abs(Mu) / phi_Mn if phi_Mn > 0 else 999
                                        status = "✓ OK" if ratio <= 1.0 else "✗ NG"
                                        equation = "Mu/φMn"
                                        mode = flex_result['Case'].split('-')[1].strip() if '-' in flex_result['Case'] else flex_result['Case']
                                    else:
                                        ratio = 999
                                        status = "ERROR"
                                        equation = "N/A"
                                        mode = "Calc Error"
                                
                                # ==================== BEAM-COLUMN (COMPRESSION) ====================
                                elif member_type == "Beam-Column (Compression)":
                                    comp_result = aisc_360_16_e3_compression_design(
                                        df, df_mat, section, material, KL, KL
                                    )
                                    flex_result = aisc_360_16_f2_flexural_design(
                                        df, df_mat, section, material, Lb, Cb
                                    )
                                    
                                    if comp_result and flex_result:
                                        phi_Pn = comp_result['phi_Pn']
                                        phi_Mnx = 0.9 * flex_result['Mn']
                                        phi_Mny = 0.9 * 0.9 * Fy * Zy / 100000.0  # Minor axis
                                        phi_Mn = phi_Mnx
                                        
                                        if Pu < 0:
                                            # Tension case - use simplified linear interaction
                                            tension_capacity = 0.9 * Fy * Ag / 1000.0
                                            axial_ratio = abs(Pu) / tension_capacity
                                            moment_ratio = abs(Mu) / phi_Mnx
                                            ratio = axial_ratio + moment_ratio
                                            status = "✓ OK" if ratio <= 1.0 else "✗ NG"
                                            equation = "Tu/φTn + Mu/φMn"
                                            mode = "Tension+Flexure"
                                        else:
                                            # Compression case - AISC H1 interaction
                                            interaction_result = aisc_360_16_h1_interaction(
                                                abs(Pu), phi_Pn, abs(Mu), phi_Mnx, 0, phi_Mny
                                            )
                                            
                                            if interaction_result:
                                                ratio = interaction_result['interaction_ratio']
                                                status = "✓ OK" if interaction_result['design_ok'] else "✗ NG"
                                                equation = interaction_result['equation']
                                                mode = f"Comp+Flex ({comp_result['buckling_mode']})"
                                            else:
                                                ratio = 999
                                                status = "ERROR"
                                                equation = "N/A"
                                                mode = "Calc Error"
                                    else:
                                        ratio = 999
                                        status = "ERROR"
                                        equation = "N/A"
                                        mode = "Calc Error"
                                
                                # ==================== BEAM-COLUMN (TENSION) ====================
                                elif member_type == "Beam-Column (Tension)":
                                    flex_result = aisc_360_16_f2_flexural_design(
                                        df, df_mat, section, material, Lb, Cb
                                    )
                                    
                                    if flex_result:
                                        phi_Mn = 0.9 * flex_result['Mn']
                                        tension_capacity = 0.9 * Fy * Ag / 1000.0
                                        phi_Pn = tension_capacity
                                        
                                        if Pu >= 0:
                                            # Compression case - need compression check
                                            comp_result = aisc_360_16_e3_compression_design(
                                                df, df_mat, section, material, KL, KL
                                            )
                                            if comp_result:
                                                phi_Pn_comp = comp_result['phi_Pn']
                                                phi_Mny = 0.9 * 0.9 * Fy * Zy / 100000.0
                                                
                                                interaction_result = aisc_360_16_h1_interaction(
                                                    Pu, phi_Pn_comp, abs(Mu), phi_Mn, 0, phi_Mny
                                                )
                                                
                                                if interaction_result:
                                                    ratio = interaction_result['interaction_ratio']
                                                    status = "✓ OK" if interaction_result['design_ok'] else "✗ NG"
                                                    equation = interaction_result['equation']
                                                    mode = "Comp+Flex (Reversed)"
                                                    phi_Pn = phi_Pn_comp
                                                else:
                                                    ratio = 999
                                                    status = "ERROR"
                                                    equation = "N/A"
                                                    mode = "Calc Error"
                                            else:
                                                ratio = 999
                                                status = "ERROR"
                                                equation = "N/A"
                                                mode = "Calc Error"
                                        else:
                                            # Tension case - linear interaction
                                            axial_ratio = abs(Pu) / tension_capacity
                                            moment_ratio = abs(Mu) / phi_Mn
                                            ratio = axial_ratio + moment_ratio
                                            status = "✓ OK" if ratio <= 1.0 else "✗ NG"
                                            equation = "Tu/φTn + Mu/φMn"
                                            mode = "Tension+Flexure"
                                    else:
                                        ratio = 999
                                        status = "ERROR"
                                        equation = "N/A"
                                        mode = "Calc Error"
                                
                                results.append({
                                    'LC': lc,
                                    'Mu (t·m)': Mu,
                                    'Pu (tons)': Pu,
                                    'φPn (tons)': phi_Pn if phi_Pn else 0,
                                    'φMn (t·m)': phi_Mn if phi_Mn else 0,
                                    'Ratio': ratio if ratio else 999,
                                    'Status': status if status else "ERROR",
                                    'Equation': equation if equation else "N/A",
                                    'Mode': mode if mode else "Unknown"
                                })
                                
                            except Exception as e:
                                results.append({
                                    'LC': lc,
                                    'Mu (t·m)': Mu,
                                    'Pu (tons)': Pu,
                                    'φPn (tons)': 0,
                                    'φMn (t·m)': 0,
                                    'Ratio': 999,
                                    'Status': f"ERROR",
                                    'Equation': "N/A",
                                    'Mode': str(e)[:30]
                                })
                        
                        # Store results
                        df_results = pd.DataFrame(results)
                        all_results[member] = {
                            'results': df_results,
                            'config': config
                        }
                        
                        progress_bar.progress((i + 1) / len(members_to_analyze))
                    
                    status_text.text("✅ Analysis Complete!")
                    st.session_state.analysis_results_tab5 = all_results
                    st.success(f"✅ Completed analysis for {len(all_results)} members")
            
            # Display Results
            st.markdown("---")
            st.markdown("### 📊 Analysis Results")
            
            if len(st.session_state.analysis_results_tab5) > 0:
                # Results selector
                analyzed_members = list(st.session_state.analysis_results_tab5.keys())
                selected_result_member = st.selectbox(
                    "Select Member to View:",
                    analyzed_members,
                    key="result_member_select"
                )
                
                if selected_result_member:
                    member_result = st.session_state.analysis_results_tab5[selected_result_member]
                    df_result = member_result['results']
                    config = member_result['config']
                    
                    # Member info card
                    st.markdown(f"""
                    <div class="metric-card">
                    <h4>Member: {selected_result_member}</h4>
                    <p><b>Section:</b> {config['section']} | <b>Material:</b> {config['material']} | <b>Type:</b> {config['member_type']}</p>
                    <p><b>Lb:</b> {config['Lb']:.2f}m | <b>KL:</b> {config['KL']:.2f}m | <b>Cb:</b> {config['Cb']:.2f}</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Summary metrics
                    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
                    
                    # Find governing case
                    valid_results = df_result[df_result['Ratio'] < 900]
                    if len(valid_results) > 0:
                        governing_idx = valid_results['Ratio'].idxmax()
                        governing_lc = df_result.loc[governing_idx, 'LC']
                        governing_ratio = df_result.loc[governing_idx, 'Ratio']
                    else:
                        governing_lc = "N/A"
                        governing_ratio = 999
                    
                    passing = len(df_result[df_result['Status'].str.contains('✓', na=False)])
                    
                    with col_m1:
                        st.metric("Total LCs", len(df_result))
                    with col_m2:
                        st.metric("Passing", passing, delta=f"{passing/len(df_result)*100:.0f}%")
                    with col_m3:
                        st.metric("Governing LC", int(governing_lc) if governing_lc != "N/A" else "N/A")
                    with col_m4:
                        st.metric("Max Ratio", f"{governing_ratio:.3f}",
                                 delta="✓ OK" if governing_ratio <= 1.0 else "✗ NG")
                    
                    # Overall status
                    if governing_ratio <= 1.0:
                        st.markdown(f"""
                        <div class="success-box">
                        <h3>✅ SECTION ADEQUATE FOR ALL LOAD COMBINATIONS</h3>
                        <p><b>Maximum Strength Ratio:</b> {governing_ratio:.3f} (LC {int(governing_lc) if governing_lc != 'N/A' else 'N/A'})</p>
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.markdown(f"""
                        <div class="error-box">
                        <h3>❌ SECTION INADEQUATE</h3>
                        <p><b>Governing Load Combination:</b> LC {int(governing_lc) if governing_lc != 'N/A' else 'N/A'} with ratio {governing_ratio:.3f}</p>
                        <p><b>Recommendation:</b> Increase section size or check design assumptions</p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # Results table with styling
                    st.markdown("#### 📋 Detailed Results by Load Combination")
                    
                    def style_status(val):
                        if '✓' in str(val):
                            return 'background-color: #C8E6C9; color: #2E7D32; font-weight: bold'
                        elif '✗' in str(val):
                            return 'background-color: #FFCDD2; color: #C62828; font-weight: bold'
                        elif '⚠️' in str(val):
                            return 'background-color: #FFF9C4; color: #F57C00; font-weight: bold'
                        return ''
                    
                    def style_ratio(val):
                        try:
                            if float(val) <= 1.0:
                                return 'background-color: #E8F5E9'
                            elif float(val) < 900:
                                return 'background-color: #FFEBEE'
                            else:
                                return 'background-color: #FFF9C4'
                        except:
                            return ''
                    
                    styled_results = df_result.style.format({
                        'Mu (t·m)': '{:.2f}',
                        'Pu (tons)': '{:.2f}',
                        'φPn (tons)': '{:.2f}',
                        'φMn (t·m)': '{:.2f}',
                        'Ratio': '{:.3f}'
                    }).applymap(style_status, subset=['Status'])\
                      .applymap(style_ratio, subset=['Ratio'])
                    
                    st.dataframe(styled_results, use_container_width=True, height=350)
                    
                    # Strength ratio chart
                    st.markdown("#### 📊 Strength Ratio Chart")
                    
                    fig = go.Figure()
                    
                    # Filter valid ratios for chart
                    valid_for_chart = df_result[df_result['Ratio'] < 900].copy()
                    
                    if len(valid_for_chart) > 0:
                        colors = ['#4CAF50' if r <= 1.0 else '#F44336' for r in valid_for_chart['Ratio']]
                        
                        fig.add_trace(go.Bar(
                            x=valid_for_chart['LC'],
                            y=valid_for_chart['Ratio'],
                            marker_color=colors,
                            text=valid_for_chart['Ratio'].apply(lambda x: f'{x:.3f}'),
                            textposition='outside',
                            name='Strength Ratio',
                            hovertemplate='<b>LC %{x}</b><br>Ratio: %{y:.3f}<br>%{customdata}<extra></extra>',
                            customdata=valid_for_chart['Mode']
                        ))
                        
                        # Unity line
                        fig.add_hline(y=1.0, line_dash="dash", line_color='#FF9800', line_width=2,
                                     annotation_text="Unity (Ratio = 1.0)")
                        
                        layout = get_enhanced_plotly_layout()
                        layout['title'] = f"Strength Ratio - Member {selected_result_member} ({config['section']})"
                        layout['xaxis']['title'] = "Load Combination"
                        layout['yaxis']['title'] = "Strength Ratio"
                        layout['height'] = 450
                        
                        fig.update_layout(layout)
                        st.plotly_chart(fig, use_container_width=True, config=create_enhanced_plotly_config())
            else:
                st.info("💡 Run the design analysis to see results")
    
    # ==================== SUB-TAB 4: SUMMARY REPORT ====================
    with subtab4:
        st.markdown("### 📊 Summary Report & Export")
        
        if len(st.session_state.analysis_results_tab5) == 0:
            st.warning("⚠️ Please run design analysis first in the 'Design Check' tab")
        else:
            all_results = st.session_state.analysis_results_tab5
            
            # Overall summary
            st.markdown("#### 📋 Overall Design Summary")
            
            summary_rows = []
            for member, data in all_results.items():
                df_result = data['results']
                config = data['config']
                
                valid_results = df_result[df_result['Ratio'] < 900]
                if len(valid_results) > 0:
                    max_ratio = valid_results['Ratio'].max()
                    governing_idx = valid_results['Ratio'].idxmax()
                    governing_lc = df_result.loc[governing_idx, 'LC']
                    governing_mode = df_result.loc[governing_idx, 'Mode']
                else:
                    max_ratio = 999
                    governing_lc = "N/A"
                    governing_mode = "Error"
                
                passing = len(df_result[df_result['Status'].str.contains('✓', na=False)])
                
                summary_rows.append({
                    'Member': member,
                    'Section': config['section'],
                    'Type': config['member_type'].split('(')[0].strip(),
                    '# LCs': len(df_result),
                    'Passing': passing,
                    'Gov. LC': governing_lc,
                    'Max Ratio': max_ratio,
                    'Mode': governing_mode,
                    'Status': '✓ OK' if max_ratio <= 1.0 else '✗ NG'
                })
            
            df_summary = pd.DataFrame(summary_rows)
            
            # Statistics
            col_s1, col_s2, col_s3, col_s4 = st.columns(4)
            
            total_members = len(df_summary)
            passing_members = len(df_summary[df_summary['Max Ratio'] <= 1.0])
            failing_members = total_members - passing_members
            max_overall_ratio = df_summary['Max Ratio'].max()
            
            with col_s1:
                st.metric("Total Members", total_members)
            with col_s2:
                st.metric("Passing", passing_members, delta=f"{passing_members/total_members*100:.0f}%")
            with col_s3:
                st.metric("Failing", failing_members)
            with col_s4:
                st.metric("Max Ratio", f"{max_overall_ratio:.3f}" if max_overall_ratio < 900 else "N/A")
            
            # Summary table
            def style_summary_status(val):
                if '✓' in str(val):
                    return 'background-color: #C8E6C9; color: #2E7D32; font-weight: bold'
                elif '✗' in str(val):
                    return 'background-color: #FFCDD2; color: #C62828; font-weight: bold'
                return ''
            
            styled_summary = df_summary.style.format({
                'Max Ratio': '{:.3f}'
            }).applymap(style_summary_status, subset=['Status'])
            
            st.dataframe(styled_summary, use_container_width=True, height=400)
            
            # Visualization
            st.markdown("#### 📊 Summary Visualization")
            
            col_chart1, col_chart2 = st.columns(2)
            
            with col_chart1:
                # Bar chart of max ratios
                valid_summary = df_summary[df_summary['Max Ratio'] < 900]
                
                if len(valid_summary) > 0:
                    colors = ['#4CAF50' if r <= 1.0 else '#F44336' for r in valid_summary['Max Ratio']]
                    
                    fig_bar = go.Figure()
                    fig_bar.add_trace(go.Bar(
                        x=valid_summary['Member'],
                        y=valid_summary['Max Ratio'],
                        marker_color=colors,
                        text=valid_summary['Max Ratio'].apply(lambda x: f'{x:.2f}'),
                        textposition='outside'
                    ))
                    
                    fig_bar.add_hline(y=1.0, line_dash="dash", line_color='#FF9800', line_width=2)
                    
                    fig_bar.update_layout(
                        title="Maximum Strength Ratio by Member",
                        xaxis_title="Member",
                        yaxis_title="Max Ratio",
                        height=400,
                        template='plotly_white'
                    )
                    
                    st.plotly_chart(fig_bar, use_container_width=True)
            
            with col_chart2:
                # Pie chart of pass/fail
                fig_pie = go.Figure()
                fig_pie.add_trace(go.Pie(
                    values=[passing_members, failing_members],
                    labels=['Passing', 'Failing'],
                    marker_colors=['#4CAF50', '#F44336'],
                    hole=0.4
                ))
                
                fig_pie.update_layout(
                    title="Pass/Fail Distribution",
                    height=400
                )
                
                st.plotly_chart(fig_pie, use_container_width=True)
            
            # Export options
            st.markdown("---")
            st.markdown("#### 💾 Export Results")
            
            col_exp1, col_exp2, col_exp3 = st.columns(3)
            
            with col_exp1:
                # Summary CSV
                csv_summary = df_summary.to_csv(index=False)
                st.download_button(
                    label="📥 Summary (CSV)",
                    data=csv_summary,
                    file_name=f"Design_Summary_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    key="export_summary_csv"
                )
            
            with col_exp2:
                # Detailed CSV (all load combinations)
                all_detailed = []
                for member, data in all_results.items():
                    df_result = data['results'].copy()
                    df_result['Member'] = member
                    df_result['Section'] = data['config']['section']
                    df_result['Material'] = data['config']['material']
                    df_result['Type'] = data['config']['member_type']
                    all_detailed.append(df_result)
                
                df_all_detailed = pd.concat(all_detailed, ignore_index=True)
                csv_detailed = df_all_detailed.to_csv(index=False)
                
                st.download_button(
                    label="📥 Detailed (CSV)",
                    data=csv_detailed,
                    file_name=f"Design_Detailed_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    key="export_detailed_csv"
                )
            
            with col_exp3:
                # Excel export
                if EXCEL_AVAILABLE:
                    buffer = BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        # Summary sheet
                        df_summary.to_excel(writer, sheet_name='Summary', index=False)
                        
                        # Individual member sheets
                        for member, data in all_results.items():
                            sheet_name = str(member)[:31]  # Excel limit
                            data['results'].to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Format sheets
                        wb = writer.book
                        for ws_name in wb.sheetnames:
                            ws = wb[ws_name]
                            # Header formatting
                            for cell in ws[1]:
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            # Auto-width
                            for column in ws.columns:
                                max_length = 0
                                column_letter = get_column_letter(column[0].column)
                                for cell in column:
                                    try:
                                        if cell.value:
                                            max_length = max(max_length, len(str(cell.value)))
                                    except:
                                        pass
                                ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
                    
                    buffer.seek(0)
                    st.download_button(
                        label="📥 Full Report (Excel)",
                        data=buffer,
                        file_name=f"Design_Report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="export_excel"
                    )
                else:
                    st.warning("⚠️ Excel export requires openpyxl")

# ==================== FOOTER ====================
st.markdown("---")
st.markdown("""
<div style='text-align: center; padding: 2rem; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
            color: white; border-radius: 15px; margin-top: 2rem;'>
    <h3 style='margin: 0; font-weight: 700;'>AISC 360-16 Steel Design v7.0</h3>
    <p style='margin: 0.5rem 0; font-size: 1.1rem;'>🎯 UI/UX | 📊 Advanced Export | 📈 Enhanced Visualizations</p>
    <p style='margin: 0.5rem 0;'>📐 Full AISC Compliance: F2 Flexural | E3 Compression | H1 Combined Forces</p>
    <p style='margin: 0.5rem 0; font-size: 0.9rem; opacity: 0.9;'>
        <i>© 2024 - Structural Engineering Tool</i>
    </p>
</div>
""", unsafe_allow_html=True)
